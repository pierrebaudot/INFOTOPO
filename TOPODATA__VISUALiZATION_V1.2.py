# -*- coding: utf-8 -*-
"""
Created on Thu Nov  5 16:06:46 2015
Copyright (2007-2017) Pierre Baudot, Daniel Bennequin, Jean-Marc Goaillard, Monica Tapia 
The terms of the licence GNU GPL is reproduced in the file "COPYING" of INFOTOPO distribution.
 
This file is part of INFOTOPO.
 
  INFOTOPO is free software: you can redistribute it and/or modify
  it under the terms of the GNU General Public License as published by
  the Free Software Foundation, either version 3 of the License, or
  (at your option) any later version.
 
  INFOTOPO is distributed in the hope that it will be useful,
  but WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU General Public License for more details.
 
  You should have received a copy of the GNU General Public License
  along with INFOTOPO.  If not, see <http://www.gnu.org/licenses/>. 
  
@author: Pierre Baudot, Daniel Bennequin, Jean-Marc Goaillard, Monica Tapia 

GENERAL PHILOSOPHY AND AUTORS REQUEST: 
Giving the authorization of free use of this program also gives you the responsibility 
of its use: we trust you. Every tool can be used to make things better for all, but also worse.
We ask any users of the present program to use and develop it such that it profits to the 
sustainable development of all, societies and environment and avoid development
in selfish interest, conflict involvement, or individual freedom restrictions or survey. 
As a Human in the world, please be conscious of what you do, of its consequence on others, and try to make it better.
This tools comes from mathematic and pertain to it, mathematic do not pertain to anybody in particular 
and according to its original developments promotes harmony and sustainable developments on short 
and long time and space scales, please respect this ancestral tradition, old as humanity, and develop it in this respect. 
More information on the practical use and mathematical framework and bibliographic citations:
http://www.biorxiv.org/content/early/2017/07/26/168740 
http://www.mdpi.com/1099-4300/17/5/3253 
and README file. 
Baudot wrote the code and participated to theorems and algorithms, Bennequin participated to the theory and main theorems that the algorithm implement, 
Goaillard and Tapia participated to the development of the algorithms and program. 
For any requests, questions, improvements, developments (etc.) contact pierre.baudot [at] gmail.com
Friendly ergonomic interface will be developped afterward, sorry, please see the README file and/or contact me for use. 
INFOTOPO has been developed since 2007 thanks to grants-fundings-hosting of:
_ Institut de Mathématiques de Jussieu-Paris Rive Gauche (IMJ-PRG) (2006-2010)
_ ISC-PIF (Complex system institute Paris Ile de France) (2007-2013) 
_ Max Planck Institute for Mathematic in the Sciences (MPI-MIS, Leipzig)   (2013-2015)
_ Inserm Unis 1072 - ERC channelomics (2015-2017)
We thank those public and non-profit scientific organization for their support.
""" 

# ************************************************************
# Reading of datafiles (specific of .CSV files or XLS file)
#csvkit manual: http://csvkit.readthedocs.org/en/latest/cli.html
# ************************************************************

import math
from math import *
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.colors import LogNorm 
from matplotlib.colors import SymLogNorm
import numpy as np
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter.messagebox import *
from tkinter import filedialog
import pickle
import itertools
from operator import itemgetter
from collections import OrderedDict
import networkx as nx

#################################################################
#################################################################
#################################################################
### THE INPUT  PARAMETERS OF THE PROGRAM  #######################
#################################################################
#################################################################
#################################################################

#in order to use this program read the comments in grey

#Give the path of the file where the .xlsx file and .plk files are
directory_path =  os.path.abspath("/home/baudot/LABOMARSEILLE/RESULTAT ANALYSE PCR/ACTUAL RESEARCH/ANALYSE 21RNA MARS/CLUSTER_10_10SHUFFLE/")
#directory_path =  os.path.abspath("/home/baudot/LABOMARSEILLE/RESULTAT ANALYSE PCR/ACTUAL RESEARCH/ANALYSE 21RNA MARS/DOPA_n=8_oct2017SHUFFLE/")
#give the name of the .xlsx worksheet on which the data are :
Name_worksheet="DOPA"
# the number of different values that can take each random variable, the resampling scale.
Nb_bins=8
# Nb var is the number n used for the computation of n-tuple information, 
# it is the number n of random variable: the total dimension of the data space.
Nb_var=20
# Nb var tot is the total number m of variable used for the computation of n-tuple information
# we have necessarily m>n
Nb_var_tot=20

#Shuffle are making a derangement permutation of all the rows values of the data matrix, rows by rows, and then compute all informations functions on the randomiszed sample
# if you computed shuffles of the data and whish to realize significance test of independence choose True
compute_shuffle = False
# the Shuffle can be made nb_of_shuffle times, with each time computing informations
nb_of_shuffle=17
# p value is the significiativity of the test of independence computed using the shuffles
# p_value=0.05 means 5 percent of the shuffled independent distribution of the Ik values 
p_value=0.05

#number of bins in the histograms of info values (between max and min)
resultion_histo=50 

##################################################################
# Choose the file .plk to load.  By default choose True for all load_results_and false for compute_different_  
##################################################################
# add a number in the name of the .plk file to load 
# compute_different_m will compute all entropie and mutualinfo for different values of Nb_trials=m (SAMPLE SIZE m)
compute_different_m = False
Nb_of_m = 10 # the number of Nb_trials_max used ranges from Nb_trials_max to int(Nb_trials_max/(Nb_of_m)) with stepps of int(Nb_trials_max/(Nb_of_m))
#Nb_trials =Nb_trials_max- k*int(Nb_trials_max/(Nb_of_m)) 
#ex: if Nb_of_m=10  and Nb_trials_max=111, the program will run for Nb_trials=111,100,89,78,67,56,45,34,23,12
# compute_different_N will compute all entropie and mutualinfo for a value of Nb_bins=N (GRAINING SIZE N )
# with Nb_bins=number_load+2
compute_different_N = False
Nb_of_N = 17 
choose_number_load = False
number_load=0   



load_results_MATRIX = False
load_results_ENTROPY = True
load_results_ENTROPY_ORDERED = True
load_results_INFOMUT = True
load_results_ENTROPY_SUM = True
load_results_INFOMUT_ORDERED = True
load_results_INFOMUT_SUM = True

###########################
# Choose the type of data analysis and representation (figure)    
###########################

# Computes and display various entropies, means,  efficiencies etc...    
SHOW_results_ENTROPY = False
# Computes and display various infomut, means, etc...    
SHOW_results_INFOMUT = False
# Computes and display cond infomut, cond infomut landscape, etc... 
SHOW_results_COND_INFOMUT = False
# Computes and display HISTOGRAMS ENTROPY, ENTROPY landscape, etc...
SHOW_results_ENTROPY_HISTO = False
# Computes and display INFOMUT HISTOGRAMS , INFOMUT landscape, etc...
SHOW_results_INFOMUT_HISTO = False
# Computes and display INFOMUT PATHS, etc...
SHOW_results_INFOMUT_path = False
# Computes and display SCAFFOLDS (RING representation) of mutual info
# currently only for I2 ( pairwise infomut...), Ik to be developped soon
SHOW_results_SCAFOLD =  False
# Computes and display INFOMUT-ENTROPY-k landscape (ENERGY VS. ENTROPY) ...
SHOW_results_INFO_entropy_landscape = True
# Computes the mean Ik as a function os the binning graining N and dim k ..
SHOW_results_INFOMUT_Normed_per_bins = False
# Computes the mean Ik as a function os the sample size m and dim k ..
SHOW_results_INFOMUT_Normed_per_samplesize = False
# Reload the information landscapes saved in pkl file
SHOW_results_RELOAD_LANDSCAPES =  False
# print data saved in pkl file
SHOW_results_PRINT_PLK_file  =  False
degree_to_print= 4
variable_x=4 -1
variable_y=12 -1

SHOW_HARD_DISPLAY= False
SHOW_HARD_DISPLAY_medium= False

#display_figure = True
display_figure = True
#matplotlib.use('agg')
#matplotlib.use('svg')
Format_SVG_SAVE_Figure = False
save_results = False
save_landscape_object = True
#save_results = False

#################################################################
### Procedure for saving and loading objects ####################
### Procedure for saving FIGURES             ####################
#################################################################

def save_obj(obj, name ):
#    with open('obj/'+ name + '.pkl', 'wb') as f:
    with open(directory_path + '/' + name + '.pkl', 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)

def load_obj(name ):
    print('File loaded: ', directory_path + '/' + name + '.pkl')  
    with open(directory_path + '/' + name + '.pkl', 'rb') as f:       
        return pickle.load(f)
      

def save_FIGURE(num_fig2, name, Format_SVG_SAVE_Fig ):
#    with open('obj/'+ name + '.pkl', 'wb') as f:
#   mng = plt.get_current_fig_manager(num_fig2)                                         
#   mng.window.showMaximized(num_fig2)
   plt.tight_layout(num_fig2)   
   if Format_SVG_SAVE_Fig == False: 
         plt.savefig(os.path.join(directory_path,name), format="png")
   if Format_SVG_SAVE_Fig == True:       
         plt.savefig(os.path.join(directory_path,name), format="svg")   


#################################################################
### dialog open file XLS DATA FILE ##############################
#################################################################

root = Tk()

root.filename =  filedialog.askopenfilename(initialdir = directory_path,title = "choose your file",filetypes = (("xls files","*.xlsx"),("all files","*.*")))
root.destroy()

workbook_data = load_workbook(root.filename)

########################################################################
### Tools to manipulate WORKBOOK  ######################################
# https://openpyxl.readthedocs.org/en/2.3.3/index.html #################
########################################################################


#Name_worksheet="NONDOPA_MONICA_SIMONE"

print(workbook_data.get_sheet_names())
d = workbook_data[Name_worksheet].cell(row = 2, column = 5)
print(d.value)
c = workbook_data[Name_worksheet]['B1':'B239']
print(c)


###############################################################
### procedure for resampling and displaying  ################## 
###     probability distributions            ##################
###############################################################


Nb_column=1
for i in range(2,10000):
    if workbook_data[Name_worksheet].cell(row = 1, column = i).value == None:
       Nb_column=i-2
       break
   


Nb_trials=Nb_column
Nb_trials_max=Nb_trials

# compute_different_m will compute all entropie and mutualinfo for different values of Nb_trials=m (SAMPLE SIZE m)
if compute_different_m == True :
    Nb_trials =Nb_trials_max- number_load*int(Nb_trials_max/(Nb_of_m)) 
print('the number of column (or trials n) is: ', Nb_trials)       
#ex: if Nb_of_m=10  and Nb_trials_max=111, the program will run for Nb_trials=111,100,89,78,67,56,45,34,23,12

if compute_different_N == True :
    Nb_bins=(number_load+2)
print('the number of bins in the GRAINING is: ', Nb_bins)    

#We put the XLSSheet into a numpy matrix nparray #############
# Variables (RNA) are the lines #
# Trials are the columns  #
# Matrix_data contains the RAW DATA
# Matrix_data contains the  DATA resampled and rescaled

Matrix_data = np.zeros((Nb_var_tot,Nb_trials))
Matrix_data2 = np.zeros((Nb_var_tot,Nb_trials))

# WE PUT ALL THE 0 value to nan  #
for col in range(Nb_trials):
  for row in range(Nb_var_tot):
      if workbook_data[Name_worksheet].cell(row = (row+2), column = (col+2)).value == 0:
         Matrix_data[row,col] = np.nan
      else:
         Matrix_data[row,col] = workbook_data[Name_worksheet].cell(row = (row+2), column = (col+2)).value



Min_matrix = np.nanmin(Matrix_data, axis=1)
Max_matrix = np.nanmax(Matrix_data, axis=1) 
Max_matrix = Max_matrix + ((Max_matrix - Min_matrix)/100000)
Ampl_matrix = Max_matrix - Min_matrix

# WE RESCALE ALL MATRICES AND SAMPLE IT into  Nb_bins (Parameter N Graining) #
for col in range(Nb_trials):
  for row in range(Nb_var_tot):
      if np.isnan(Matrix_data[row,col]):
         Matrix_data[row,col] = 0
         Matrix_data2[row,col] = 0
         workbook_data["DOPA_SAMPLED"].cell(row = (row+2), column = (col+2)).value = 0 
      else:
         if Ampl_matrix[row] !=0 :
             Matrix_data2[row,col] = int(((Matrix_data[row,col]-Min_matrix[row])*(Nb_bins))/(Ampl_matrix[row]))+1
         else:    
             Matrix_data2[row,col] = 0
         workbook_data["DOPA_SAMPLED"].cell(row = (row+2), column = (col+2)).value = Matrix_data2[row,col]          
        
         


###############################################################
########          SOME FUNCTIONS USEFULLS            ##########
###          AT ALL ORDERS On SET OF SUBSETS          #########
###############################################################     


# Fonction factorielle 

def factorial(x):
     if x < 2:
         return 1
     else:
         return x * factorial(x-1)
         
# Fonction coeficient binomial (nombre de combinaison de k elements dans [1,..,n]) 
   
def binomial(n,k):
    return factorial(n)/(factorial(k)*factorial(n-k))
    
    
    
def compute_Ninfomut_cond(Nentropie_input):    
#    Ninfomut={}   
    for x,y in Nentropie_input.items():
        for k in range(1, len(x)+1):
           for subset in itertools.combinations(x, k):
#              print(subset)  
              Ninfomut[x]=Ninfomut.get(x,0)+ ((-1)**(len(subset)+1))*Nentropie_input[subset] 
    return (Ninfomut)     
    
    
          

###############################################################
### AFFICHAGE ET ENREGISTREMENT DES DATA RESAMPLEES  ##########
###                  ET RESCALEES            ##################
###############################################################

num_fig=0

workbook_data.save(os.path.join(directory_path,"Sampled_DATA.xlsx"))     
#plt.figure(1)
num_fig=num_fig+1
plt.matshow(Matrix_data, cmap='autumn')
if save_results == True: 
   if Format_SVG_SAVE_Figure == False: 
      plt.savefig(os.path.join(directory_path,'Matrix_RAW_VALUE'), format="png")
   if Format_SVG_SAVE_Figure == True:       
      plt.savefig(os.path.join(directory_path,'Matrix_RAW_VALUE'), format="svg")  

#plt.figure(2)
plt.matshow(Matrix_data2, cmap='autumn')
num_fig=num_fig+1
if save_results == True:
   if Format_SVG_SAVE_Figure == False: 
      plt.savefig(os.path.join(directory_path,'Matrix_RESCALED_VALUE'), format="png")
   if Format_SVG_SAVE_Figure == True:       
      plt.savefig(os.path.join(directory_path,'Matrix_RESCALED_VALUE'), format="svg")  

num_fig=num_fig+1
plt.figure(num_fig)
hist, bin_edges = np.histogram(Matrix_data2, bins = Nb_bins,range=(0,Nb_bins+1))
plt.bar(bin_edges[:-1], hist, width = 1)
plt.xlim(min(bin_edges), max(bin_edges))
if display_figure == True:
   plt.show()
if save_results == True: 
   if Format_SVG_SAVE_Figure == False: 
      plt.savefig(os.path.join(directory_path,'histogram_RESCALED_VALUE'), format="png")
   if Format_SVG_SAVE_Figure == True:       
      plt.savefig(os.path.join(directory_path,'histogram_RESCALED_VALUE'), format="svg")  

num_fig=num_fig+1   
plt.figure(num_fig)
hist, bin_edges = np.histogram(Matrix_data, bins = 52,range=(0,26))
plt.bar(bin_edges[:-1], hist, width = 0.5)
plt.xlim(min(bin_edges), max(bin_edges))
if display_figure == True:
   plt.show()
if save_results == True: 
   if Format_SVG_SAVE_Figure == False: 
      plt.savefig(os.path.join(directory_path,'histogram_RAW_VALUE'), format="png")
   if Format_SVG_SAVE_Figure == True:       
      plt.savefig(os.path.join(directory_path,'histogram_RAW_VALUE'), format="svg")  
   


###############################################################
###                LOAD SAVED FILE           ##################
###                                          ##################
###############################################################


if load_results_MATRIX == True: 
    name_object= 'Matrix_data2'
    if choose_number_load :
        name_object= name_object+str(number_load)
    Matrix_data2 = load_obj(name_object)

if load_results_ENTROPY == True: 
    name_object= 'ENTROPY'
    if choose_number_load :
        name_object= name_object+str(number_load)
    Nentropie = load_obj(name_object)   

if load_results_ENTROPY_ORDERED == True: 
    name_object= 'ENTROPY_ORDERED'
    if choose_number_load :
        name_object= name_object+str(number_load)
    Nentropy_per_order_ordered = load_obj(name_object) 

if load_results_INFOMUT == True: 
   name_object= 'INFOMUT'
   if choose_number_load :
        name_object= name_object+str(number_load)
   Ninfomut = load_obj(name_object) 

if load_results_ENTROPY_SUM == True:  
    name_object= 'ENTROPY_SUM'
    if choose_number_load :
        name_object= name_object+str(number_load)
    entropy_sum_order = load_obj(name_object)   

if load_results_INFOMUT_ORDERED == True: 
    name_object= 'INFOMUT_ORDERED'
    if choose_number_load :
        name_object= name_object+str(number_load)
    Ninfomut_per_order_ordered = load_obj(name_object)
    name_object= 'INFOMUT_ORDEREDList'
    if choose_number_load :
        name_object= name_object+str(number_load)
    infomut_per_order = load_obj(name_object)

if load_results_INFOMUT_SUM == True: 
    name_object= 'INFOMUT_SUM'
    if choose_number_load :
        name_object= name_object+str(number_load)
    infomut_sum_order = load_obj(name_object)
    
if load_results_INFOMUT_SUM == True: 
    name_object= 'INFOMUT_SUM'
    if choose_number_load :
        name_object= name_object+str(number_load)
    infomut_sum_order = load_obj(name_object)    
    
 
###############################################################
###                  LOAD SHUFFLES           ##################
###           AND SHUFFLE ANALYSIS           ##################
############################################################### 
'''
if compute_shuffle == True:
   for k in range(nb_of_shuffle): 
       if load_results_MATRIX == True: 
           name_object= 'Matrix_data2'+str(k)
           Matrix_data2 = load_obj(name_object)

       if load_results_ENTROPY == True: 
           name_object= 'ENTROPY'
           Nentropie = load_obj(name_object)      

       if load_results_ENTROPY_ORDERED == True: 
           name_object= 'ENTROPY_ORDERED'+str(k)
           Nentropy_per_order_ordered = load_obj(name_object) 
  
       if load_results_INFOMUT == True: 
          name_object= 'INFOMUT'+str(k)
          Ninfomut = load_obj(name_object) 

       if load_results_ENTROPY_SUM == True:  
           name_object= 'ENTROPY_SUM'+str(k)
           entropy_sum_order = load_obj(name_object)
    
       if load_results_INFOMUT_ORDERED == True: 
           name_object= 'INFOMUT_ORDERED'+str(k)
           Ninfomut_per_order_ordered = load_obj(name_object)
           name_object= 'INFOMUT_ORDEREDList'+str(k)
           infomut_per_order = load_obj(name_object)

       if load_results_INFOMUT_SUM == True: 
           name_object= 'INFOMUT_SUM'+str(k)
           infomut_sum_order = load_obj(name_object)'''
   
###############################################################
###         PRINT   infomut_sum_order        ##################
###                 EXAMPLE                  ##################
###############################################################    

#(infomut_sum_order_abs,infotot_absbis) = compute_infomut_sum_order_abs(Ninfomut)
#for x in range(0,(2**Nb_var_tot)-1):     
#    print('les infos mutuelles à ', x ,'sont:')
#    print(Ninfomut_per_order_ordered[x])    
#print('la somme des val abs de toutes le infomut est:')
#print(infotot_absbis)         
#print('les valeur abs(info mutuelles) sommées par ordre sont:')
#print(infomut_sum_order_abs)

#print('les info mutuelles sommées par ordre sont:')
#print(infomut_sum_order) 



#########################################################################
#########################################################################
#########################################################################
#########################################################################
######      GRAPHICS AND DISPLAY           ##############################
#########################################################################
#########################################################################
#########################################################################
#########################################################################
# Visualisation Graphique Distribution des valeurs


#########################################################################
#########################################################################
######      JOINT ENTROPIES VARIOUS calculus    #########################
######                FIGURE 1                  #########################
#########################################################################
#########################################################################


#########################################################################
######      JOINT ENTROPIES Values              #########################
#########################################################################

if SHOW_results_ENTROPY == True: 
   
   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   if SHOW_HARD_DISPLAY_medium== True:
      for x,y in Nentropie.items():
         plt.plot(len(x), y, 'ro')
         if y>maxordonnee:
             maxordonnee=y
         if y<minordonnee:
             minordonnee=y
#      xp = np.linspace(minordonnee-((maxordonnee-minordonnee)*0.03), maxordonnee+((maxordonnee-minordonnee)*0.03), 100)
#      plt.plot(len(x), y, 'ro',xp, polyn_fit(xp), '-')
       
      plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
      plt.ylabel('(Bits)')
      plt.title('Entropy')
      plt.grid(True) 
#      if display_figure == True: 
#         plt.show(num_fig)
#      else: plt.close(num_fig)        
      if save_results == True: 
          save_FIGURE(num_fig, 'DIAGRAM_ENTROPY', Format_SVG_SAVE_Figure )
   


#########################################################################
###### Sum of Joint-ENtropy Values per degrees n   ######################
#########################################################################   

   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   for x,y in entropy_sum_order.items():
      plt.plot(x, y, 'ro',x, y, 'b-')
      if y>maxordonnee:
          maxordonnee=y
      if y<minordonnee:
          minordonnee=y
   plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(bits)')
   plt.title('Sum entropy per order')
   plt.grid(True)
#   if display_figure == True: 
#      plt.show(num_fig)
#   else: plt.close(num_fig)        
   if save_results == True: 
       save_FIGURE(num_fig, 'DIAGRAM_ENTROPY', Format_SVG_SAVE_Figure )


#########################################################################
######       Average Entropy rate                  ######################
######################################################################### 
 

   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   for x,y in entropy_sum_order.items():
      rate=y/x
      plt.plot(x, rate, 'ro',x, y, 'b-')
      if rate>maxordonnee:
          maxordonnee=rate
      if rate<minordonnee:
          minordonnee=rate 
   plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(Bits/symbols)')
   plt.title('Average Entropy Rate')
   plt.grid(True)
#   if display_figure == True: 
#      plt.show(num_fig)
#   else: plt.close(num_fig)        
   if save_results == True: 
       save_FIGURE(num_fig, 'Average Entropy Rate', Format_SVG_SAVE_Figure )
       
       
#########################################################################
###########                MEAN ENTROPY                       ########### 
###### Average entropy normalised by Binomial coefficients    ###########
###########       (MEASURE OF CONTRIBUTION PER ORDER)         ########### 
#########################################################################        

   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   average_entropy=[]
   for x,y in entropy_sum_order.items():
      rate=y/binomial(Nb_var,x)
      plt.plot(x, rate,  linestyle='--', marker='o', color='b')
      average_entropy.append(rate)
#      plt.plot(x, rate, 'ro',x, y, 'b-')
      if rate>maxordonnee:
          maxordonnee=rate
      if rate<minordonnee:
          minordonnee=rate 
   plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(Bits/symbols)')
   plt.title('Average entropy normalised by Binomial coef')
   plt.grid(True)
   print('Average entropy normalised by Binomial coef')
   print(average_entropy)
#   if display_figure == True: 
#      plt.show(num_fig)
#   else: plt.close(num_fig)        
   if save_results == True: 
       save_FIGURE(num_fig, 'Average entropy normalised by Binomial coef', Format_SVG_SAVE_Figure )



#########################################################################
######               Entropy rate                  ######################
######################################################################### 


#plt.subplot(323)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Nentropie.items():
#   rate=y/len(x)
#   plt.plot(len(x), rate, 'ro')
#   if rate>maxordonnee:
#       maxordonnee=rate
#   if rate<minordonnee:
#       minordonnee=rate 
#plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(Bits/symbols)')
#plt.title('Entropy Rate')
#plt.grid(True)

# Sum of Joint-entropy values per degrees n


# MULTI-INFO  
#plt.subplot(324)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Nentropie.items():
#    SumH1=0.00
##    print 'le code de sequence encours est:', x
#    for v in range(0,len(x)):
#        Mierda=(x[v],)
#        print 'les composantes marg :', x[v]
 #       print 'lentropy de la marg :', Nentropie[Mierda]
#        SumH1=SumH1+Nentropie[Mierda]
#    print 'la somme tot des H1 :',SumH1 
#    print 'lentropy tot de sequence en cours :',y
#    multiInfo=SumH1-y    
#    plt.plot(len(x), multiInfo, 'ro')
#    if multiInfo>maxordonnee:
#       maxordonnee=multiInfo
#    if multiInfo<minordonnee:
#       minordonnee=multiInfo
#plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(bits)')
#plt.title('MultiInfo = sumH(1)-Htot')
#plt.grid(True)

#########################################################################
######           Order Entropy Duality: H(n-k)-H(k)          ############
######################################################################### 


   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   for x,y in entropy_sum_order.items():
      dualdiff=entropy_sum_order.get(Nb_var-x+1,0)-y
      plt.plot(x, dualdiff, 'ro')
      if dualdiff>maxordonnee:
          maxordonnee=dualdiff
      if dualdiff<minordonnee:
          minordonnee=dualdiff
   plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(bits)')
   plt.title('Order Entropy Duality: H(n-k)-H(k)')
   plt.grid(True)
#   if display_figure == True: 
#      plt.show(num_fig)
#   else: plt.close(num_fig)        
   if save_results == True: 
       save_FIGURE(num_fig, 'Order Entropy Duality: H(n-k)-H(k)', Format_SVG_SAVE_Figure )

#########################################################################
######           AVERAGE MEAN INFORMATION  EFFICIENCY        ############
######################################################################### 

   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   for x,y in entropy_sum_order.items():
# Hmax is = nlog(nb_bin) ...     
      redundancy=1-(y/(binomial(Nb_var,x)*x*math.log(Nb_bins)/math.log(2) ) )
      plt.plot(x,redundancy, 'ro')
      if redundancy>maxordonnee:
          maxordonnee=redundancy
      if redundancy<minordonnee:
          minordonnee=redundancy
   plt.axis([0, Nb_var+1,minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('[0,1]')
   plt.title('Mean Efficiency - Redundancy')
   plt.grid(True)
#   if display_figure == True: 
#      plt.show(num_fig)
 #  else: plt.close(num_fig)        
   if save_results == True: 
       save_FIGURE(num_fig, 'Mean Efficiency - Redundancy', Format_SVG_SAVE_Figure )

# INFORMATION EFFICIENCY 
#plt.subplot(323)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Nentropie.items():
# Hmax is = nlog(nb_bin) ...     
#   redundancy=1-(y/(len(x)*math.log(Nb_bins)/math.log(2) ) )
#   plt.plot(len(x),redundancy, 'ro')
#   if redundancy>maxordonnee:
#       maxordonnee=redundancy
#   if redundancy<minordonnee:
#       minordonnee=redundancy
#plt.axis([0, Nb_var+1,minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('[0,1]')
#plt.title('Efficiency - Redundancy')
#plt.grid(True)



#########################################################################
#########################################################################
######      MUTUAL INFORMATIONS VARIOUS calculus    #####################
######                FIGURE 2                  #########################
#########################################################################
#########################################################################

#########################################################################
######               MUTUAL INFORMATIONS                     ############
######################################################################### 


if SHOW_results_INFOMUT == True: 
   num_fig=num_fig+1
   plt.figure(num_fig)
   if SHOW_HARD_DISPLAY_medium== True:
      import matplotlib as mpl
      mpl.rcParams['agg.path.chunksize'] = 3000000  
      maxordonnee=-1000000.00
      minordonnee=1000000.00
      x_absss = np.array([])
      y_absss = np.array([]) 
      for x,y in Ninfomut.items():
         if len(x)>0:
#             print(len(x))
#             print(y)
             x_absss=np.append(x_absss,[len(x)])
             y_absss=np.append(y_absss,[y])
#             plt.plot(len(x), y, 'ro',len(x), y, 'b-',len(x),0,'r-')
             if y>maxordonnee:
                maxordonnee=y
             if y<minordonnee:
                minordonnee=y
#      plt.plot(x_absss, y_absss, 'ro',len(x), y, 'b-',len(x),0,'r-')     
#      plt.plot(x_absss, y_absss, 'ro',alpha=(-1.8*x_absss)/(Nb_var-2)+(Nb_var-0.2)/(Nb_var-2))          
      plt.scatter(x_absss, y_absss, marker= "_", s= 1000)
#      print(x_absss)
#      print(y_absss)
      plt.draw()       
      plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
      plt.ylabel('(Bits)')
      plt.title('Mutual Information')
      plt.grid(True)
#      if display_figure == True: 
#        plt.show(num_fig)        
#      else: plt.close(num_fig)        
      if save_results == True: 
         save_FIGURE(num_fig, 'DIAGRAM_INFOMUT', Format_SVG_SAVE_Figure )

#########################################################################
######        SUM OF MUTUAL INFORMATIONS Per degrees         ############
######################################################################### 

   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   for x,y in infomut_sum_order.items():
       if x>1:
           plt.plot(x, y, 'ro',x, y, 'b-',x, 0, 'r-')
           if y>maxordonnee:
              maxordonnee=y
           if y<minordonnee:
              minordonnee=y
   plt.axis([1, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(Bits)')
   plt.title('Sum Mutual Information per Order')
   plt.grid(True)
#   print('tamerenslip-2')
#   if display_figure == True: 
#        plt.show(num_fig)
#        print('tamerenslip-1')
#   else: plt.close(num_fig)        
   if save_results == True: 
         save_FIGURE(num_fig, 'Sum Mutual Information per Order', Format_SVG_SAVE_Figure )

 
#########################################################################
######      Average MUTUAL INFORMATIONS RATE Per degrees     ############
######################################################################### 

   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   for x,y in infomut_sum_order.items():
       rate=y/x
       if x>0:
           plt.plot(x, rate, 'ro',x, y, 'b-',x, 0, 'r-')
           if rate>maxordonnee:
              maxordonnee=rate
           if rate<minordonnee:
              minordonnee=rate
   plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(Bits)')
   plt.title('Averaged Mutual Information rate')
   plt.grid(True)
   print('tamerenslip0')
#   if display_figure == True: 
#        plt.show(num_fig)
 #       print('tamerenslip1')
 #  else: plt.close(num_fig)        
   if save_results == True: 
         save_FIGURE(num_fig, 'Averaged Mutual Information rate', Format_SVG_SAVE_Figure )

#########################################################################
######     Average MUTUAL INFORMATIONS NORMALISED Per degrees ###########
#########################################################################

 
   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   normed_informut={} 
   for x,y in infomut_sum_order.items():
       rate=y/binomial(Nb_var,x)
       normed_informut[x]=y/binomial(Nb_var,x)
       if x>0:
           plt.plot(x, rate, 'ro',x, y, 'b-',x, 0, 'r-')
           if rate>maxordonnee:
              maxordonnee=rate
           if rate<minordonnee:
              minordonnee=rate
   plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(Bits)')
   plt.title('Averaged Mutual Information Normalised by Binomial coef')
   plt.grid(True)   
   plt.show(num_fig)
#   if display_figure == True: 
#        plt.show(num_fig)
#        print('tamerenslip')
#   else: plt.close(num_fig)        
   if save_results == True: 
         save_FIGURE(num_fig, 'Averaged Mutual Information Normalised by Binomial coef', Format_SVG_SAVE_Figure )
         save_obj(normed_informut,'normed_informut'+str(Nb_bins))


#########################################################################
######     INFORMATION DISTANCE and n-VOLUME: d= H-I          ###########
#########################################################################
# Information distance-pseudovol d= H-I 

   if SHOW_HARD_DISPLAY== True:
      num_fig=num_fig+1
      plt.figure(num_fig)   
      maxordonnee=-1000000.00
      minordonnee=1000000.00
      for x,y in Ninfomut.items():
          dist=Nentropie[x]-Ninfomut[x]
          if len(x)>0:
             plt.plot(len(x), dist, 'ro',len(x), dist, 'b-',len(x),0,'r-')
             if dist>maxordonnee:
                maxordonnee=dist
             if dist<minordonnee:
                minordonnee=dist
      plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
      plt.ylabel('(Bits)')
      plt.title('Info 2-Distance - n-Volume')
      plt.grid(True)
#      if display_figure == True: 
#           plt.show(num_fig)
#      else: plt.close(num_fig)        
      if save_results == True: 
            save_FIGURE(num_fig, 'Info 2-Distance - n-Volume', Format_SVG_SAVE_Figure )
         
#plt.show()
 
#########################################################################
######            VARIOUS KEPT FOR MEMORY                     ###########
#########################################################################
 
#plt.subplot(323)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Ninfomut.items():
#   if len(x)>1:
#       percent=100*math.fabs(y)/infomut_sum_order_abs[len(x)]
#       plt.plot(len(x), percent, 'ro')
#       if percent>maxordonnee:
#          maxordonnee=percent
#       if percent<minordonnee:
#          minordonnee=percent
#plt.axis([1, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(%)')
#plt.title('InfoMut(n)/SumInfomutOrder(n)')
#plt.grid(True)

# Calcul de Ninfomut_contrib:
# % de participation de |In| par rapport a la somme de 
# toutes les |Ik| (k=<n) avec les mêmes variables internes

#Ninfomut_contrib={}
#for x,y in Ninfomut.items():
#   Denominator=0.00 
#   for w,z in Ninfomut.items():
#         test_in=0         
#         for v in range(0,len(w)):
#             if  w[v] in x: 
#                test_in= test_in+0
#             else: 
#                test_in= test_in+1 
#         if test_in!= 0 :
#             Denominator=Denominator
#            Ninfomut[x]=Ninfomut.get(x,0) 
# I suppose the last lign was not very useful     pass?      
#         else:
#             Denominator=Denominator+math.fabs(Ninfomut[w])
 #            Ninfomut[x]=Ninfomut.get(x,0)+ ((-1)**(len(w)+1))*Nentropie[w]
#   if Denominator!=0.00:
#       Ninfomut_contrib[x]=math.fabs(Ninfomut[x])/Denominator 
   

#plt.subplot(324)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Ninfomut_contrib.items():
#   if len(x)>1:
#      plt.plot(len(x), y, 'ro')
#      if y>maxordonnee:
#         maxordonnee=y
#      if y<minordonnee:
#         minordonnee=y
#plt.axis([1, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(%)')
#plt.title('abs(InfoMut(n))/Sum_abs(InfomutOrder)(1->n)')
#plt.grid(True)


#plt.subplot(325)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Ninfomut.items():
#   if len(x)>1:  
#      percent=100*math.fabs(y)/infotot_absbis
#      plt.plot(len(x), percent, 'ro')
#      if percent>maxordonnee:
#         maxordonnee=percent
#      if percent<minordonnee:
#         minordonnee=percent
#plt.axis([1, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(%)')
#plt.title('InfoMut(n)/SumInfomutTOT')
#plt.grid(True)

#plt.subplot(326)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Nentropie.items():
#    if len(x)>1:
#        SumH1=0.00
#        print 'le code de sequence encours est:', x
#        for v in range(0,len(x)):
#            Mierda=(x[v],)
#            print 'les composantes marg :', x[v]
 #           print 'lentropy de la marg :', Nentropie[Mierda]
#            SumH1=SumH1+Nentropie[Mierda]
#        print 'la somme tot des H1 :',SumH1 
#        print 'lentropy tot de sequence en cours :',y
#        multiInfo=((SumH1-y)**0.5)   
#        plt.plot(len(x), multiInfo, 'ro')
#        if multiInfo>maxordonnee:
#           maxordonnee=multiInfo
#        if multiInfo<minordonnee:
#           minordonnee=multiInfo
#plt.axis([1, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(sqrt(HoRh))')
#plt.title('MultiInfo = sqrt(sumHi-Htot)')
#plt.grid(True)

#plt.figure(7)

#plt.subplot(321)
#maxordonnee=-1000000.00/home/baudot/LABOMARSEILLE/RESULTAT ANALYSE PCR/ACTUAL RESEARCH/ANALYSE 21RNA MARS/TEST PROGRAMME
#minordonnee=1000000.00
#for x,y in Nentropie.items():
#    if len(x)>1:
#        SumH1=0.00
#        print 'le code de sequence encours est:', x
#        for v in range(0,len(x)):
#            Mierda=(x[v],)
#            print 'les composantes marg :', x[v]
 #           print 'lentropy de la marg :', Nentropie[Mierda]
#            SumH1=SumH1+Nentropie[Mierda]
#        print 'la somme tot des H1 :',SumH1 
#        print 'lentropy tot de sequence en cours :',y
#        multiInfo=(math.log(SumH1-y))   
#        plt.plot(len(x), multiInfo, 'ro')
#        if multiInfo>maxordonnee:
#           maxordonnee=multiInfo
#        if multiInfo<minordonnee:
#           minordonnee=multiInfo
#plt.axis([1, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(log(HoRh)/n)')
#plt.title('MultiInfo = log(sumHi-Htot)/n')
#plt.grid(True)


#########################################################################
#########################################################################
######        CONDITIONAL MUTUAL INFORMATION    #########################
######                FIGURE 3                  #########################
#########################################################################
#########################################################################
# FIGURE 3.5 
# Nentropy and Ninfomut are dictionaries (x,y) with x a list of kind (1,2,5) and y a value in bit 


if SHOW_results_COND_INFOMUT == True: 
   num_fig=num_fig+1
   plt.figure(num_fig)
   moyenne={}
   nbpoint={}
   matrix_distrib_infomut=np.array([])
   x_absss = np.array([])
   y_absss = np.array([]) 
# Display every Histo with its own scales:   
#   minim={}
#  maxim={}
   ListInfomutcond={}
   maxima_tot=-1000000.00
   minima_tot=1000000.00
   Ninfomut_COND=[]
   Ninfomut_CONDtot=[]
   Ninfomut_COND_perorder=[]
   for i in range(1,Nb_var+1):
      ListInfomutcond[i]=[]
      Ninfomut_CONDperORDER=[]
      Ninfomut_COND.append(Ninfomut_CONDperORDER)
      dicobis={} 
      Ninfomut_CONDtot.append(dicobis)
      for j in range(1,Nb_var+1):
          dico={} 
          Ninfomut_COND[i-1].append(dico)

   for i in range(1,Nb_var+1):
      for x,y in Ninfomut.items():
         if len(x)>1: 
            for b in x:
                 if (b==i):
                    xbis= tuple(a for a in x if (a!=i)) 
#                    print('x:',x,'xbis:', xbis)
                    cond= Ninfomut[xbis]-y
                    if cond>maxima_tot:
                       maxima_tot=cond
                    if cond<minima_tot:
                       minima_tot=cond 
# for conditioning per degree                       
                    ListInfomutcond[len(x)-1].append(cond)                  
# for conditioning per variable                    
#                    ListInfomutcond[i].append(cond) 
                    Ninfomut_COND[len(x)-1][i-1][xbis]=cond
                    xter = xbis + ((i),)
                    Ninfomut_CONDtot[len(x)-1][xter]=cond
# The last term in the tuple is the conditionning variable                    
#      moyenne[len(x)]=moyenne.get(len(x),0)+y
#      nbpoint[len(x)]=nbpoint.get(len(x),0)+1
         if SHOW_HARD_DISPLAY_medium== True:
             if len(x)>0:
#             print(len(x))
#             print(y)
                    x_absss=np.append(x_absss,[len(x)-0.5])
                    y_absss=np.append(y_absss,[y])
   

 

   for a in range(1,Nb_var+1):
       if Nb_var<9 :
           plt.subplot(3,3,a)
       else :    
           if Nb_var<=16 :
               plt.subplot(4,4,a)
           else : 
              if Nb_var<=20 : 
                   plt.subplot(5,4,a) 
              else :
                  plt.subplot(5,5,a)
       ListInfomutcond[a].append(minima_tot-0.1)
       ListInfomutcond[a].append(maxima_tot+0.1)           
#       Numgraph=
       resultion_histo=100  
       n, bins, patches = plt.hist(ListInfomutcond[a], resultion_histo, facecolor='r')
       plt.axis([minima_tot, maxima_tot,0,n.max()])
#      print(n)
       if a==1 :
          matrix_distrib_infomut=n
       else: 
          matrix_distrib_infomut=np.c_[matrix_distrib_infomut,n]
#          print(matrix_distrib_infomut)
#       matrix_distrib_infomut=np.vstack((matrix_distrib_infomut,n))
#      plt.axis([minim[a]-((maxim[a]-minim[a])*0.07), maxim[a]+((maxim[a]-minim[a])*0.07), 0,nbpoint[a]+10000])
#       plt.ylabel('(count)')
#       plt.title('Histogram of Mutual info at order '+str(a))
#       plt.text(moyenne[a]/nbpoint[a], 1,moyenne[a]/nbpoint[a])
       plt.grid(True)
#   if display_figure == True: 
#   print(matrix_distrib_infomut)   
    
#   mng = plt.get_current_fig_manager(num_fig)                                         
#   mng.window.showMaximized(num_fig)
#   plt.show(num_fig)   
#   plt.tight_layout(num_fig)    
   if save_results == True:    
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="svg")  
   
   
   num_fig=num_fig+1
   plt.figure(num_fig)  
   matrix_distrib_infomut=np.flipud(matrix_distrib_infomut)
   multiply_by_info_values= False
   if multiply_by_info_values :
        for i in range(0,resultion_histo):    
           matrix_distrib_infomut[i,:]=matrix_distrib_infomut[i,:] * ((maxima_tot-minima_tot+0.2) *(resultion_histo-i)/(resultion_histo-1) +minima_tot-0.1)
           print((maxima_tot-minima_tot+0.2) *(resultion_histo-i) /(resultion_histo-1) +minima_tot-0.1)
           print('i is ')
           print((resultion_histo-i))
           print(matrix_distrib_infomut[i,:])
#   print('min')
#   print(minima_tot-0.1)
#   print(maxima_tot+0.1)
   
#   plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=Nb_var/(2*resultion_histo),extent=[1,Nb_var,minima_tot-0.1,maxima_tot+0.1])
#   norm = matplotlib.colors.Normalize(vmin = 1, vmax = np.max(matrix_distrib_infomut))
   if multiply_by_info_values :
        if ((-1)*matrix_distrib_infomut.min()) <= (matrix_distrib_infomut.max()): 
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=-matrix_distrib_infomut.max(), vmax=matrix_distrib_infomut.max())
        else :
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=matrix_distrib_infomut.min(), vmax=-matrix_distrib_infomut.min())
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=norm)
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   else :
#        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm(vmin=1, vmax=200000))
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm())
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   plt.colorbar()
   plt.grid(False)     
   if SHOW_HARD_DISPLAY_medium== True:        
       plt.scatter(x_absss, y_absss, marker= "o" , facecolor= 'w')
   for i in range(1,6):
       print('ORDER =',(i-1))
       Ninfomut_CONDtot[i-1]=OrderedDict(sorted(Ninfomut_CONDtot[i-1].items(), key=lambda t: t[1]))
       print(Ninfomut_CONDtot[i-1])
#       for j  in range(1,Nb_var+1):
#              print('COND PER =',(j))
#             Ninfomut_COND[i-1][j-1]=OrderedDict(sorted(Ninfomut_COND[i-1][j-1].items(), key=lambda t: t[1]))
#              print(Ninfomut_COND[i-1][j-1])
#   plt.show(num_fig)       
#   print('tamere12')
#else: plt.close(num_fig)    
     

#########################################################################
#########################################################################
######             Histogramms ENTROPY          #########################
######           &  ENTROPY LANDSCAPES          #########################
######                FIGURE 4                  #########################
#########################################################################
#########################################################################

if SHOW_results_ENTROPY_HISTO == True:
   num_fig=num_fig+1
   plt.figure(num_fig)  
   moyenne={}
   nbpoint={}
   matrix_distrib_info=np.array([])
   x_absss = np.array([])
   y_absss = np.array([]) 
# Display every Histo with its own scales:   
#   minim={}
#   maxim={}
   
   maxima_tot=-1000000.00
   minima_tot=1000000.00   
   ListEntropyordre={}
   undersampling_percent=np.array([])

   for i in range(1,Nb_var+1):
     ListEntropyordre[i]=[]

   for x,y in Nentropie.items():
      ListEntropyordre[len(x)].append(y)
      moyenne[len(x)]=moyenne.get(len(x),0)+y
      nbpoint[len(x)]=nbpoint.get(len(x),0)+1
      if SHOW_HARD_DISPLAY_medium== True:
          if len(x)>0:
#             print(len(x))
#             print(y)
                 x_absss=np.append(x_absss,[len(x)-0.5])
                 y_absss=np.append(y_absss,[y])
# Display every Histo with its own scales:          
#      minim[len(x)]=minim.get(len(x),1000000.00)
#      maxim[len(x)]=maxim.get(len(x),-1000000.00)
# Display every Histo with its own scales:      
#      if y>maxim[len(x)]:
#          maxim[len(x)]=y
#     if y<minim[len(x)]:
#          minim[len(x)]=y    
      if y>maxima_tot:
          maxima_tot=y
      if y<minima_tot:
          minima_tot=y     
   for a in range(1,Nb_var+1):
       if Nb_var<=9 :
#       Numgraph=330+a
           plt.subplot(3,3,a)
       else :    
           if Nb_var<=16 :
               plt.subplot(4,4,a)
           else : 
              if Nb_var<=20 : 
                   plt.subplot(5,4,a) 
              else :
                  plt.subplot(5,5,a)
 # compute the Ku undersampling bound                                   
       nb_undersampling_point=0
       for x in range(0,len(ListEntropyordre[a])): 
#           if ListEntropyordre[a][x] >= ((math.log(Nb_trials)/math.log(2))-0.1):
           if ListEntropyordre[a][x] >= ((math.log(Nb_trials)/math.log(2))-0.00000001):
               nb_undersampling_point=nb_undersampling_point+1
               if a ==1:
                   print(ListEntropyordre[a][x])
       undersampling_percent = np.hstack((undersampling_percent,100*nb_undersampling_point/binomial(Nb_var,a)))       
       print('undersampling_percent in dim ',a,' = ', 100*nb_undersampling_point/binomial(Nb_var,a))             
       ListEntropyordre[a].append(minima_tot-0.1)
       ListEntropyordre[a].append(maxima_tot+0.1)           
#    print 'listentropie à lordre:',a
#    print ListEntropyordre[a]
#       resultion_histo=200  
       n, bins, patches = plt.hist(ListEntropyordre[a], resultion_histo, facecolor='g')
       plt.axis([minima_tot, maxima_tot,0,n.max()])
 
            
       if a==1 :
          matrix_distrib_info=n
       else: 
          matrix_distrib_info=np.c_[matrix_distrib_info,n]
#    plt.axis([minim[a]-((maxim[a]-minim[a])*0.07), maxim[a]+((maxim[a]-minim[a])*0.07), 0,nbpoint[a]+10000])
#       plt.ylabel('(count)')
#       plt.title('Histogram of Entropy at order '+str(a))
#       plt.text(moyenne[a]/nbpoint[a], 1,moyenne[a]/nbpoint[a])
       plt.grid(True)
    
#   mng = plt.get_current_fig_manager(num_fig)                                         
#   mng.window.showMaximized(num_fig)
#   plt.tight_layout(num_fig) 
   if save_results == True: 
      plt.savefig(os.path.join(directory_path,'Histogram_ENTROPY'))    
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'Histogram_ENTROPY'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'Histogram_ENTROPY'), format="svg")   
   num_fig=num_fig+1
   plt.figure(num_fig)     
   abssice_degree=np.linspace(1, Nb_var, Nb_var)
   plt.plot(abssice_degree,undersampling_percent)
   plt.ylabel('(percent of undersampled points)')
   plt.title('undersampling bound')
   plt.grid(True)
#   print('WAS HERE')
#   plt.show(num_fig)        
   
         
   num_fig=num_fig+1
   plt.figure(num_fig)  
   matrix_distrib_info=np.flipud(matrix_distrib_info)
   multiply_by_info_values= False
   if multiply_by_info_values :
        for i in range(0,resultion_histo):    
           matrix_distrib_info[i,:]=matrix_distrib_info[i,:] * ((maxima_tot-minima_tot+0.2) *(resultion_histo-i)/(resultion_histo-1) +minima_tot-0.1)
           print((maxima_tot-minima_tot+0.2) *(resultion_histo-i) /(resultion_histo-1) +minima_tot-0.1)
           print('i est ')
           print((resultion_histo-i))
           print(matrix_distrib_info[i,:])
#   print('min')
#   print(minima_tot-0.1)
#   print(maxima_tot+0.1)
   
#   plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=Nb_var/(2*resultion_histo),extent=[1,Nb_var,minima_tot-0.1,maxima_tot+0.1])
#   norm = matplotlib.colors.Normalize(vmin = 1, vmax = np.max(matrix_distrib_infomut))
   if multiply_by_info_values :
        if ((-1)*matrix_distrib_info.min()) <= (matrix_distrib_info.max()): 
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=-matrix_distrib_info.max(), vmax=matrix_distrib_info.max())
        else :
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=matrix_distrib_info.min(), vmax=-matrix_distrib_info.min())
        plt.matshow(matrix_distrib_info, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=norm)
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   else :
#        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm(vmin=1, vmax=200000))
        plt.matshow(matrix_distrib_info, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm())
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   plt.colorbar()
   plt.grid(False)     
   
   if SHOW_HARD_DISPLAY_medium== True:        
       plt.scatter(x_absss, y_absss, marker= "o" , facecolor= 'w')         
   plt.show(num_fig)
    

#else: plt.close(num_fig)         



#########################################################################
#########################################################################
######      Histogramms MUTUAL INFORMATION      #########################
######           &  INFOMUT LANDSCAPES          #########################
######                FIGURE 5                  #########################
#########################################################################
#########################################################################


if SHOW_results_INFOMUT_HISTO == True:


   
   num_fig=num_fig+1
   plt.figure(num_fig)   
   moyenne={}
   nbpoint={}
   matrix_distrib_infomut=np.array([])
   x_absss = np.array([])
   y_absss = np.array([]) 
   
#######################################################  
#   COMPUTE THE LIST OF INFOMUT VALUES FOR EACH DEGREE   
# Display every Histo with its own scales:  
####################################################### 

# Compute the list of Infomut at each degree 
   ListInfomutordre={}
   maxima_tot=-1000000.00
   minima_tot=1000000.00
   for i in range(1,Nb_var+1):
     ListInfomutordre[i]=[]

   for x,y in Ninfomut.items():
      ListInfomutordre[len(x)].append(y)     
      if SHOW_HARD_DISPLAY_medium== True:
          if len(x)>0:
                 x_absss=np.append(x_absss,[len(x)-0.5])
                 y_absss=np.append(y_absss,[y])
      if y>maxima_tot:
          maxima_tot=y
      if y<minima_tot:
          minima_tot=y 
   Ninfomut.clear()
   del(Ninfomut)


# Compute the list of Infomut at each degree for each SHUFFLE and sums the distributions 
# (we sum the districbution because the original n-shuffles vectors would be too big )
   if compute_shuffle == True:
       hist_sum_SHUFFLE = {}
       for a in range(1,Nb_var+1):
           hist_sum_SHUFFLE[a]=[]
         
       for k in range(nb_of_shuffle):     
           print('k=',k)
           name_object= 'INFOMUT'+str(k)
           Ninfomut = load_obj(name_object) 
           ListInfomutordreSHUFFLE={}
           for i in range(1,Nb_var+1):
              ListInfomutordreSHUFFLE[i]=[]
#              hist_sum_SHUFFLE[i]=[]
           for x,y in Ninfomut.items():
               ListInfomutordreSHUFFLE[len(x)].append(y)             
               if y>maxima_tot:
                   maxima_tot=y
               if y<minima_tot:
                   minima_tot=y   
           for i in range(1,Nb_var+1):  
                nSHUFFLE,bin_edgesSHUFFLE  = np.histogram(ListInfomutordreSHUFFLE[i], resultion_histo, (minima_tot,maxima_tot))
#               nSHUFFLE, binSHUFFLE, patcheSHUFFLE = plt.hist(ListInfomutordreSHUFFLE[a], resultion_histo, facecolor='r')
#                print('nSHUFFLE')    
#                print(nSHUFFLE)
#                print('bin_edgesSHUFFLE') 
#                print(len(bin_edgesSHUFFLE))
#                print(bin_edgesSHUFFLE)                   
                if k == 0 :
                   hist_sum_SHUFFLE[i] = nSHUFFLE
#                   print('hist_sum_SHUFFLE[i]')    
#                   print(hist_sum_SHUFFLE[i]) 
                else:    
                   hist_sum_SHUFFLE[i] = np.sum([hist_sum_SHUFFLE[i],nSHUFFLE],axis=0)  
           Ninfomut.clear()
           del(Ninfomut)      
           ListInfomutordreSHUFFLE.clear()
           del(ListInfomutordreSHUFFLE)
#       print('bin_edgesSHUFFLE') 
#       print(len(bin_edgesSHUFFLE))
#       print(bin_edgesSHUFFLE) 
       for i in range(1,Nb_var+1): 
#           hist_sum_SHUFFLE[i] = np.append(hist_sum_SHUFFLE[i], 0) 
           hist_sum_SHUFFLE[i]=np.concatenate([[0],hist_sum_SHUFFLE[i]])
#           print('hist_sum_SHUFFLE[i]') 
#           print(len(hist_sum_SHUFFLE[i]))
#           print(hist_sum_SHUFFLE[i]) 
           
           if Nb_var<9 :  
               plt.subplot(3,3,i)
           else :    
               if Nb_var<=16 :
                   plt.subplot(4,4,i)
               else : 
                  if Nb_var<=20 : 
                       plt.subplot(5,4,i) 
                  else :
                      plt.subplot(5,5,i)
           plt.plot(bin_edgesSHUFFLE,hist_sum_SHUFFLE[i])
#           n_SHUFFLE, bins_SHUFFLE, patches_SHUFFLE = plt.hist(hist_sum_SHUFFLE[i], bin_edgesSHUFFLE)
           plt.axis([minima_tot, maxima_tot,0,hist_sum_SHUFFLE[i].max()])
   
#   COMPUTE THE HISTOGRAMS OF THE LIST OF INFOMUT VALUES FOR EACH DEGREE  
#   If shuffle is true it also compute the signifiance test against independence null hypothesis  
   num_fig=num_fig+1
   fig_Histo_infomut = plt.figure(num_fig)
   if compute_shuffle == True:
       low_signif_bound={}
       high_signif_bound={}
       lign_signif={}
   for a in range(1,Nb_var+1):
       if Nb_var<9 :  
           plt.subplot(3,3,a)
       else :    
           if Nb_var<=16 :
               plt.subplot(4,4,a)
           else : 
              if Nb_var<=20 : 
                   plt.subplot(5,4,a) 
              else :
                  plt.subplot(5,5,a)
#       if a == 2:
#         print(ListInfomutordre[a])           
       ListInfomutordre[a].append(minima_tot-0.1)
       ListInfomutordre[a].append(maxima_tot+0.1)                   
       n, bins, patches = plt.hist(ListInfomutordre[a], resultion_histo, facecolor='r')
       
       if compute_shuffle == True:
           plt.plot(bin_edgesSHUFFLE,hist_sum_SHUFFLE[a]*(1/nb_of_shuffle),color="blue")
           cumul=0
           first_signif = True
           second_signif = False
           lign_signif[a]=[]
           lign_signif[a] = np.zeros_like(hist_sum_SHUFFLE[a])
           for x in range(0,len(hist_sum_SHUFFLE[a])):
               cumul=cumul+hist_sum_SHUFFLE[a][x] 
#               print('cumul = ',  cumul)
               if first_signif == True:
                   if cumul >= (binomial(Nb_var,a)*nb_of_shuffle*p_value):
#                       print('nb low signif = ',  (binomial(Nb_var,a)*nb_of_shuffle*p_value))                       
                       low_signif_bound[a] = bin_edgesSHUFFLE[x]
                       lign_signif[a][x] =hist_sum_SHUFFLE[a].max()
                       first_signif = False
                       second_signif = True
                       print('low_signif_bound in dim',a, ' = ',  low_signif_bound[a])
               if second_signif == True: 
                   if cumul >= (binomial(Nb_var,a)*nb_of_shuffle*(1-p_value)):
                       high_signif_bound[a] = bin_edgesSHUFFLE[x]
                       lign_signif[a][x] =hist_sum_SHUFFLE[a].max()
                       second_signif = False
                       print('high_signif_bound in dim',a, ' = ',  high_signif_bound[a]) 
           nb_of_signif_low = 0     
           nb_of_signif_high = 0
           for x in range(0,len(ListInfomutordre[a])): 
               if ListInfomutordre[a][x] <= low_signif_bound[a] :                  
                  nb_of_signif_low = nb_of_signif_low+1 
               if ListInfomutordre[a][x] >= high_signif_bound[a] :
                  nb_of_signif_high = nb_of_signif_high+1               
           print('nb_of_signif_low in dim',a, ' = ',  nb_of_signif_low)
           print('nb_of_signif_high in dim',a, ' = ',  nb_of_signif_high) 
           plt.plot(bin_edgesSHUFFLE,lign_signif[a]*(1/nb_of_shuffle),color="green")             
       plt.axis([minima_tot, maxima_tot,0,n.max()])
       if a==1 :
          matrix_distrib_infomut=n
       else: 
# np.c_ concatenates  along the second axis.           
          matrix_distrib_infomut=np.c_[matrix_distrib_infomut,n]

       plt.grid(True)
   if save_landscape_object == True: 
       save_obj(fig_Histo_infomut,'Infomut_distrib')
   if save_results == True:    
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="svg")  

   
 #   COMPUTE THE INFOMUT LANDSCAPE FROM THE HISTOGRAMS AS THE MATRIX  matrix_distrib_infomut
#   If shuffle is true it also plots the signifiance test against independence null hypothesis  
    
   num_fig=num_fig+1
   fig_infolandscape =plt.figure(num_fig)  
   matrix_distrib_infomut=np.flipud(matrix_distrib_infomut)
   multiply_by_info_values= False
   if multiply_by_info_values :
        for i in range(0,resultion_histo):    
           matrix_distrib_infomut[i,:]=matrix_distrib_infomut[i,:] * ((maxima_tot-minima_tot+0.2) *(resultion_histo-i)/(resultion_histo-1) +minima_tot-0.1)
           print((maxima_tot-minima_tot+0.2) *(resultion_histo-i) /(resultion_histo-1) +minima_tot-0.1)
           print('i is ')
           print((resultion_histo-i))
           print(matrix_distrib_infomut[i,:])
   
#   plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=Nb_var/(2*resultion_histo),extent=[1,Nb_var,minima_tot-0.1,maxima_tot+0.1])
#   norm = matplotlib.colors.Normalize(vmin = 1, vmax = np.max(matrix_distrib_infomut))
   if multiply_by_info_values :
        if ((-1)*matrix_distrib_infomut.min()) <= (matrix_distrib_infomut.max()): 
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=-matrix_distrib_infomut.max(), vmax=matrix_distrib_infomut.max())
        else :
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=matrix_distrib_infomut.min(), vmax=-matrix_distrib_infomut.min())
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=norm)
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   else :
#        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm(vmin=1, vmax=200000))
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm())
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   if compute_shuffle == True:     
       abssice=np.linspace(0.5, Nb_var-0.5, Nb_var)
       low_ordinate=[]
       high_ordinate=[]
       for a in range(1,Nb_var+1):
           low_ordinate.append(low_signif_bound[a])
           high_ordinate.append(high_signif_bound[a])
#       print('abssice')
#       print(abssice)
#       print('low_ordinate')
#      print(low_ordinate)
#       print('high_ordinate')
#       print(high_ordinate)
       plt.plot(abssice, low_ordinate, marker='o', color='black')
       plt.plot(abssice, high_ordinate, marker='o',color='black')     
   plt.colorbar()
   plt.grid(False)
   if save_landscape_object == True: 
       save_obj(fig_infolandscape,'Infolandscape')

   if SHOW_HARD_DISPLAY_medium== True:        
       plt.scatter(x_absss, y_absss, marker= "o" , facecolor= 'w')  



###############################################################  
###############################################################
########          FIGURE INFORMATION PATH            ##########
########                 FIGURE 6                    ##########
###############################################################  
###############################################################    

# Nentropy and Ninfomut are dictionaries (x,y) with x a list of kind (1,2,5) and y a value in bit 


if SHOW_results_INFOMUT_path == True: 
   num_fig=num_fig+1
   plt.figure(num_fig)
   moyenne={}
   nbpoint={}
   matrix_distrib_infomut=np.array([])
   x_absss = np.array([])
   y_absss = np.array([]) 
# Display every Histo with its own scales:   
#   minim={}
#  maxim={}
   ListInfomutcond={}
   maxima_tot=-1000000.00
   minima_tot=1000000.00
   Ninfomut_COND=[]
   Ninfomut_CONDtot=[]
   Ninfomut_COND_perorder=[]
   
   infocond=1000000.00
   listartbis=[] 
   infomutmax_path_VAR=[] 
   infomutmax_path_VALUE=[]
   infomutmin_path_VAR=[] 
   infomutmin_path_VALUE=[]
   number_of_max_and_min=Nb_var #explore the 2 max an min marginals (of degree 1 information) 
   items = list(Ninfomut_per_order_ordered[1].items())
   for inforank in range(0,number_of_max_and_min): 
       infomutmax_path_VAR.append([])
       infomutmax_path_VALUE.append([])
       infomutmin_path_VAR.append([])
       infomutmin_path_VALUE.append([])
       liststart=[]
#       print(items[inforank])
#       print(items[Nb_var-inforank-1])
       xstart=items[inforank][0]
       xstartmin=items[Nb_var-inforank-1][0]
#       print(xstart)
#       print(xstart[0])
       infomutmax_path_VAR[-1].append(xstart[0])
       infomutmax_path_VALUE[-1].append(items[inforank][1])
       infomutmin_path_VAR[-1].append(xstartmin[0])
       infomutmin_path_VALUE[-1].append(items[Nb_var-inforank-1][1])
       degree=1
       infocond=1000000.00
       while infocond >=0 :
          maxima_tot=-1000000.00
          minima_tot=1000000.00 
          degree=degree+1 
          for i in range(1,Nb_var+1) :
             if i in infomutmax_path_VAR[-1] :     
                del listartbis[:]  
             else:    
                del listartbis[:]
                listartbis=infomutmax_path_VAR[-1][:]
                listartbis.append(i)
                listartbis.sort() 
#                print('listartbis')
#                print(listartbis)
                tuplestart=tuple(listartbis)
                if infomut_per_order[degree][tuplestart]>maxima_tot:
                      maxima_tot=infomut_per_order[degree][tuplestart]
                      tuplemax=tuplestart
                      igood= i
#          print(tuplemax)
#          print(maxima_tot)
          infomutmax_path_VAR[-1].append(igood)
          infomutmax_path_VALUE[-1].append(maxima_tot)
          infocond= infomutmax_path_VALUE[-1][-2]- infomutmax_path_VALUE[-1][-1]
#          print('infocond')
#          print(infocond)
       del infomutmax_path_VAR[-1][-1]
       del infomutmax_path_VALUE[-1][-1]
       print('The path of maximal mutual-info Nb',inforank+1,' is :')   
       print(infomutmax_path_VAR[-1])   
       
       degree=1
       infocond=1000000.00
       while infocond >=0 :
          maxima_tot=-1000000.00
          minima_tot=1000000.00 
          degree=degree+1 
          for i in range(1,Nb_var+1) :
             if i in infomutmin_path_VAR[-1] :     
                del listartbis[:]  
             else:    
                del listartbis[:]
                listartbis=infomutmin_path_VAR[-1][:]
                listartbis.append(i)
                listartbis.sort() 
#                print('listartbis')
#                print(listartbis)
                tuplestart=tuple(listartbis)
                if infomut_per_order[degree][tuplestart]<minima_tot:
                      minima_tot=infomut_per_order[degree][tuplestart]
                      tuplemax=tuplestart
                      igood= i
#          print(tuplemax)
#          print(minima_tot)
          infomutmin_path_VAR[-1].append(igood)
          infomutmin_path_VALUE[-1].append(minima_tot)
          infocond= infomutmin_path_VALUE[-1][-2]- infomutmin_path_VALUE[-1][-1]
#          print('infocond')
#          print(infocond)
       del infomutmin_path_VAR[-1][-1]
       del infomutmin_path_VALUE[-1][-1]   
       print('The path of minimal mutual-info Nb',inforank+1,' is :')   
       print(infomutmin_path_VAR[-1])    

# COMPUTE THE HISTOGRAMS OF INFORMATION           
# Display every Histo with its own scales:   
#  minim={}
#  maxim={}
   ListInfomutordre={}
   maxima_tot=-1000000.00
   minima_tot=1000000.00
   for i in range(1,Nb_var+1):
     ListInfomutordre[i]=[]

   for x,y in Ninfomut.items():
      ListInfomutordre[len(x)].append(y)
      moyenne[len(x)]=moyenne.get(len(x),0)+y
      nbpoint[len(x)]=nbpoint.get(len(x),0)+1
      if SHOW_HARD_DISPLAY_medium== True:
          if len(x)>0:
#             print(len(x))
#             print(y)
                 x_absss=np.append(x_absss,[len(x)-0.5])
                 y_absss=np.append(y_absss,[y])
# Display every Histo with its own scales:
      
      if y>maxima_tot:
          maxima_tot=y
      if y<minima_tot:
          minima_tot=y 

   for a in range(1,Nb_var+1):
       if Nb_var<9 :
           plt.subplot(3,3,a)
       else :    
           if Nb_var<=16 :
               plt.subplot(4,4,a)
           else : 
              if Nb_var<=20 : 
                   plt.subplot(5,4,a) 
              else :
                  plt.subplot(5,5,a)
       ListInfomutordre[a].append(minima_tot-0.1)
       ListInfomutordre[a].append(maxima_tot+0.1)           
#       Numgraph=
       resultion_histo=100  
       n, bins, patches = plt.hist(ListInfomutordre[a], resultion_histo, facecolor='r')
       plt.axis([minima_tot, maxima_tot,0,n.max()])
#      print(n)
       if a==1 :
          matrix_distrib_infomut=n
       else: 
          matrix_distrib_infomut=np.c_[matrix_distrib_infomut,n]
#          print(matrix_distrib_infomut)
#       matrix_distrib_infomut=np.vstack((matrix_distrib_infomut,n))
#      plt.axis([minim[a]-((maxim[a]-minim[a])*0.07), maxim[a]+((maxim[a]-minim[a])*0.07), 0,nbpoint[a]+10000])
#       plt.ylabel('(count)')
#       plt.title('Histogram of Mutual info at order '+str(a))
#       plt.text(moyenne[a]/nbpoint[a], 1,moyenne[a]/nbpoint[a])
       plt.grid(True)
#   if display_figure == True: 
#   print(matrix_distrib_infomut)   
    
#   mng = plt.get_current_fig_manager(num_fig)                                         
#   mng.window.showMaximized(num_fig)
#   plt.show(num_fig)  
   print('onestdes')    
#   plt.tight_layout(num_fig)    
   if save_results == True:    
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="svg")  

# COMPUTE THE MATRIX OF INFORMATION LANDSACPES   
   
   num_fig=num_fig+1
#   plt.figure(num_fig)  
   plt.figure(num_fig,figsize=(50, 30))   
   matrix_distrib_infomut=np.flipud(matrix_distrib_infomut)
   multiply_by_info_values= False
   if multiply_by_info_values :
        for i in range(0,resultion_histo):    
           matrix_distrib_infomut[i,:]=matrix_distrib_infomut[i,:] * ((maxima_tot-minima_tot+0.2) *(resultion_histo-i)/(resultion_histo-1) +minima_tot-0.1)
           print((maxima_tot-minima_tot+0.2) *(resultion_histo-i) /(resultion_histo-1) +minima_tot-0.1)
           print('i est ')
           print((resultion_histo-i))
           print(matrix_distrib_infomut[i,:])

   if multiply_by_info_values :
        if ((-1)*matrix_distrib_infomut.min()) <= (matrix_distrib_infomut.max()): 
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=-matrix_distrib_infomut.max(), vmax=matrix_distrib_infomut.max())
        else :
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=matrix_distrib_infomut.min(), vmax=-matrix_distrib_infomut.min())
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=norm)
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   else :
# for figure paper with colar scale up to 200000       
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm(vmin=1, vmax=200000))
# for autoscaled color scale - general case 
#        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm())
        plt.axis([0,Nb_var,minima_tot,maxima_tot])  
#   plt.axes().set_aspect(0.5)     
   plt.colorbar()
   plt.grid(False)     
   if SHOW_HARD_DISPLAY_medium== True:        
       plt.scatter(x_absss, y_absss, marker= "o" , facecolor= 'w')
 

# COMPUTE THE INFORMATION PATHS    
   x_infomax=[] 
   x_infomin=[]    
   maxima_x=-10
   maxima_tot=-1000000.00
   minima_tot=1000000.00
   for inforank in range(0,number_of_max_and_min): 
       x_infomax.append([])
       j=-0.5
       for y in  range(0,len(infomutmax_path_VALUE[inforank])): 
          j=j+1 
          x_infomax[-1].append(j) 
          if j > maxima_x:
              maxima_x=j
          if infomutmax_path_VALUE[inforank][int(j-0.5)]>maxima_tot :
              maxima_tot=infomutmax_path_VALUE[inforank][int(j-0.5)]
          if infomutmax_path_VALUE[inforank][int(j-0.5)]<minima_tot :
              minima_tot=infomutmax_path_VALUE[inforank][int(j-0.5)]    
              
       x_infomin.append([])
       j=-0.5
       for y in  range(0,len(infomutmin_path_VALUE[inforank])): 
          j=j+1 
          x_infomin[-1].append(j) 
          if j > maxima_x:
              maxima_x=j
          if infomutmin_path_VALUE[inforank][int(j-0.5)]>maxima_tot :
              maxima_tot=infomutmin_path_VALUE[inforank][int(j-0.5)]
          if infomutmin_path_VALUE[inforank][int(j-0.5)]<minima_tot :
              minima_tot=infomutmin_path_VALUE[inforank][int(j-0.5)]     
              
       plt.plot(x_infomax[inforank], infomutmax_path_VALUE[inforank], marker='o', color='red')
       plt.plot(x_infomin[inforank], infomutmin_path_VALUE[inforank], marker='o',color='blue')
       plt.axis([0,maxima_x+0.5,minima_tot-0.2,maxima_tot+0.2])
      
#       plt.xlim(0, maxima_x+0.5)
   display_labelnodes=False
   if display_labelnodes :    
       for inforank in range(0,number_of_max_and_min):
           for label, x,y in zip(infomutmax_path_VAR[inforank], x_infomax[inforank], infomutmax_path_VALUE[inforank]):
                plt.annotate(label,xy=(x, y), xytext=(0, 0),textcoords='offset points')
           for label, x,y in zip(infomutmin_path_VAR[inforank], x_infomin[inforank], infomutmin_path_VALUE[inforank]):
                plt.annotate(label,xy=(x, y), xytext=(0, 0),textcoords='offset points')    
                
               
 
#########################################################################
#########################################################################
######      Histogramms ENTROPY & MUTUAL INFORMATION   ##################
###### computes entropy vs information for each degree ##################
######       ENTROPY  &  INFOMUT LANDSCAPES         #####################
######                FIGURE 6                  #########################
#########################################################################
#########################################################################
### ENTROPY VS ENERGY VS VOL  Willard Gibbs' 1873 figures two and three 
# (above left and middle) used by Scottish physicist James Clerk Maxwell 
# in 1874 to create a three-dimensional entropy (x), volume (y), energy (z) 
# thermodynamic surface diagram 


if SHOW_results_INFO_entropy_landscape == True:


   
   num_fig=num_fig+1
   plt.figure(num_fig)   
   moyenne={}
   nbpoint={}
   matrix_distrib_infomut=np.array([])
   x_absss = np.array([])
   y_absss = np.array([]) 
   
#######################################################  
#   COMPUTE THE LIST OF INFOMUT VALUES FOR EACH DEGREE   
# Display every Histo with its own scales:  
####################################################### 

# Compute the list of Infomut at each degree 
   ListInfomutordre={}
   ListEntropyordre={}
   maxima_tot=-1000000.00
   minima_tot=1000000.00
   for i in range(1,Nb_var+1):
     ListInfomutordre[i]=[]
     ListEntropyordre[i]=[]

   for x,y in Ninfomut.items():
      ListInfomutordre[len(x)].append(y)   
      ListEntropyordre[len(x)].append(Nentropie.get(x))
      if SHOW_HARD_DISPLAY_medium== True:
          if len(x)>0:
                 x_absss=np.append(x_absss,[len(x)-0.5])
                 y_absss=np.append(y_absss,[y])
      if y>maxima_tot:
          maxima_tot=y
      if y<minima_tot:
          minima_tot=y 
   Ninfomut.clear()
   del(Ninfomut)


# Compute the list of Infomut at each degree for each SHUFFLE and sums the distributions 
# (we sum the districbution because the original n-shuffles vectors would be too big )
   if compute_shuffle == True:
       hist_sum_SHUFFLE = {}
       for a in range(1,Nb_var+1):
           hist_sum_SHUFFLE[a]=[]
         
       for k in range(nb_of_shuffle):     
           print('k=',k)
           name_object= 'INFOMUT'+str(k)
           Ninfomut = load_obj(name_object) 
           ListInfomutordreSHUFFLE={}
           for i in range(1,Nb_var+1):
              ListInfomutordreSHUFFLE[i]=[]
#              hist_sum_SHUFFLE[i]=[]
           for x,y in Ninfomut.items():
               ListInfomutordreSHUFFLE[len(x)].append(y)             
               if y>maxima_tot:
                   maxima_tot=y
               if y<minima_tot:
                   minima_tot=y   
           for i in range(1,Nb_var+1):  
                nSHUFFLE,bin_edgesSHUFFLE  = np.histogram(ListInfomutordreSHUFFLE[i], resultion_histo, (minima_tot,maxima_tot))
#               nSHUFFLE, binSHUFFLE, patcheSHUFFLE = plt.hist(ListInfomutordreSHUFFLE[a], resultion_histo, facecolor='r')
#                print('nSHUFFLE')    
#                print(nSHUFFLE)
#                print('bin_edgesSHUFFLE') 
#                print(len(bin_edgesSHUFFLE))
#                print(bin_edgesSHUFFLE)                   
                if k == 0 :
                   hist_sum_SHUFFLE[i] = nSHUFFLE
#                   print('hist_sum_SHUFFLE[i]')    
#                   print(hist_sum_SHUFFLE[i]) 
                else:    
                   hist_sum_SHUFFLE[i] = np.sum([hist_sum_SHUFFLE[i],nSHUFFLE],axis=0)  
           Ninfomut.clear()
           del(Ninfomut)      
           ListInfomutordreSHUFFLE.clear()
           del(ListInfomutordreSHUFFLE)
#       print('bin_edgesSHUFFLE') 
#       print(len(bin_edgesSHUFFLE))
#       print(bin_edgesSHUFFLE) 
       for i in range(1,Nb_var+1): 
#           hist_sum_SHUFFLE[i] = np.append(hist_sum_SHUFFLE[i], 0) 
           hist_sum_SHUFFLE[i]=np.concatenate([[0],hist_sum_SHUFFLE[i]])
#           print('hist_sum_SHUFFLE[i]') 
#           print(len(hist_sum_SHUFFLE[i]))
#           print(hist_sum_SHUFFLE[i]) 
           
           if Nb_var<9 :  
               plt.subplot(3,3,i)
           else :    
               if Nb_var<=16 :
                   plt.subplot(4,4,i)
               else : 
                  if Nb_var<=20 : 
                       plt.subplot(5,4,i) 
                  else :
                      plt.subplot(5,5,i)
           plt.plot(bin_edgesSHUFFLE,hist_sum_SHUFFLE[i])
#           n_SHUFFLE, bins_SHUFFLE, patches_SHUFFLE = plt.hist(hist_sum_SHUFFLE[i], bin_edgesSHUFFLE)
           plt.axis([minima_tot, maxima_tot,0,hist_sum_SHUFFLE[i].max()])
   
#   COMPUTE THE HISTOGRAMS OF THE LIST OF INFOMUT VALUES FOR EACH DEGREE  
#   If shuffle is true it also compute the signifiance test against independence null hypothesis  
   num_fig=num_fig+1
   fig_Histo_infomut = plt.figure(num_fig)
   if compute_shuffle == True:
       low_signif_bound={}
       high_signif_bound={}
       lign_signif={}
   for a in range(1,Nb_var+1):
       if Nb_var<9 :  
           plt.subplot(3,3,a)
       else :    
           if Nb_var<=16 :
               plt.subplot(4,4,a)
           else : 
              if Nb_var<=20 : 
                   plt.subplot(5,4,a) 
              else :
                  plt.subplot(5,5,a)
#       if a == 2:
#         print(ListInfomutordre[a])           
       ListInfomutordre[a].append(minima_tot-0.1)
       ListInfomutordre[a].append(maxima_tot+0.1)                   
       n, bins, patches = plt.hist(ListInfomutordre[a], resultion_histo, facecolor='r')
       
       if compute_shuffle == True:
           plt.plot(bin_edgesSHUFFLE,hist_sum_SHUFFLE[a]*(1/nb_of_shuffle),color="blue")
           cumul=0
           first_signif = True
           second_signif = False
           lign_signif[a]=[]
           lign_signif[a] = np.zeros_like(hist_sum_SHUFFLE[a])
           for x in range(0,len(hist_sum_SHUFFLE[a])):
               cumul=cumul+hist_sum_SHUFFLE[a][x] 
#               print('cumul = ',  cumul)
               if first_signif == True:
                   if cumul >= (binomial(Nb_var,a)*nb_of_shuffle*p_value):
#                       print('nb low signif = ',  (binomial(Nb_var,a)*nb_of_shuffle*p_value))                       
                       low_signif_bound[a] = bin_edgesSHUFFLE[x]
                       lign_signif[a][x] =hist_sum_SHUFFLE[a].max()
                       first_signif = False
                       second_signif = True
                       print('low_signif_bound in dim',a, ' = ',  low_signif_bound[a])
               if second_signif == True: 
                   if cumul >= (binomial(Nb_var,a)*nb_of_shuffle*(1-p_value)):
                       high_signif_bound[a] = bin_edgesSHUFFLE[x]
                       lign_signif[a][x] =hist_sum_SHUFFLE[a].max()
                       second_signif = False
                       print('high_signif_bound in dim',a, ' = ',  high_signif_bound[a]) 
           nb_of_signif_low = 0     
           nb_of_signif_high = 0
           for x in range(0,len(ListInfomutordre[a])): 
               if ListInfomutordre[a][x] <= low_signif_bound[a] :                  
                  nb_of_signif_low = nb_of_signif_low+1 
               if ListInfomutordre[a][x] >= high_signif_bound[a] :
                  nb_of_signif_high = nb_of_signif_high+1               
           print('nb_of_signif_low in dim',a, ' = ',  nb_of_signif_low)
           print('nb_of_signif_high in dim',a, ' = ',  nb_of_signif_high) 
           plt.plot(bin_edgesSHUFFLE,lign_signif[a]*(1/nb_of_shuffle),color="green")             
       plt.axis([minima_tot, maxima_tot,0,n.max()])
       if a==1 :
          matrix_distrib_infomut=n
       else: 
# np.c_ concatenates  along the second axis.           
          matrix_distrib_infomut=np.c_[matrix_distrib_infomut,n]

       plt.grid(True)
   if save_landscape_object == True: 
       save_obj(fig_Histo_infomut,'Infomut_distrib')
   if save_results == True:    
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="svg")  

   
 #   COMPUTE THE INFOMUT LANDSCAPE FROM THE HISTOGRAMS AS THE MATRIX  matrix_distrib_infomut
#   If shuffle is true it also plots the signifiance test against independence null hypothesis  
    
   num_fig=num_fig+1
   fig_infolandscape =plt.figure(num_fig)  
   matrix_distrib_infomut=np.flipud(matrix_distrib_infomut)
   multiply_by_info_values= False
   if multiply_by_info_values :
        for i in range(0,resultion_histo):    
           matrix_distrib_infomut[i,:]=matrix_distrib_infomut[i,:] * ((maxima_tot-minima_tot+0.2) *(resultion_histo-i)/(resultion_histo-1) +minima_tot-0.1)
           print((maxima_tot-minima_tot+0.2) *(resultion_histo-i) /(resultion_histo-1) +minima_tot-0.1)
           print('i is ')
           print((resultion_histo-i))
           print(matrix_distrib_infomut[i,:])
   
#   plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=Nb_var/(2*resultion_histo),extent=[1,Nb_var,minima_tot-0.1,maxima_tot+0.1])
#   norm = matplotlib.colors.Normalize(vmin = 1, vmax = np.max(matrix_distrib_infomut))
   if multiply_by_info_values :
        if ((-1)*matrix_distrib_infomut.min()) <= (matrix_distrib_infomut.max()): 
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=-matrix_distrib_infomut.max(), vmax=matrix_distrib_infomut.max())
        else :
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=matrix_distrib_infomut.min(), vmax=-matrix_distrib_infomut.min())
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=norm)
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   else :
#        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm(vmin=1, vmax=200000))
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm())
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   if compute_shuffle == True:     
       abssice=np.linspace(0.5, Nb_var-0.5, Nb_var)
       low_ordinate=[]
       high_ordinate=[]
       for a in range(1,Nb_var+1):
           low_ordinate.append(low_signif_bound[a])
           high_ordinate.append(high_signif_bound[a])
#       print('abssice')
#       print(abssice)
#       print('low_ordinate')
#      print(low_ordinate)
#       print('high_ordinate')
#       print(high_ordinate)
       plt.plot(abssice, low_ordinate, marker='o', color='black')
       plt.plot(abssice, high_ordinate, marker='o',color='black')     
   plt.colorbar()
   plt.grid(False)
   if save_landscape_object == True: 
       save_obj(fig_infolandscape,'Infolandscape')

   if SHOW_HARD_DISPLAY_medium== True:        
       plt.scatter(x_absss, y_absss, marker= "o" , facecolor= 'w')         



##########################################
##########################################
##########################################
##########################################


   x_absss = np.array([])
   y_absss = np.array([]) 
# Display every Histo with its own scales:   
#   minim={}
#   maxim={}
   
   maxima_tot_entropy=-1000000.00
   minima_tot_entropy=1000000.00   
#   ListEntropyordre={}
   undersampling_percent=np.array([])

#   for i in range(1,Nb_var+1):
#     ListEntropyordre[i]=[]

   for x,y in Nentropie.items():
#      ListEntropyordre[len(x)].append(y)
      if SHOW_HARD_DISPLAY_medium== True:
          if len(x)>0:
#             print(len(x))
#             print(y)
                 x_absss=np.append(x_absss,[len(x)-0.5])
                 y_absss=np.append(y_absss,[y])
# Display every Histo with its own scales:          
#      minim[len(x)]=minim.get(len(x),1000000.00)
#      maxim[len(x)]=maxim.get(len(x),-1000000.00)
# Display every Histo with its own scales:      
#      if y>maxim[len(x)]:
#          maxim[len(x)]=y
#     if y<minim[len(x)]:
#          minim[len(x)]=y    
      if y>maxima_tot_entropy:
          maxima_tot_entropy=y
      if y<minima_tot_entropy:
          minima_tot_entropy=y     
   for a in range(1,Nb_var+1):
       if Nb_var<=9 :
#       Numgraph=330+a
           plt.subplot(3,3,a)
       else :    
           if Nb_var<=16 :
               plt.subplot(4,4,a)
           else : 
              if Nb_var<=20 : 
                   plt.subplot(5,4,a) 
              else :
                  plt.subplot(5,5,a)           
       ListEntropyordre[a].append(minima_tot_entropy-0.1)
       ListEntropyordre[a].append(maxima_tot_entropy+0.1)     
       plt.hist2d(ListEntropyordre[a], ListInfomutordre[a], bins=resultion_histo, norm=LogNorm())
       cbar = plt.colorbar()
       cbar.ax.set_ylabel('Counts')
#    print 'listentropie à lordre:',a
#    print ListEntropyordre[a]
       plt.axis([minima_tot_entropy, maxima_tot_entropy,minima_tot,maxima_tot])
 
            
    
#   mng = plt.get_current_fig_manager(num_fig)                                         
#   mng.window.showMaximized(num_fig)
#   plt.tight_layout(num_fig) 
   if save_results == True: 
      plt.savefig(os.path.join(directory_path,'INFOMUT_ENTROPY'))    
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'INFOMUT_ENTROPY'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'INFOMUT_ENTROPY'), format="svg")   
 
   
###############################################################
########          Mean Info vs. Graining N           ##########
###                mean Mutual information           #########
######  as a function of graining (temperature)      #########
######          and dimension k  <Ik>=f(k,N)         #########
###############################################################       
   
if SHOW_results_INFOMUT_Normed_per_bins == True: 
    num_fig=num_fig+1
    fig_meaninfo_per_binsize =plt.figure(num_fig) 
    Matrix_meaninfomut_per_binsize = np.zeros((Nb_var+2,Nb_of_N+2))
    for k in range(Nb_of_N):     
       print('Nb of bins N =',k+2)
       name_object= 'INFOMUT_SUM'+str(k)
       infomut_sum_order = load_obj(name_object)     
       maxordonnee=-1000000.00
       minordonnee=1000000.00
#       normed_informut={} 
       for x,y in infomut_sum_order.items():
           rate=y/binomial(Nb_var,x)
           print('Nb of bins N =',k+2, 'dim k =',x, 'meaninfomut =',y/binomial(Nb_var,x))
           Matrix_meaninfomut_per_binsize[x,k+2]=y/binomial(Nb_var,x)
       infomut_sum_order.clear()
       del(infomut_sum_order)    
    Matrix_meaninfomut_per_binsize=Matrix_meaninfomut_per_binsize.transpose()
#    Matrix_meaninfomut_per_binsize=np.flipud(Matrix_meaninfomut_per_binsize)
    norm = mpl.colors.Normalize(vmin=Matrix_meaninfomut_per_binsize.min(),vmax=Matrix_meaninfomut_per_binsize.max())
#    norm=mpl.colors.SymLogNorm(linthresh=1, vmin=-Matrix_meaninfomut_per_binsize.max(), vmax=Matrix_meaninfomut_per_binsize.max())   
#    plt.matshow(Matrix_meaninfomut_per_binsize, cmap='jet', aspect=3, extent=[1,Nb_var,2,Nb_of_N+2],norm=norm)
    plt.matshow(Matrix_meaninfomut_per_binsize, cmap='jet', aspect=3,norm=norm)
    

    plt.axis([0.5,Nb_var+0.5,1.5,Nb_of_N+1.5])    
#    plt.axes().set_aspect('auto')
    plt.ylabel('(# Bin N)')
    plt.xlabel('(Dim k)')
    plt.title('Mean Information vs dimension and graining. <Ik>=f(k,N)')
    plt.colorbar()
#    plt.axes().set_aspect('equal')
    
    plt.grid(True) 
    
    
    
###############################################################
########          Mean Info vs. Sample-SIZE m        ##########
###                mean Mutual information            #########
######  as a function of Sample size (m number of trials)######
#####          and dimension k  <Ik>=f(k,m)           #########
###############################################################       
   
if SHOW_results_INFOMUT_Normed_per_samplesize == True: 
    num_fig=num_fig+1
    fig_meaninfo_per_samplesize =plt.figure(num_fig) 
    Matrix_meaninfomut_per_samplesize = np.zeros((Nb_var+2,Nb_of_m+2))
    Nb_trials_max=Nb_trials
    for k in range(Nb_of_m):  
       Nb_trials =Nb_trials_max- k*int(Nb_trials_max/(Nb_of_m))
       print('Nb_trials=',Nb_trials)
       name_object= 'INFOMUT_SUM'+str(k)
       infomut_sum_order = load_obj(name_object)     
       maxordonnee=-1000000.00
       minordonnee=1000000.00
#       normed_informut={} 
       for x,y in infomut_sum_order.items():
           rate=y/binomial(Nb_var,x)
           print('Nb_trials=',Nb_trials, 'dim k =',x, 'meaninfomut =',y/binomial(Nb_var,x))
           Matrix_meaninfomut_per_samplesize[x,k+2]=y/binomial(Nb_var,x)
       infomut_sum_order.clear()
       del(infomut_sum_order)    
    Matrix_meaninfomut_per_samplesize=Matrix_meaninfomut_per_samplesize.transpose()
#    Matrix_meaninfomut_per_binsize=np.flipud(Matrix_meaninfomut_per_binsize)
    norm = mpl.colors.Normalize(vmin=Matrix_meaninfomut_per_samplesize.min(),vmax=Matrix_meaninfomut_per_samplesize.max())
#    norm=mpl.colors.SymLogNorm(linthresh=1, vmin=-Matrix_meaninfomut_per_binsize.max(), vmax=Matrix_meaninfomut_per_binsize.max())   
#    plt.matshow(Matrix_meaninfomut_per_binsize, cmap='jet', aspect=3, extent=[1,Nb_var,2,Nb_of_N+2],norm=norm)
    plt.matshow(Matrix_meaninfomut_per_samplesize, cmap='jet', aspect=3,norm=norm)
    

    plt.axis([0.5,Nb_var+0.5,1.5,Nb_of_m+1.5])    
#    plt.axes().set_aspect('auto')
    plt.ylabel('(# sampleSize m)')
    plt.xlabel('(Dim k)')
    plt.title('Mean Information vs dimension and sampleSize. <Ik>=f(k,m)')
    plt.colorbar()
#    plt.axes().set_aspect('equal')
    
    plt.grid(True) 
        
    



###############################################################
########          Ring representation               ##########
###                Mutual information               #########
###### http://networkx.readthedocs.org/en/stable/   #########
###############################################################    
#import networkx as nx

#plt.figure(9)


#if SHOW_results_SCAFOLD == True:

#   netring = nx.Graph()
#   list_of_node=[]
#   list_of_size=[]
#   list_of_edge=[]
#   list_of_width=[]


#   for x,y in infomut_per_order[1].items():
#        tuple_interim=(workbook_data[Name_worksheet].cell(row = x[0]+1, column = 1).value )
#        list_of_size.append((infomut_per_order[1][x]**2) *500)
#        list_of_node.append(tuple_interim)

#   tuple_interim=()
#   for x,y in infomut_per_order[2].items():    
#        tuple_interim=(workbook_data[Name_worksheet].cell(row = x[0]+1, column = 1).value,workbook_data[Name_worksheet].cell(row = x[1]+1, column = 1).value)
#        list_of_edge.append(tuple_interim)
#        netring.add_edge(tuple_interim[0], tuple_interim[1], weight= (infomut_per_order[2][x]) )
#        list_of_width.append((infomut_per_order[2][x]*10))
#     list_of_width.append(0.5*(infomut_per_order[2][x]*10)**2)

    
#   netring.add_nodes_from(list_of_node)
    

#   nx.draw_circular(netring, with_labels=True, nodelist = list_of_node,edgelist = list_of_edge, width= list_of_width, node_size = list_of_size)

#   mng = plt.get_current_fig_manager(9)                                         
#   mng.window.showMaximized(9)
#plt.tight_layout()     
#   if save_results == True: 
#      plt.savefig(os.path.join(directory_path,'SCAFOLD_INFOMUT_PAIR'), format="svg") 
#      if Format_SVG_SAVE_Figure == False: 
#         plt.savefig(os.path.join(directory_path,'SCAFOLD_INFOMUT_PAIR'), format="png")
#      if Format_SVG_SAVE_Figure == True:       
#         plt.savefig(os.path.join(directory_path,'SCAFOLD_INFOMUT_PAIR'), format="svg")  
#   if display_figure == True:
#      plt.show(9)
#   else: plt.close(9)  
#else: plt.close(9)    






if SHOW_results_SCAFOLD == True:
   num_fig=num_fig+1
   plt.figure(num_fig)  

   netring = nx.Graph()
   list_of_node=[]
   list_of_size=[]
   list_of_edge=[]
   list_of_width=[]
   list_of_labels={}
   mapping={}
   
   for x in range(1,Nb_var+1):
       tuple_interim=(workbook_data[Name_worksheet].cell(row = x, column = 1).value )
       list_of_node.append(x)
       list_of_labels[x]=tuple_interim
       cagoenlaputa=[x]
       mapping[x]=tuple_interim
       code_var=tuple(cagoenlaputa)
       list_of_size.append((infomut_per_order[1][code_var]**2) *500)
       netring.add_node(x)
       
#   for x,y in infomut_per_order[1].items():
#       tuple_interim=(workbook_data[Name_worksheet].cell(row = x[0]+1, column = 1).value )
#        list_of_size.append((infomut_per_order[1][x]**2) *500)
#        list_of_node.append(tuple_interim)

   tuple_interim=()
   for x,y in infomut_per_order[2].items():    
        tuple_interim=(workbook_data[Name_worksheet].cell(row = x[0]+1, column = 1).value,workbook_data[Name_worksheet].cell(row = x[1]+1, column = 1).value)
#        list_of_edge.append(tuple_interim)
        list_of_edge.append(x)
        var_1=x[0]
        var_2=x[1]
        netring.add_edge(var_1, var_2, weight= (infomut_per_order[2][x]) )
        list_of_width.append((infomut_per_order[2][x]*10))
#     list_of_width.append(0.5*(infomut_per_order[2][x]*10)**2)

   print(list_of_node) 
#   netring=nx.complete_graph(Nb_var+5)
#   netring.add_nodes_from(list_of_node)
  
  
#   G=nx.relabel_nodes(netring,mapping)
   nx.draw_circular(netring, with_labels= True, nodelist = list_of_node,edgelist = list_of_edge, width= list_of_width, node_size = list_of_size)
#   G=nx.relabel_nodes(G,mapping)

#   mng = plt.get_current_fig_manager(10)                                         
#   mng.window.showMaximized(10)
#plt.tight_layout()     
   if save_results == True: 
      plt.savefig(os.path.join(directory_path,'SCAFOLD_INFOMUT_PAIR'), format="svg") 
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'SCAFOLD_INFOMUT_PAIR'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'SCAFOLD_INFOMUT_PAIR'), format="svg")  
   if display_figure == True:
      plt.show()
   else: plt.close(num_fig)  
#else: plt.close(num_fig)    


##############################
#infomut_per_order=[]
#Ninfomut_per_order_ordered=[]
#for x in range(Nb_var+1):
#    infomut_dicoperoder={} 
#    infomut_per_order.append(infomut_dicoperoder)
#    Ninfomut_per_order_ordered.append(infomut_dicoperoder)
    
#for x,y in Ninfomut.items():
#    print('lenghtx=',len(x))
#    infomut_per_order[len(x)][x]=Ninfomut[x]

#for x in range(Nb_var+1):
#    Ninfomut_per_order_ordered[x]=OrderedDict(sorted(infomut_per_order[x].items(), key=lambda t: t[1]))
#    print('les infos mutuelles à ', x ,'sont:')
#    print(Ninfomut_per_order_ordered[x])

#Ninfomut_ordered=OrderedDict(sorted(Ninfomut.items(), key=lambda t: t[1]))





#########################################################################
#########################################################################
#########################################################################
#########################################################################
######      RELOAD INFO LANDSCAPES FIGURES           ####################
#########################################################################
#########################################################################
#########################################################################
#########################################################################


if SHOW_results_RELOAD_LANDSCAPES == True:   
    
    num_fig=num_fig+1
    fig_infolandscape =plt.figure(num_fig) 
    name_object= 'Infolandscape'
    fig_infolandscape = load_obj(name_object)   
    fig_infolandscape.show()
    num_fig=num_fig+1
    fig_Histo_infomut =plt.figure(num_fig) 
    name_object= 'Infomut_distrib'
    fig_Histo_infomut = load_obj(name_object)  
    fig_Histo_infomut.show()
    
if SHOW_results_PRINT_PLK_file == True:   
    
    if load_results_MATRIX == True:
        name_object= 'Matrix_data2'
        if choose_number_load :
            name_object= name_object+str(number_load)
        object_to_print = load_obj(name_object)  
        print(object_to_print)
        num_fig=num_fig+1
        fig_mat =plt.figure(num_fig)       
        plt.matshow(object_to_print)
        plt.title('Matrix reloaded')
        fig_mat.show()
        num_fig=num_fig+1
        fig_distrib_2var =plt.figure(num_fig)        
        VarX_sampled=[]
        VarY_sampled=[]
        VarX_raw=[]
        VarY_raw=[]
        for col in range(Nb_trials):
            VarX_raw.append(Matrix_data[variable_x,col])
            VarY_raw.append(Matrix_data[variable_y,col])
            VarX_sampled.append(object_to_print[variable_x,col])
            VarY_sampled.append(object_to_print[variable_y,col])
        fig_distrib_2var = plt.hist2d(VarX_sampled, VarY_sampled, bins=Nb_bins, cmap='binary')
# Problem to check !!!  
#        fig_distrib_2var = plt.hist2d(VarX_sampled, VarY_sampled, bins=Nb_bins+1, cmap='binary')
        cbar = plt.colorbar()
        cbar.ax.set_ylabel('Counts')
#        plt.matshow(object_to_print)
        plt.title('Joint distribution')
#        fig_distrib_2var.show()
        num_fig=num_fig+1        
        fig2 = plt.figure(num_fig)
        ax2 = fig2.add_subplot(111)
        pts2=ax2.scatter(VarX_raw, VarY_raw,c= 'red' , marker='8')
        ax2.legend()
        ax2.grid(True)
        
        num_fig=num_fig+1
        plt.figure(num_fig)
        hist, bin_edges = np.histogram(VarX_sampled, bins = Nb_bins+1,range=(0,Nb_bins+1))
        plt.bar(bin_edges[:-1], hist, width = 1)
        plt.xlim(min(bin_edges), max(bin_edges))
        num_fig=num_fig+1
        plt.figure(num_fig)
        hist, bin_edges = np.histogram(VarY_sampled, bins = Nb_bins+1,range=(0,Nb_bins+1))
        plt.bar(bin_edges[:-1], hist, width = 1)
        plt.xlim(min(bin_edges), max(bin_edges))
#        plt.show()
#    print 'listentropie à lordre:',a
#    print ListEntropyordre[a]
        
    if load_results_ENTROPY == True: 
        name_object= 'ENTROPY'
        if choose_number_load :
            name_object= name_object+str(number_load)
        object_to_print = load_obj(name_object)
        print(object_to_print[degree_to_print])

    if load_results_ENTROPY_ORDERED == True: 
        name_object= 'ENTROPY_ORDERED'
        if choose_number_load :
            name_object= name_object+str(number_load)
        object_to_print = load_obj(name_object)
        print(object_to_print[degree_to_print])

    if load_results_INFOMUT == True: 
       name_object= 'INFOMUT'
       if choose_number_load :
            name_object= name_object+str(number_load)
       object_to_print = load_obj(name_object) 
       print(object_to_print[degree_to_print])

    if load_results_ENTROPY_SUM == True:  
        name_object= 'ENTROPY_SUM'
        if choose_number_load :
            name_object= name_object+str(number_load)
        object_to_print = load_obj(name_object)
        print(object_to_print[degree_to_print])

    if load_results_INFOMUT_ORDERED == True: 
        name_object= 'INFOMUT_ORDERED'
        if choose_number_load :
            name_object= name_object+str(number_load)
        object_to_print = load_obj(name_object)
        name_object= 'INFOMUT_ORDEREDList'
        if choose_number_load :
            name_object= name_object+str(number_load)
        infomut_per_order = load_obj(name_object)
        print(object_to_print[degree_to_print])

    if load_results_INFOMUT_SUM == True: 
        name_object= 'INFOMUT_SUM'
        if choose_number_load :
            name_object= name_object+str(number_load)
        object_to_print = load_obj(name_object)
        print(object_to_print[degree_to_print])
    
    if load_results_INFOMUT_SUM == True: 
        name_object= 'INFOMUT_SUM'
        if choose_number_load :
            name_object= name_object+str(number_load)
        object_to_print = load_obj(name_object)    
        print(object_to_print[degree_to_print])
    



   


print('boulot fini (Halt)')
