# -*- coding: utf-8 -*-
"""
Created on Thu Nov  5 16:06:46 2015

@author: baudot
"""

# ************************************************************
# Reading of datafiles (specific of .CSV files or XLS file)
#csvkit manual: http://csvkit.readthedocs.org/en/latest/cli.html
# ************************************************************

import math
from math import *
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter.messagebox import *
from tkinter import filedialog
import pickle
import itertools

#################################################################
### THE INPUT  PARAMETERS OF THE PROGRAM  #######################
#################################################################

directory_path =  os.path.abspath("/home/baudot/LABOMARSEILLE/RESULTAT ANALYSE PCR/ACTUAL RESEARCH/ANALYSE 21RNA MARS/CLUSTER_10_10/")
Name_worksheet="DOPA"
Nb_bins=8
# Nb var is the number n used for the computation of n-tuple information
Nb_var=9
# Nb var tot is the total number m of variable used for the computation of n-tuple information
# we have necessarily m>n
Nb_var_tot=9

#display_figure = True
display_figure = False
#matplotlib.use('agg')
#matplotlib.use('svg')
Format_SVG_SAVE_Figure = True
save_results = True
#save_results = False

#################################################################
### Procedure for saving and loading objects ####################
#################################################################

def save_obj(obj, name ):
#    with open('obj/'+ name + '.pkl', 'wb') as f:
    with open(directory_path + '/' + name + '.pkl', 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)

def load_obj(name ):
    with open(directory_path + '/' + name + '.pkl', 'rb') as f:
        return pickle.load(f)

#################################################################
### dialog open file XLS DATA FILE ##############################
#################################################################

root = Tk()

root.filename =  filedialog.askopenfilename(initialdir = directory_path,title = "choose your file",filetypes = (("xls files","*.xlsx"),("all files","*.*")))
root.destroy()

workbook_data = load_workbook(root.filename)

########################################################################
### Tools to manipulate WORKBOOK  ######################################
# https://openpyxl.readthedocs.org/en/2.3.3/index.html #################
########################################################################


#Name_worksheet="NONDOPA_MONICA_SIMONE"

print(workbook_data.get_sheet_names())
d = workbook_data[Name_worksheet].cell(row = 2, column = 5)
print(d.value)
c = workbook_data[Name_worksheet]['B1':'B239']
print(c)


###############################################################
### procedure for resampling and displaying  ################## 
###     probability distributions            ##################
###############################################################


Nb_column=1
for i in range(2,10000):
    if workbook_data[Name_worksheet].cell(row = 1, column = i).value == None:
       Nb_column=i-2
       break
   
print('the number of column (or trials n) is: ', Nb_column)   

Nb_trials=Nb_column


#We put the XLSSheet into a numpy matrix nparray #############
# Variables (RNA) are the lines #
# Trials are the columns  #
# Matrix_data contains the RAW DATA
# Matrix_data contains the  DATA resampled and rescaled

Matrix_data = np.zeros((Nb_var_tot,Nb_trials))
Matrix_data2 = np.zeros((Nb_var_tot,Nb_trials))

# WE PUT ALL THE 0 value to nan  #
for col in range(Nb_trials):
  for row in range(Nb_var_tot):
      if workbook_data[Name_worksheet].cell(row = (row+2), column = (col+2)).value == 0:
         Matrix_data[row,col] = np.nan
      else:
         Matrix_data[row,col] = workbook_data[Name_worksheet].cell(row = (row+2), column = (col+2)).value



Min_matrix = np.nanmin(Matrix_data, axis=1)
Max_matrix = np.nanmax(Matrix_data, axis=1) 
Ampl_matrix = Max_matrix - Min_matrix

# WE RESCALE ALL MATRICES AND SAMPLE IT into  Nb_bins #
for col in range(Nb_trials):
  for row in range(Nb_var_tot):
      if np.isnan(Matrix_data[row,col]):
         Matrix_data[row,col] = 0
         Matrix_data2[row,col] = 0
         workbook_data["DOPA_SAMPLED"].cell(row = (row+2), column = (col+2)).value = 0 
      else:
         if Ampl_matrix[row] !=0 :
             Matrix_data2[row,col] = int(((Matrix_data[row,col]-Min_matrix[row])*(Nb_bins-2))/(Ampl_matrix[row]))+1
         else:    
             Matrix_data2[row,col] = 0
         workbook_data["DOPA_SAMPLED"].cell(row = (row+2), column = (col+2)).value = Matrix_data2[row,col]          
        
         

###############################################################
### AFFICHAGE ET ENREGISTREMENT DES DATA RESAMPLEES  ##########
###                  ET RESCALEES            ##################
###############################################################



workbook_data.save(os.path.join(directory_path,"Sampled_DATA.xlsx"))     
#plt.figure(1)
plt.matshow(Matrix_data)
if save_results == True: 
   if Format_SVG_SAVE_Figure == False: 
      plt.savefig(os.path.join(directory_path,'Matrix_RAW_VALUE'), format="png")
   if Format_SVG_SAVE_Figure == True:       
      plt.savefig(os.path.join(directory_path,'Matrix_RAW_VALUE'), format="svg")  

#plt.figure(2)
plt.matshow(Matrix_data2)
if save_results == True:
   if Format_SVG_SAVE_Figure == False: 
      plt.savefig(os.path.join(directory_path,'Matrix_RESCALED_VALUE'), format="png")
   if Format_SVG_SAVE_Figure == True:       
      plt.savefig(os.path.join(directory_path,'Matrix_RESCALED_VALUE'), format="svg")  

plt.figure(3)
hist, bin_edges = np.histogram(Matrix_data2, bins = Nb_bins,range=(0,Nb_bins))
plt.bar(bin_edges[:-1], hist, width = 1)
plt.xlim(min(bin_edges), max(bin_edges))
if display_figure == True:
   plt.show()
if save_results == True: 
   if Format_SVG_SAVE_Figure == False: 
      plt.savefig(os.path.join(directory_path,'histogram_RESCALED_VALUE'), format="png")
   if Format_SVG_SAVE_Figure == True:       
      plt.savefig(os.path.join(directory_path,'histogram_RESCALED_VALUE'), format="svg")  
   
plt.figure(4)
hist, bin_edges = np.histogram(Matrix_data, bins = 52,range=(0,26))
plt.bar(bin_edges[:-1], hist, width = 0.5)
plt.xlim(min(bin_edges), max(bin_edges))
if display_figure == True:
   plt.show()
if save_results == True: 
   if Format_SVG_SAVE_Figure == False: 
      plt.savefig(os.path.join(directory_path,'histogram_RAW_VALUE'), format="png")
   if Format_SVG_SAVE_Figure == True:       
      plt.savefig(os.path.join(directory_path,'histogram_RAW_VALUE'), format="svg")  
   




    
###############################################################
########          Calculus probability               ##########
###              probability distributions            #########
###############################################################

#### To avoid to have to explore all  possible  probability (sparse data)
# we encode probability as dictionanry, each existing probability
# has a key  
def compute_probability(Nb_trials_input,Nb_var_input,Matrix_data2_input):
    probability={}

    for col in range(Nb_trials_input): 
        x=''
        for row in range(Nb_var_input):
           x= x+str(int((Matrix_data2_input[row,col])))
        probability[x]=probability.get(x,0)+1

#    print(probability) 
    Nbtot=0  
    for i in probability.items():
        Nbtot=Nbtot+i[1]
#    print('le nombre total doccurence  est:',Nbtot)
    for i,j in probability.items():
           probability[i]=j/float(Nbtot)   
    return (probability)  
      
      
       
###############################################################
########          SOME FUNCTIONS USEFULLS            ##########
###          AT ALL ORDERS On SET OF SUBSETS          #########
###############################################################       
       
       
       
def information(x):
	return -x*math.log(x)/math.log(2) 
 
# Calcul proba et entropie à tous les ordres 

# for i in range(1,Longueurtot+1):  
Nentropie={} 



# Fonction factorielle 

def factorial(x):
     if x < 2:
         return 1
     else:
         return x * factorial(x-1)
   
# Fonction coeficient binomial (nombre de combinaison de k elements dans [1,..,n]) 

def binomial(n,k):
    return factorial(n)/(factorial(k)*factorial(n-k))
    
#############################################################################    
# Fonction decode(x,n,k,combinat)
#--> renvoie la combinatoire combinat de k variables dans n codée par x
# dans combinat
#les combinaisons de k élements dans [1,..,n] sont en bijection avec 
# les entiers x de [0,...,n!/(k!(n-k)!)-1]  
# attention numerotation part de 0     
#############################################################################
    
def decode(x,n,k,combinat):
    if x<0 or n<=0 or k<=0:
        return 
    b= binomial(n-1,k-1)
    if x<b:
        decode(x,n-1,k-1,combinat)
        combinat.append(n)
    else:
        decode(x-b,n-1,k,combinat) 

        
#############################################################################
# Fonction decode_all(x,n,k,combinat)
#--> renvoie la combinatoire (combinat) et l'ordre k associé au code x 
# x varie de 0 à (2^n)-1, les n premiers x code pour 1 parmis n
# les suivants codent pour 2 parmis n
# etc... jusquà x=(2^n)-1 qui code pour n parmis n
#les combinaisons de k élements dans [1,..,n] sont en bijection avec 
# les entiers x de [0,...,n!/(k!(n-k)!)-1]  
# attention numerotation part de 0   
#############################################################################

def decode_all(x,n,order,combinat):
    sumtot=n
    order=1
    Code_order=x
    while x>=sumtot:
       order=order+1
       sumtot=sumtot+binomial(n,order) 
       Code_order=Code_order-binomial(n,order-1)
    k=order
    decode(Code_order,n,k,combinat)  
     



###############################################################
###  COMBINATORIAL SUB MATRIX SELECTION             ##########
###                                          ##################
###############################################################
import sys

def compute_sub_matrix_aux(x_input,Nb_var_tot_input,Nb_var_input,Matrix_data2_input,combinat):
    decode(x_input,Nb_var_tot_input,Nb_var_input,combinat)  
         
    
def compute_sub_matrix(x_input,Nb_var_tot_input,Nb_var_input,Nb_trials_input,Matrix_data2_input,combinat_input):
      
#    compute_sub_matrix_aux(x_input,Nb_var_tot,Nb_var,Matrix_data2,ntuple1)
#    decode(x_input,Nb_var_tot_input,Nb_var_input,combinat_input)
#    print('matrix depard', Matrix_data2_input)
#    if len(combinat_input) ==   Nb_var_input :      
#            print('cobinat is', combinat_input)
            global Matrix_data_temp
            Matrix_data_temp = np.zeros((Nb_var_input,Nb_trials_input))
            for num_row in range(Nb_var_input):
#                print('num_row', num_row)
                val_row = combinat_input[num_row]-1
                for col in range(Nb_trials_input):
                   #col = col-2
#                   print('col', col)
#                   print('lign', combinat_input)
#                   print('val matrix raw', Matrix_data[val_row,col])
#                   print('val matrix renorm', Matrix_data2[val_row,col])
 #                  print('cobinat is', combinat_input)    
                   Matrix_data_temp[num_row,col] = Matrix_data2_input[val_row,col]
                   #if col > 5 :
                      #sys.exit()
#            print('cobinat is', combinat_input)     
#            print('matrix depard', Matrix_data2_input)
#            print('matrix data temp', Matrix_data_temp)
            return (Matrix_data_temp)        

##################################################################################
###############    COMPUTE ENTROPY                         #######################
##################################################################################   



def compute_entropy(Nb_var_input,probability_input,ntuple1_input):
#    Nentropie={} 
    sys.getrecursionlimit()
    ntuple=[]        
    for x in range(0,(2**Nb_var_input)-1): 
        ntuple=[]
        orderInf=0
        decode_all(x,Nb_var_input,orderInf,ntuple)   
#        print len(ntuple)
#        print ntuple
        tuple_code=()        
        probability2={}
        for z in range(0,len(ntuple)):
            concat=()
#            concat=(ntuple1_input[z],)
            concat=(ntuple1_input[ntuple[z]-1],)
            tuple_code=tuple_code+concat
#            print('code dans boucle:',tuple_code)
#        print('code local:',tuple_code)
#        print('code global:',ntuple1_input)
 #       tuple_code = change_code(tuple_code,ntuple1_input)
        for x,y in probability_input.items():
            Codeproba=''
            length=0
            for w in range(1,Nb_var_input+1):
#                print 'w='
#                print w
#                print length
#                print len(ntuple)
#                print ntuple[length]            
                if ntuple[length]!=w:
                    Codeproba=Codeproba+'0'                
                else: 
                    Codeproba=Codeproba+x[ntuple[length]-1:ntuple[length]]
                    if length<(len(ntuple)-1):
                        length=length+1
#                print Codeproba        
            probability2[Codeproba]=probability2.get(Codeproba,0)+probability_input.get(x,0)
#        print('probability:',probability2)  
        Nentropie[tuple_code]=0
# to change: the program computes too many times the entropy:  m*2**n instead of          
        for x,y in probability2.items():
#             if Nentropie[tuple_code]==0:
#                print('ta mere en short celui la y est pas:')
                Nentropie[tuple_code]=Nentropie.get(tuple_code,0)+information(probability2[x])   
        probability2={}  
        probability2.clear()
        del(probability2)
#    return (Nentropie)

##################################################################################
###############    SUMM  ENTROPY per order                 #######################
##################################################################################   
# somme des entropies à chaque ordre

def compute_sum_entropy_order(Nentropie_input):
    entropy_sum_order={} 
    for x,y in Nentropie_input.items():
        entropy_sum_order[len(x)]=entropy_sum_order.get(len(x),0)+Nentropie_input[x]
    return (entropy_sum_order)
        
##################################################################################
###############    COMPUTE MUTUAL INFORMATION              #######################
##################################################################################                
# Calcul des informations mutuelles à tous les ordres




#def compute_Ninfomut(Nentropie_input):    

#    for x,y in Nentropie_input.items():
     
#         for w,z in Nentropie_input.items():
#             test_in=0
#             for v in range(0,len(w)):

#                 if  w[v] in x: 
#                    test_in= test_in+0
#                 else: 
#                    test_in= test_in+1 
#             if test_in!=0 :
#               Ninfomut[x]=Ninfomut.get(x,0) 

#             else:   

#                 Ninfomut[x]=Ninfomut.get(x,0)+ ((-1)**(len(w)+1))*Nentropie_input[w]

#    return (Ninfomut)
    
#############################################################################
# Function binomial_subgroups COMBINAT Gives all binomial k subgroup of a group 
#############################################################################
     
    
    
def compute_Ninfomut(Nentropie_input):    
#    Ninfomut={}   
    for x,y in Nentropie_input.items():
        for k in range(1, len(x)+1):
           for subset in itertools.combinations(x, k):
#              print(subset)  
              Ninfomut[x]=Ninfomut.get(x,0)+ ((-1)**(len(subset)+1))*Nentropie_input[subset] 
    return (Ninfomut)  
    
    
##################################################################################
############### RANKING of the n higher Mutual information #######################
##################################################################################   

from operator import itemgetter
from collections import OrderedDict

# regular unsorted dictionary
# dictionary sorted by value
# Ninfo_per_order_ordered[x]  is ordered firstby each order x of information
# and second ordered by value of information 
def compute_Ninfo_per_order(Nb_var_input,Ninfo_input): 
    info_per_order=[]
    Ninfo_per_order_ordered=[]
    for x in range(Nb_var_input+1):
        info_dicoperoder={} 
        info_per_order.append(info_dicoperoder)
        Ninfo_per_order_ordered.append(info_dicoperoder)
    
    for x,y in Ninfo_input.items():
#       print('lenghtx=',len(x))
        info_per_order[len(x)][x]=Ninfo_input[x]

    for x in range(Nb_var_input+1):
        Ninfo_per_order_ordered[x]=OrderedDict(sorted(info_per_order[x].items(), key=lambda t: t[1]))
#        print('les infos mutuelles à ', x ,'sont:')
#        print(Ninfo_per_order_ordered[x])
    return (Ninfo_per_order_ordered,info_per_order)
#Ninfomut_ordered=OrderedDict(sorted(Ninfomut.items(), key=lambda t: t[1]))



##################################################################################
###############    Sum of mutual info per order            #######################
################################################################################## 

# somme des information à chaque ordre

def compute_infomut_sum_order(Ninfomut_input): 
    infomut_sum_order={} 
    for x,y in Ninfomut_input.items():
        infomut_sum_order[len(x)]=infomut_sum_order.get(len(x),0)+Ninfomut_input[x]

#    print('les info mutuelles sommées par ordre sont:')
#    print(infomut_sum_order)     

    infototbis=0.00    
    for x,y in infomut_sum_order.items():
        infototbis=infototbis+((-1)**(x+1))*infomut_sum_order.get(x,0)

#    print('l entropie calculée par les sommes est de:')
#    print(infototbis)   
    return (infomut_sum_order,infototbis)

# somme des valeur absolues des information à chaque ordre
def compute_infomut_sum_order_abs(Ninfomut_input): 
    infomut_sum_order_abs={} 
    for x,y in Ninfomut_input.items():
        infomut_sum_order_abs[len(x)]=infomut_sum_order_abs.get(len(x),0)+math.fabs(Ninfomut_input[x])

#    print('les valeur abs(info mutuelles) sommées par ordre sont:')
#    print(infomut_sum_order_abs)     

    infotot_absbis=0.00    
    for x,y in infomut_sum_order_abs.items():
        infotot_absbis=infotot_absbis+infomut_sum_order_abs.get(x,0)   

#    print('la somme des val abs de toutes le infomut est:')
#    print(infotot_absbis)         
    return (infomut_sum_order_abs,infotot_absbis)

#########################################################################
#########################################################################
###### CENTRAL PROGRAM ITERATION COMBINAT  ##############################
#########################################################################
#########################################################################



ntuple1=[]
# Nentropy and Ninfomut are dictionaries (x,y) with x a list of kind (1,2,5) and y a value in bit 
Nentropie={}  
Ninfomut={}
print('Ta mere en slip courre le marathon 1')
for x in range(0,(2**Nb_var_tot)-1): 
    ntuple1=[]  
    compute_sub_matrix_aux(x,Nb_var_tot,Nb_var,Matrix_data2,ntuple1)
    if len(ntuple1) ==   Nb_var :
        Matrix_data_temp = np.zeros((Nb_var,Nb_trials))
        Matrix_data_temp = compute_sub_matrix(x,Nb_var_tot,Nb_var,Nb_trials,Matrix_data2,ntuple1)
        probability = compute_probability(Nb_trials,Nb_var,Matrix_data_temp)
#        print('procedure CENTRAl combinat is', ntuple1)     
#        print('matrix depard', Matrix_data2)
#        print('matrix data temp', Matrix_data_temp)
#        print('the probability matrix is:')
#        print(probability)
#        Nentropie = compute_entropy(Nb_var,probability,ntuple1)
        compute_entropy(Nb_var,probability,ntuple1)
#        print('ENTROPY VALUES LIST',Nentropie)     
#        quit()
#       entropy_sum_order = compute_sum_entropy_order(Nentropie)
print('Ta mere en slip courre le marathon 2')
if save_results == True: 
    save_obj(Nentropie,'ENTROPY')   
    
    
print('Ta mere en slip courre le marathon 2.5')   
(Nentropy_per_order_ordered,entropy_per_order) = compute_Ninfo_per_order(Nb_var,Nentropie)
if save_results == True: 
    save_obj(Nentropy_per_order_ordered,'ENTROPY_ORDERED')

print('Ta mere en slip courre le marathon 2.7')      
Ninfomut = compute_Ninfomut(Nentropie)
if save_results == True: 
    save_obj(Ninfomut,'INFOMUT')
    
print('Ta mere en slip courre le marathon 3')
entropy_sum_order = compute_sum_entropy_order(Nentropie)
if save_results == True: 
    save_obj(entropy_sum_order,'ENTROPY_SUM')  
    
print('Ta mere en slip courre le marathon 4')
(Ninfomut_per_order_ordered,infomut_per_order) = compute_Ninfo_per_order(Nb_var,Ninfomut)
if save_results == True: 
    save_obj(Ninfomut_per_order_ordered,'INFOMUT_ORDERED')
    save_obj(infomut_per_order,'INFOMUT_ORDEREDList')
    
print('Ta mere en slip courre le marathon 5')
(infomut_sum_order,infototbis) = compute_infomut_sum_order(Ninfomut) 
if save_results == True: 
    save_obj(infomut_sum_order,'INFOMUT_SUM')

print('Ta mere en slip courre le marathon 6')
#(infomut_sum_order_abs,infotot_absbis) = compute_infomut_sum_order_abs(Ninfomut)


#for x in range(0,(2**Nb_var_tot)-1):     
#    print('les infos mutuelles à ', x ,'sont:')
#    print(Ninfomut_per_order_ordered[x])    
#print('la somme des val abs de toutes le infomut est:')
#print(infotot_absbis)         
#print('les valeur abs(info mutuelles) sommées par ordre sont:')
#print(infomut_sum_order_abs)
print('l entropie calculée par les sommes est de:')
print(infototbis) 
print('les info mutuelles sommées par ordre sont:')
print(infomut_sum_order) 


    

print('boulot fini')