# -*- coding: utf-8 -*-
"""
Created on Thu Nov  5 16:06:46 2015
Copyright (2007-2017) Pierre Baudot, Daniel Bennequin, Jean-Marc Goaillard, Monica Tapia 
The terms of the licence GNU GPL is reproduced in the file "COPYING" of INFOTOPO distribution.
 
This file is part of INFOTOPO.
 
  INFOTOPO is free software: you can redistribute it and/or modify
  it under the terms of the GNU General Public License as published by
  the Free Software Foundation, either version 3 of the License, or
  (at your option) any later version.
 
  INFOTOPO is distributed in the hope that it will be useful,
  but WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU General Public License for more details.
 
  You should have received a copy of the GNU General Public License
  along with INFOTOPO.  If not, see <http://www.gnu.org/licenses/>. 
  
@author: Pierre Baudot, Daniel Bennequin, Jean-Marc Goaillard, Monica Tapia 

GENERAL PHILOSOPHY AND AUTORS REQUEST: 
Giving the authorization of free use of this program also gives you the responsibility 
of its use: we trust you. Every tool can be used to make things better for all, but also worse.
We ask any users of the present program to use and develop it such that it profits to the 
sustainable development of all, societies and environment and avoid development
in selfish interest, conflict involvement, or individual freedom restrictions or survey. 
As a Human in the world, please be conscious of what you do, of its consequence on others, and try to make it better.
This tools comes from mathematic and pertain to it, mathematic do not pertain to anybody in particular 
and according to its original developments promotes harmony and sustainable developments on short 
and long time and space scales, please respect this ancestral tradition, old as humanity, and develop it in this respect. 
More information on the practical use and mathematical framework and bibliographic citations:
http://www.biorxiv.org/content/early/2017/07/26/168740 
http://www.mdpi.com/1099-4300/17/5/3253 
and README file. 
Baudot wrote the code and participated to theorems and algorithms, Bennequin participated to the theory and main theorems that the algorithm implement, 
Goaillard and Tapia participated to the development of the algorithms and program. 
For any requests, questions, improvements, developments (etc.) contact pierre.baudot [at] gmail.com
Friendly ergonomic interface will be developped afterward, sorry, please see the README file and/or contact me for use. 
INFOTOPO has been developed since 2007 thanks to grants-fundings-hosting of:
_ Institut de Mathématiques de Jussieu-Paris Rive Gauche (IMJ-PRG) (2006-2010)
_ ISC-PIF (Complex system institute Paris Ile de France) (2007-2013) 
_ Max Planck Institute for Mathematic in the Sciences (MPI-MIS, Leipzig)   (2013-2015)
_ Inserm Unis 1072 - ERC channelomics (2015-2017)
We thank those public and non-profit scientific organization for their support.
""" 

# ************************************************************
# Reading of datafiles (specific of .CSV files or XLS file)
#csvkit manual: http://csvkit.readthedocs.org/en/latest/cli.html
# ************************************************************

import math
from math import *
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter.messagebox import *
from tkinter import filedialog
import pickle
import itertools
from itertools import permutations
from random import shuffle

#################################################################
#################################################################
#################################################################
### THE INPUT  PARAMETERS OF THE PROGRAM  #######################
#################################################################
#################################################################
#################################################################

# string of the path where the .XlSX file is and where the .plk will be saved
directory_path =  os.path.abspath("/home/baudot/LABOMARSEILLE/RESULTAT ANALYSE PCR/ACTUAL RESEARCH/ANALYSE 21RNA MARS/infopathmax6/")
Name_worksheet="DOPA"
Nb_bins=8
# Nb var is the number n used for the computation of n-tuple information
Nb_var=6
# Nb var tot is the total number nbis of variable used for the computation of n-tuple information
# we have necessarily nbis>n
Nb_var_tot=6


### CHOOSE THE KIND OF COMPUTATION YOU WANT TO DO  #######################
# compute_data_only means just 1 computation no shuffle, no change of Nb_bins, no change of Nb_bins
compute_data_only = False
#Shuffle are making a derangement permutation of all the rows values of the data matri, rows by rows, and then compute all informations functions on the randomiszed sample
compute_shuffle = True
nb_of_shuffle=17 # the Shuffle can be made nb_of_shuffle times, with each time computing informations
# compute_different_N will compute all entropie and mutualinfo for different values of Nb_bins=N (GRAINNING N)
compute_different_N = False
Nb_of_N = 10 # the number of bins used for the graining ranges from 2 to Nb_of_N+1 with stepps of 1 bins, ex: if Nb_of_N=5, the program will run for Nb_bins=2,3,4,5,6
# compute_different_m will compute all entropie and mutualinfo for different values of Nb_trials=m (SAMPLE SIZE m)
compute_different_m = False
Nb_of_m = 10 # the number of Nb_trials_max used ranges from Nb_trials_max to int(Nb_trials_max/(Nb_of_m)) with stepps of int(Nb_trials_max/(Nb_of_m))
#Nb_trials =Nb_trials_max- k*int(Nb_trials_max/(Nb_of_m)) 
#ex: if Nb_of_m=10  and Nb_trials_max=111, the program will run for Nb_trials=111,100,89,78,67,56,45,34,23,12

### SAVING OPTIONS  #######################
#display_figure = True
display_figure = False
#matplotlib.use('agg')
#matplotlib.use('svg')
Format_SVG_SAVE_Figure = True
save_results = True
#save_results = False

#################################################################
### Procedure for saving and loading objects ####################
#################################################################

def save_obj(obj, name ):
#    with open('obj/'+ name + '.pkl', 'wb') as f:
    with open(directory_path + '/' + name + '.pkl', 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)

def load_obj(name ):
    with open(directory_path + '/' + name + '.pkl', 'rb') as f:
        return pickle.load(f)

#################################################################
### dialog open file XLS DATA FILE ##############################
#################################################################

root = Tk()

root.filename =  filedialog.askopenfilename(initialdir = directory_path,title = "choose your file",filetypes = (("xls files","*.xlsx"),("all files","*.*")))
root.destroy()

workbook_data = load_workbook(root.filename)

########################################################################
### Tools to manipulate WORKBOOK  ######################################
# https://openpyxl.readthedocs.org/en/2.3.3/index.html #################
########################################################################



print(workbook_data.get_sheet_names())
d = workbook_data[Name_worksheet].cell(row = 2, column = 5)
print(d.value)
c = workbook_data[Name_worksheet]['B1':'B239']
print(c)


###############################################################
### READ AND COPY THE .XLSX FILE SPREADSHIT  ################## 
###     in a NUMPY MATRIX  Matrix_data       ##################
###############################################################


Nb_column=1
for i in range(2,10000):
    if workbook_data[Name_worksheet].cell(row = 1, column = i).value == None:
       Nb_column=i-2
       break
   
print('the number of column (or trials n) is: ', Nb_column)   

Nb_trials=Nb_column


#We put the XLSSheet into a numpy matrix nparray #############
# Variables (RNA) are the lines #
# Trials are the columns  #
# Matrix_data contains the RAW DATA
# Matrix_data contains the  DATA resampled and rescaled


Matrix_data = np.zeros((Nb_var_tot,Nb_trials))
# WE PUT ALL THE 0 value to nan  #
for col in range(Nb_trials):
   for row in range(Nb_var_tot):
      if workbook_data[Name_worksheet].cell(row = (row+2), column = (col+2)).value == 0:
          Matrix_data[row,col] = np.nan
      else:
          Matrix_data[row,col] = workbook_data[Name_worksheet].cell(row = (row+2), column = (col+2)).value

###############################################################
### procedure for resampling and displaying  ################## 
###     probability distributions            ##################
###############################################################

def Resample_data_matrix(Nb_var_tot_temp,Nb_trials_temp,Matrix_data_temp,Nb_bins_temp):


    Matrix_data2_temp = np.zeros((Nb_var_tot_temp,Nb_trials_temp))
    
    Min_matrix = np.nanmin(Matrix_data_temp, axis=1)
    Max_matrix = np.nanmax(Matrix_data_temp, axis=1) 
#ugly correction for a bad rounding of nanmax that exclude the maximal point
    Max_matrix = Max_matrix + ((Max_matrix - Min_matrix)/100000)
    Ampl_matrix = Max_matrix - Min_matrix

# WE RESCALE ALL MATRICES AND SAMPLE IT into  Nb_bins #
    for col in range(Nb_trials_temp):
      for row in range(Nb_var_tot_temp):
          if np.isnan(Matrix_data_temp[row,col]):
#             Matrix_data_temp[row,col] = 0
             Matrix_data2_temp[row,col] = 0
# line to put if you want to save the resampled data in the .xlxs file              
#             workbook_data["DOPA_SAMPLED"].cell(row = (row+2), column = (col+2)).value = 0 
          else:
             if Ampl_matrix[row] !=0 :
                 Matrix_data2_temp[row,col] = int(((Matrix_data_temp[row,col]-Min_matrix[row])*(Nb_bins_temp))/(Ampl_matrix[row]))+1
             else:    
                 Matrix_data2_temp[row,col] = 0
    return (Matrix_data2_temp)
# lines to put if you want to save the resampled data in the .xlxs file   
#             workbook_data["DOPA_SAMPLED"].cell(row = (row+2), column = (col+2)).value = Matrix_data2[row,col]          
#    workbook_data.save(os.path.join(directory_path,"Sampled_DATA.xlsx"))           
         


###############################################################
### AFFICHAGE ET ENREGISTREMENT DES DATA RESAMPLEES  ##########
###                  ET RESCALEES            ##################
###############################################################


   
   
def Display_data_matrix(Matrix_data_input,Matrix_data2_input):   
  
#plt.figure(1)
   plt.matshow(Matrix_data_input)
   if save_results == True: 
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'Matrix_RAW_VALUE'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'Matrix_RAW_VALUE'), format="svg")  

#plt.figure(2)
   plt.matshow(Matrix_data2_input)
   if save_results == True:
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'Matrix_RESCALED_VALUE'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'Matrix_RESCALED_VALUE'), format="svg")  

   plt.figure(3)
   hist, bin_edges = np.histogram(Matrix_data2_input, bins = Nb_bins+10,range=(0,Nb_bins+10))
   plt.bar(bin_edges[:-1], hist, width = 1)
   plt.xlim(min(bin_edges), max(bin_edges))
   if display_figure == True:
      plt.show()
   if save_results == True: 
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'histogram_RESCALED_VALUE'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'histogram_RESCALED_VALUE'), format="svg")  
   
   plt.figure(4)
   hist, bin_edges = np.histogram(Matrix_data_input, bins = 52,range=(0,26))
   plt.bar(bin_edges[:-1], hist, width = 0.5)
   plt.xlim(min(bin_edges), max(bin_edges))
   if display_figure == True:
      plt.show()
   if save_results == True: 
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'histogram_RAW_VALUE'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'histogram_RAW_VALUE'), format="svg")  
   




    
###############################################################
########          Calculus probability               ##########
###              probability distributions            #########
###############################################################

#### To avoid to have to explore all  possible  probability (sparse data)
# we encode probability as dictionanry, each existing probability
# has a key  'x' : x is the index of the probability Pijk...n ex:P111111
# where i is the first variable (row X1) and varies from 1 to nb_bins=N 
# where j is the second variable (row X2) and varies from 1 to nb_bins=N  
# where n=nb_var is the nth and last variable (row Xn) and varies from 1 to nb_bins


def compute_probability(Nb_trials_input,Nb_var_input,Matrix_data2_input):
    probability={}

    for col in range(Nb_trials_input): 
        x=''
        for row in range(Nb_var_input):
           x= x+str(int((Matrix_data2_input[row,col])))
        probability[x]=probability.get(x,0)+1
       

#    print(probability) 
    Nbtot=0  
    for i in probability.items():
        Nbtot=Nbtot+i[1]
#    print('le nombre total doccurence  est:',Nbtot)
    for i,j in probability.items():
           probability[i]=j/float(Nbtot)
#           print('P',i,'=',j/float(Nbtot))
    return (probability)  
      
      
       
###############################################################
########          SOME FUNCTIONS USEFULLS            ##########
###          AT ALL ORDERS On SET OF SUBSETS          #########
###############################################################       
       
       
       
def information(x):
	return -x*math.log(x)/math.log(2) 
 
# Compute probability and  entropies at every order/degrees 

# for i in range(1,Longueurtot+1):   



# Fonction factorielle 

def factorial(x):
     if x < 2:
         return 1
     else:
         return x * factorial(x-1)
   
# Fonction coeficient binomial (nombre de combinaison de k elements dans [1,..,n]) 

def binomial(n,k):
    return factorial(n)/(factorial(k)*factorial(n-k))
    
    

### All deranged permutations of the list row_input.... Derangements are permutatuion without fixed points. 
def Derangements(row_input):
   o=row_input[:]
   while any(x==y for x,y in zip(o,row_input)): shuffle(row_input)
   return row_input   
   
### random permutations of the list row_input
       
def Derangement(row_input):
   np.random.shuffle(row_input.flat)
   return row_input
#############################################################################    
# Fonction decode(x,n,k,combinat)
#--> renvoie la combinatoire 'combinat' de k variables dans n codée par x
# dans combinat
#les combinaisons de k élements dans [1,..,n] sont en bijection avec 
# les entiers x de [0,...,n!/(k!(n-k)!)-1]  
# attention numerotation part de 0     
#############################################################################
    
def decode(x,n,k,combinat):
    if x<0 or n<=0 or k<=0:
        return 
    b= binomial(n-1,k-1)
    if x<b:
        decode(x,n-1,k-1,combinat)
        combinat.append(n)
    else:
        decode(x-b,n-1,k,combinat) 

        
#############################################################################
# Fonction decode_all(x,n,k,combinat)
# take x in entry an integer in [0,(2^n)-1] , the code of a given k-tuple     
#--> returns the k-tuple (combinat: a k-tuple of the form [4,6,8] for a 3-tuple) and the degree k, associated to the code x 
# x varies from 0 to (2^n)-1, the n first x code for 1 in n (degree 1 1-tuple)
# the following code for 2 in n (degree 2 2-tuple)
# etc... until x=(2^n)-1  that code for n in n
#The k-tuples in [1,..,n] are in bijection with
# the integers  x of binomial  [0,...,n!/(k!(n-k)!)-1]  
# attention! numbering starts from 0   
#############################################################################

def decode_all(x,n,order,combinat):
    sumtot=n
    order=1
    Code_order=x
    while x>=sumtot:
       order=order+1
       sumtot=sumtot+binomial(n,order) 
       Code_order=Code_order-binomial(n,order-1)
    k=order
    decode(Code_order,n,k,combinat)  
     


###############################################################
###############################################################
###           SHUFFLE  OF DATA MATRIX                ##########
###############################################################
###############################################################
# Shuffling consists in applying a derangement permutation (permutation without fixed point) of each rows values rows by rows
# nobody knows what is randomness so far in mathematic but if it is a permutation it is certainly not the identity permutation
# CORRECTION: derangement are too long to compute for large Nb_trials we replace them by simple random shuffling 
def Shuffle_matrix_data(Matrix_data2_input):
    Matrix_data2_output = np.zeros((Nb_var_tot,Nb_trials))
    for row in range(Nb_var_tot):
#        print('matrix input row',row, ' is  ', Matrix_data2_input[row,0:Nb_trials])
#        print('matrix output row')
        Matrix_data2_output[row,0:Nb_trials]= Derangement(Matrix_data2_input[row,0:Nb_trials]) 
#        print('matrix output row',row, ' is  ',Matrix_data2_output[row,0:Nb_trials])
    return(Matrix_data2_output)               
 



###############################################################
###  COMBINATORIAL SUB MATRIX SELECTION             ##########
###                                          ##################
###############################################################
import sys

def compute_sub_matrix_aux(x_input,Nb_var_tot_input,Nb_var_input,Matrix_data2_input,combinat):
    decode(x_input,Nb_var_tot_input,Nb_var_input,combinat)  
         
    
def compute_sub_matrix(x_input,Nb_var_tot_input,Nb_var_input,Nb_trials_input,Matrix_data2_input,combinat_input):
      
#    compute_sub_matrix_aux(x_input,Nb_var_tot,Nb_var,Matrix_data2,ntuple1)
#    decode(x_input,Nb_var_tot_input,Nb_var_input,combinat_input)
#    print('matrix depard', Matrix_data2_input)
#    if len(combinat_input) ==   Nb_var_input :      
#            print('cobinat is', combinat_input)
            global Matrix_data_temp
            Matrix_data_temp = np.zeros((Nb_var_input,Nb_trials_input))
            for num_row in range(Nb_var_input):
#                print('num_row', num_row)
                val_row = combinat_input[num_row]-1
                for col in range(Nb_trials_input):
                   #col = col-2
#                   print('col', col)
#                   print('lign', combinat_input)
#                   print('val matrix raw', Matrix_data[val_row,col])
#                   print('val matrix renorm', Matrix_data2[val_row,col])
 #                  print('cobinat is', combinat_input)    
                   Matrix_data_temp[num_row,col] = Matrix_data2_input[val_row,col]
                   #if col > 5 :
                      #sys.exit()
#            print('cobinat is', combinat_input)     
#            print('matrix depard', Matrix_data2_input)
#            print('matrix data temp', Matrix_data_temp)
            return (Matrix_data_temp)        

##################################################################################
###############    COMPUTE ENTROPY                         #######################
##################################################################################   

# Nentropy and Ninfomut are dictionaries (x,y) with x a list of kind (1,2,5) and y a value in bit 

#ntuple1_input is usually unusefull: it is involved only in case of computation of a k-sublattice of n (k<n) (compute submatrix)
def compute_entropy(Nb_var_input,probability_input,ntuple1_input):
    
    Nentropie_input={} 
    sys.getrecursionlimit()
    ntuple=[]        
    for x in range(0,(2**Nb_var_input)-1): 
#        we scann all the k-tuple in n        
        ntuple=[]
        orderInf=0
        decode_all(x,Nb_var_input,orderInf,ntuple)   
#        ntuple contains the  k-tuple of variable ex: ntuple=[4,6,8] and orderInf=3
#        print('code dans boucle:',ntuple)
        tuple_code=()        
        probability2={}
        for z in range(0,len(ntuple)):
            concat=()
#            concat=(ntuple1_input[z],)
            concat=(ntuple1_input[ntuple[z]-1],)
            tuple_code=tuple_code+concat
#            print('code dans boucle:',tuple_code)
#        print('code local:',tuple_code)
#            print('code global:',ntuple1_input)
 #       tuple_code = change_code(tuple_code,ntuple1_input)
        for x,y in probability_input.items():
#codeproba is the code of an atomic probability ex: for ntupe [1, 3]  codeproba=102 means 1st value of X1, no value for X2, and 2nd value for X3 
# for ntupe [1,2,3]  codeproba=123 means 1st value of X1, 2nd value for X2, and 3rd value for X3 
        
            Codeproba=''
            length=0
            for w in range(1,Nb_var_input+1):
#                print 'w='
#                print w
#                print length
#                print len(ntuple)
#                print ntuple[length]            
                if ntuple[length]!=w:
                    Codeproba=Codeproba+'0'                
                else: 
                    Codeproba=Codeproba+x[ntuple[length]-1:ntuple[length]]
                    if length<(len(ntuple)-1):
                        length=length+1      
            probability2[Codeproba]=probability2.get(Codeproba,0)+probability_input.get(x,0)
#            print('probability(',Codeproba,')=',probability2[Codeproba])  
        Nentropie_input[tuple_code]=0
# to change: the program computes too many times the entropy:  m*2**n instead of          
        for x,y in probability2.items():
#             if Nentropie[tuple_code]==0:
#                print('ta mere en short celui la y est pas:')
                Nentropie_input[tuple_code]=Nentropie_input.get(tuple_code,0)+information(probability2[x])   
        probability2={}  
        probability2.clear()
        del(probability2)
    return(Nentropie_input)
    Nentropie_input.clear()
    del(Nentropie_input)

##################################################################################
###############    SUMM  ENTROPY per order                 #######################
##################################################################################   
# somme des entropies à chaque ordre

def compute_sum_entropy_order(Nentropie_input):
    entropy_sum_order={} 
    for x,y in Nentropie_input.items():
        entropy_sum_order[len(x)]=entropy_sum_order.get(len(x),0)+Nentropie_input[x]
    return (entropy_sum_order)
        
##################################################################################
###############    COMPUTE MUTUAL INFORMATION              #######################
##################################################################################                
# Calcul des informations mutuelles à tous les ordres




    
#############################################################################
# Function binomial_subgroups COMBINAT Gives all binomial k subgroup of a group 
#############################################################################
     
    
    
def compute_Ninfomut(Nentropie_input):    
#    Ninfomut={}   
    for x,y in Nentropie_input.items():
        for k in range(1, len(x)+1):
           for subset in itertools.combinations(x, k):
#              print(subset)  
              Ninfomut[x]=Ninfomut.get(x,0)+ ((-1)**(len(subset)+1))*Nentropie_input[subset] 
    return (Ninfomut)  
    
    
##################################################################################
############### RANKING of the n higher Mutual information #######################
##################################################################################   

from operator import itemgetter
from collections import OrderedDict

# regular unsorted dictionary
# dictionary sorted by value
# Ninfo_per_order_ordered[x]  is ordered firstby each order x of information
# and second ordered by value of information 
def compute_Ninfo_per_order(Nb_var_input,Ninfo_input): 
    info_per_order=[]
    Ninfo_per_order_ordered=[]
    for x in range(Nb_var_input+1):
        info_dicoperoder={} 
        info_per_order.append(info_dicoperoder)
        Ninfo_per_order_ordered.append(info_dicoperoder)
    
    for x,y in Ninfo_input.items():
#       print('lenghtx=',len(x))
        info_per_order[len(x)][x]=Ninfo_input[x]

    for x in range(Nb_var_input+1):
        Ninfo_per_order_ordered[x]=OrderedDict(sorted(info_per_order[x].items(), key=lambda t: t[1]))
#        print('les infos mutuelles à ', x ,'sont:')
#        print(Ninfo_per_order_ordered[x])
    return (Ninfo_per_order_ordered,info_per_order)
#Ninfomut_ordered=OrderedDict(sorted(Ninfomut.items(), key=lambda t: t[1]))



##################################################################################
###############    Sum of mutual info per order            #######################
################################################################################## 

# somme des information à chaque ordre

def compute_infomut_sum_order(Ninfomut_input): 
    infomut_sum_order={} 
    for x,y in Ninfomut_input.items():
        infomut_sum_order[len(x)]=infomut_sum_order.get(len(x),0)+Ninfomut_input[x]

#    print('les info mutuelles sommées par ordre sont:')
#    print(infomut_sum_order)     

    infototbis=0.00    
    for x,y in infomut_sum_order.items():
        infototbis=infototbis+((-1)**(x+1))*infomut_sum_order.get(x,0)

#    print('l entropie calculée par les sommes est de:')
#    print(infototbis)   
    return (infomut_sum_order,infototbis)

# somme des valeur absolues des information à chaque ordre
def compute_infomut_sum_order_abs(Ninfomut_input): 
    infomut_sum_order_abs={} 
    for x,y in Ninfomut_input.items():
        infomut_sum_order_abs[len(x)]=infomut_sum_order_abs.get(len(x),0)+math.fabs(Ninfomut_input[x])

#    print('les valeur abs(info mutuelles) sommées par ordre sont:')
#    print(infomut_sum_order_abs)     

    infotot_absbis=0.00    
    for x,y in infomut_sum_order_abs.items():
        infotot_absbis=infotot_absbis+infomut_sum_order_abs.get(x,0)   

#    print('la somme des val abs de toutes le infomut est:')
#    print(infotot_absbis)         
    return (infomut_sum_order_abs,infotot_absbis)

#########################################################################
#########################################################################
######################## CENTRAL PROGRAM   ##############################
#########################################################################
#########################################################################

#########################################################################
##############  PROGRAM  COMPUTE ON DATA ONLY    ########################
########    NO SHUFFLE, 1bin value, 1 nb_trial value   ##################
#########################################################################

if compute_data_only == True:       

# Nentropy and Ninfomut are dictionaries (x,y) with x a list of kind (1,2,5) and y a value in bit 
   Nentropie={}  
   Ninfomut={}
   ntuple1=[]  
   Matrix_data2 = Resample_data_matrix(Nb_var_tot,Nb_trials,Matrix_data,Nb_bins)
   Display_data_matrix(Matrix_data,Matrix_data2)  
   decode(0,Nb_var_tot,Nb_var,ntuple1)  

   print('COMPUTE PROBA step 0')
   probability = compute_probability(Nb_trials,Nb_var,Matrix_data2)
   if save_results == True: 
       save_obj(Matrix_data2,'Matrix_data2')
   print(Matrix_data2)    

   print('COMPUTE ENTROPIES step 1')
   Nentropie = compute_entropy(Nb_var,probability,ntuple1)
   if save_results == True: 
       save_obj(Nentropie,'ENTROPY')   
    
   print('COMPUTE ENTROPIES per order step 2')   
   (Nentropy_per_order_ordered,entropy_per_order) = compute_Ninfo_per_order(Nb_var,Nentropie)
   if save_results == True: 
       save_obj(Nentropy_per_order_ordered,'ENTROPY_ORDERED')

   print('COMPUTE INFOMUT step 3')      
   Ninfomut = compute_Ninfomut(Nentropie)
   if save_results == True: 
       save_obj(Ninfomut,'INFOMUT')
    
   print('COMPUTE SUM_ENTROPIES step 4')
   entropy_sum_order = compute_sum_entropy_order(Nentropie)
   if save_results == True: 
       save_obj(entropy_sum_order,'ENTROPY_SUM')  
    
   print('COMPUTE INFOMUT per order step 5')
   (Ninfomut_per_order_ordered,infomut_per_order) = compute_Ninfo_per_order(Nb_var,Ninfomut)
   if save_results == True:
       save_obj(Ninfomut_per_order_ordered,'INFOMUT_ORDERED')
       save_obj(infomut_per_order,'INFOMUT_ORDEREDList')
    
   print('COMPUTE SUM_INFOMUT step 6')
   (infomut_sum_order,infototbis) = compute_infomut_sum_order(Ninfomut) 
   if save_results == True:
       save_obj(infomut_sum_order,'INFOMUT_SUM')
   print('END COMPUTE step 7')

#########################################################################
########   PROGRAM  COMPUTE ON DATA several SHuffles     ################
#########################################################################

if compute_shuffle == True:
   for k in range(nb_of_shuffle): 
       Matrix_data2 = Resample_data_matrix(Nb_var_tot,Nb_trials,Matrix_data,Nb_bins)
       Matrix_data2_copy = np.copy(Matrix_data2)
       print('*************************')
       print('k=',k)
#       print('Matrix DATA  ',Matrix_data2)
       Matrix_data2_copy = Shuffle_matrix_data(Matrix_data2_copy)    
#       print('Matrix  DATA SHUFFLED',Matrix_data2_copy)
       if display_figure  == True: 
           Display_data_matrix(Matrix_data,Matrix_data2_copy) 
       
       Nentropie={}  
       Ninfomut={}
       ntuple1=[]  
       decode(0,Nb_var_tot,Nb_var,ntuple1)  
   
       print('COMPUTE PROBA step 0')
       probability = compute_probability(Nb_trials,Nb_var,Matrix_data2_copy)
       if save_results == True: 
           save_obj(Matrix_data2_copy,'Matrix_data2'+str(k))

       print('COMPUTE ENTROPIES step 1')
       Nentropie = compute_entropy(Nb_var,probability,ntuple1)
       if save_results == True: 
           save_obj(Nentropie,'ENTROPY'+str(k))   
    
       print('COMPUTE ENTROPIES per order step 2')   
       (Nentropy_per_order_ordered,entropy_per_order) = compute_Ninfo_per_order(Nb_var,Nentropie)
       if save_results == True: 
           save_obj(Nentropy_per_order_ordered,'ENTROPY_ORDERED'+str(k))

       print('COMPUTE INFOMUT step 3')      
       Ninfomut = compute_Ninfomut(Nentropie)
       if save_results == True: 
           save_obj(Ninfomut,'INFOMUT'+str(k))
    
       print('COMPUTE SUM_ENTROPIES step 4')
       entropy_sum_order = compute_sum_entropy_order(Nentropie)
       if save_results == True: 
           save_obj(entropy_sum_order,'ENTROPY_SUM'+str(k))  
    
       print('COMPUTE INFOMUT per order step 5')
       (Ninfomut_per_order_ordered,infomut_per_order) = compute_Ninfo_per_order(Nb_var,Ninfomut)
       if save_results == True:
           save_obj(Ninfomut_per_order_ordered,'INFOMUT_ORDERED'+str(k))
           save_obj(infomut_per_order,'INFOMUT_ORDEREDList'+str(k))
    
       print('COMPUTE SUM_INFOMUT step 6')
       (infomut_sum_order,infototbis) = compute_infomut_sum_order(Ninfomut) 
       if save_results == True:
           save_obj(infomut_sum_order,'INFOMUT_SUM'+str(k))
       print('END COMPUTE step 7')
       
#########################################################################
########   PROGRAM  COMPUTE ON DATA several Bin values   ################
#########################################################################
       
if compute_different_N == True:
   for k in range(Nb_of_N): 
       Nb_bins=(k+2)
       Matrix_data2 = Resample_data_matrix(Nb_var_tot,Nb_trials,Matrix_data,Nb_bins)
       Matrix_data2_copy = np.copy(Matrix_data2)
       print('*************************')
       print('Nb_bins=',Nb_bins)
#       print('Matrix DATA  ',Matrix_data2) 
#       print('Matrix  DATA SHUFFLED',Matrix_data2_copy)
       if display_figure  == True: 
           Display_data_matrix(Matrix_data,Matrix_data2_copy)  
       
       Nentropie={}  
       Ninfomut={}
       ntuple1=[]  
       decode(0,Nb_var_tot,Nb_var,ntuple1)  
   
       print('COMPUTE PROBA step 0')
       probability = compute_probability(Nb_trials,Nb_var,Matrix_data2_copy)
       if save_results == True: 
           save_obj(Matrix_data2_copy,'Matrix_data2'+str(k))

       print('COMPUTE ENTROPIES step 1')
       Nentropie = compute_entropy(Nb_var,probability,ntuple1)
       if save_results == True: 
           save_obj(Nentropie,'ENTROPY'+str(k))   
    
       print('COMPUTE ENTROPIES per order step 2')   
       (Nentropy_per_order_ordered,entropy_per_order) = compute_Ninfo_per_order(Nb_var,Nentropie)
       if save_results == True: 
           save_obj(Nentropy_per_order_ordered,'ENTROPY_ORDERED'+str(k))

       print('COMPUTE INFOMUT step 3')      
       Ninfomut = compute_Ninfomut(Nentropie)
       if save_results == True: 
           save_obj(Ninfomut,'INFOMUT'+str(k))
    
       print('COMPUTE SUM_ENTROPIES step 4')
       entropy_sum_order = compute_sum_entropy_order(Nentropie)
       if save_results == True: 
           save_obj(entropy_sum_order,'ENTROPY_SUM'+str(k))  
    
       print('COMPUTE INFOMUT per order step 5')
       (Ninfomut_per_order_ordered,infomut_per_order) = compute_Ninfo_per_order(Nb_var,Ninfomut)
       if save_results == True:
           save_obj(Ninfomut_per_order_ordered,'INFOMUT_ORDERED'+str(k))
           save_obj(infomut_per_order,'INFOMUT_ORDEREDList'+str(k))
    
       print('COMPUTE SUM_INFOMUT step 6')
       (infomut_sum_order,infototbis) = compute_infomut_sum_order(Ninfomut) 
       if save_results == True:
           save_obj(infomut_sum_order,'INFOMUT_SUM'+str(k))
       print('END COMPUTE step 7')

#########################################################################
#######  PROGRAM  COMPUTE ON DATA several nb_trials value   #############
#########################################################################

if compute_different_m == True:
   Nb_trials_max=Nb_trials  
   for k in range(Nb_of_m): 
       Nb_trials =Nb_trials_max- k*int(Nb_trials_max/(Nb_of_m))
       Matrix_data2 = Resample_data_matrix(Nb_var_tot,Nb_trials,Matrix_data,Nb_bins)
       Matrix_data2_copy = np.copy(Matrix_data2)
       print('*************************')
       print('Nb_trials=',Nb_trials)
#       print('Matrix DATA  ',Matrix_data2) 
#       print('Matrix  DATA SHUFFLED',Matrix_data2_copy)
       if display_figure  == True: 
           Display_data_matrix(Matrix_data,Matrix_data2_copy) 
       
       Nentropie={}  
       Ninfomut={}
       ntuple1=[]  
       decode(0,Nb_var_tot,Nb_var,ntuple1)  
   
       print('COMPUTE PROBA step 0')
       probability = compute_probability(Nb_trials,Nb_var,Matrix_data2_copy)
       if save_results == True: 
           save_obj(Matrix_data2_copy,'Matrix_data2'+str(k))

       print('COMPUTE ENTROPIES step 1')
       Nentropie = compute_entropy(Nb_var,probability,ntuple1)
       if save_results == True: 
           save_obj(Nentropie,'ENTROPY'+str(k))   
    
       print('COMPUTE ENTROPIES per order step 2')   
       (Nentropy_per_order_ordered,entropy_per_order) = compute_Ninfo_per_order(Nb_var,Nentropie)
       if save_results == True: 
           save_obj(Nentropy_per_order_ordered,'ENTROPY_ORDERED'+str(k))

       print('COMPUTE INFOMUT step 3')      
       Ninfomut = compute_Ninfomut(Nentropie)
       if save_results == True: 
           save_obj(Ninfomut,'INFOMUT'+str(k))
    
       print('COMPUTE SUM_ENTROPIES step 4')
       entropy_sum_order = compute_sum_entropy_order(Nentropie)
       if save_results == True: 
           save_obj(entropy_sum_order,'ENTROPY_SUM'+str(k))  
    
       print('COMPUTE INFOMUT per order step 5')
       (Ninfomut_per_order_ordered,infomut_per_order) = compute_Ninfo_per_order(Nb_var,Ninfomut)
       if save_results == True:
           save_obj(Ninfomut_per_order_ordered,'INFOMUT_ORDERED'+str(k))
           save_obj(infomut_per_order,'INFOMUT_ORDEREDList'+str(k))
    
       print('COMPUTE SUM_INFOMUT step 6')
       (infomut_sum_order,infototbis) = compute_infomut_sum_order(Ninfomut) 
       if save_results == True:
           save_obj(infomut_sum_order,'INFOMUT_SUM'+str(k))
       print('END COMPUTE step 7')       


print('boulot fini')
