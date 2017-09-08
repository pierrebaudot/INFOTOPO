# -*- coding: utf-8 -*-
"""
Created on Thu Nov  5 16:06:46 2015
Copyright (2007-2017) Pierre Baudot 
The terms of the licence GNU GPL is reproduced in the file "COPYING" of INFOTOPO distribution.

This file is part of INFOTOPO.

  INFOTOPO is free software: you can redistribute it and/or modify
  it under the terms of the GNU General Public License as published by
  the Free Software Foundation, either version 3 of the License, or
  (at your option) any later version.

  INFOTOPO is distributed in the hope that it will be useful,
  but WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU General Public License for more details.

  You should have received a copy of the GNU General Public License
  along with INFOTOPO.  If not, see <http://www.gnu.org/licenses/>. 
  
@author: baudot & al. 
GENERAL PHILOSOPHY AND AUTORS REQUEST: 
Giving the authorization of free use of this program also gives you the responsability 
of its use: we trust you. Every tool can be used to make things better for all, but also worse.
We ask any users of the present program to use and developp it such that it profits to the 
sustainable development of all, societies and environment and avoid development
in selfish interest, conflictual involvement, or individual freedom restrictions or survey. 
As a Human in the world, please be conscious of what you do, of its consequence on others, and try to make it better.
This tools comes from mathematic and pertain to it, mathematic do not pertain to anybody in particular 
and according to its original developments promotes harmony and sustainable developements on short 
and long time and space scales, please respect this ancestral tradition, old as humanity, and develop it in this respect. 
More information on the practical use and mathematical framework and bibliographic citations:
http://www.biorxiv.org/content/early/2017/07/26/168740 
and README file. 
For any requests, questions, improvements, developments (etc.) contact pierre.baudot [at] gmail.com
INFOTOPO has been developped since 2007 thanks to grants-fundings of:
ISC-PIF (Complex system institute Paris Ile de France) (2007-2013) 
Max Planck Institute for Mathematic in the Sciences (MPI-MIS, Leipzig)   (2013-2015)
Inserm Unis 1072 - ERC channelomics (2015-2017)
We thank those public and non-profit scientific organization for their support.
"""

# ************************************************************
# Reading of datafiles (specific of .CSV files or XLS file)
#csvkit manual: http://csvkit.readthedocs.org/en/latest/cli.html
# ************************************************************

import math
from math import *
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.colors import LogNorm 
from matplotlib.colors import SymLogNorm
import numpy as np
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter.messagebox import *
from tkinter import filedialog
import pickle
import itertools
from operator import itemgetter
from collections import OrderedDict



#################################################################
### THE INPUT  PARAMETERS OF THE PROGRAM  #######################
#################################################################

directory_path =  os.path.abspath("/home/baudot/LABOMARSEILLE/RESULTAT ANALYSE PCR/ACTUAL RESEARCH/ANALYSE 21RNA MARS/ANALYSE_nbtrial/N12/")
Name_worksheet="DOPA"
Nb_bins=8
# Nb var is the number n used for the computation of n-tuple information
Nb_var=21
# Nb var tot is the total number m of variable used for the computation of n-tuple information
# we have necessarily m>n
Nb_var_tot=21

load_results_ENTROPY = True
load_results_ENTROPY_ORDERED = True
load_results_INFOMUT = True
load_results_ENTROPY_SUM = True
load_results_INFOMUT_ORDERED = True
load_results_INFOMUT_SUM = True
    
SHOW_results_ENTROPY = False
SHOW_results_INFOMUT = False
SHOW_results_COND_INFOMUT = False
SHOW_results_ENTROPY_HISTO = True
SHOW_results_INFOMUT_HISTO = False
SHOW_results_INFOMUT_path = False
SHOW_results_SCAFOLD =  False

SHOW_HARD_DISPLAY= False
SHOW_HARD_DISPLAY_medium= False

#display_figure = True
display_figure = False
#matplotlib.use('agg')
#matplotlib.use('svg')
Format_SVG_SAVE_Figure = False
save_results = False
#save_results = False

#################################################################
### Procedure for saving and loading objects ####################
#################################################################

def save_obj(obj, name ):
#    with open('obj/'+ name + '.pkl', 'wb') as f:
    with open(directory_path + '/' + name + '.pkl', 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)

def load_obj(name ):
    with open(directory_path + '/' + name + '.pkl', 'rb') as f:       
        return pickle.load(f)

#################################################################
### dialog open file XLS DATA FILE ##############################
#################################################################

root = Tk()

root.filename =  filedialog.askopenfilename(initialdir = directory_path,title = "choose your file",filetypes = (("xls files","*.xlsx"),("all files","*.*")))
root.destroy()

workbook_data = load_workbook(root.filename)

########################################################################
### Tools to manipulate WORKBOOK  ######################################
# https://openpyxl.readthedocs.org/en/2.3.3/index.html #################
########################################################################


#Name_worksheet="NONDOPA_MONICA_SIMONE"

print(workbook_data.get_sheet_names())
d = workbook_data[Name_worksheet].cell(row = 2, column = 5)
print(d.value)
c = workbook_data[Name_worksheet]['B1':'B239']
print(c)


###############################################################
### procedure for resampling and displaying  ################## 
###     probability distributions            ##################
###############################################################


Nb_column=1
for i in range(2,10000):
    if workbook_data[Name_worksheet].cell(row = 1, column = i).value == None:
       Nb_column=i-2
       break
   
print('the number of column (or trials n) is: ', Nb_column)   

Nb_trials=Nb_column


#We put the XLSSheet into a numpy matrix nparray #############
# Variables (RNA) are the lines #
# Trials are the columns  #
# Matrix_data contains the RAW DATA
# Matrix_data contains the  DATA resampled and rescaled

Matrix_data = np.zeros((Nb_var_tot,Nb_trials))
Matrix_data2 = np.zeros((Nb_var_tot,Nb_trials))

# WE PUT ALL THE 0 value to nan  #
for col in range(Nb_trials):
  for row in range(Nb_var_tot):
      if workbook_data[Name_worksheet].cell(row = (row+2), column = (col+2)).value == 0:
         Matrix_data[row,col] = np.nan
      else:
         Matrix_data[row,col] = workbook_data[Name_worksheet].cell(row = (row+2), column = (col+2)).value



Min_matrix = np.nanmin(Matrix_data, axis=1)
Max_matrix = np.nanmax(Matrix_data, axis=1) 
Ampl_matrix = Max_matrix - Min_matrix

# WE RESCALE ALL MATRICES AND SAMPLE IT into  Nb_bins #
for col in range(Nb_trials):
  for row in range(Nb_var_tot):
      if np.isnan(Matrix_data[row,col]):
         Matrix_data[row,col] = 0
         Matrix_data2[row,col] = 0
         workbook_data["DOPA_SAMPLED"].cell(row = (row+2), column = (col+2)).value = 0 
      else:
         if Ampl_matrix[row] !=0 :
             Matrix_data2[row,col] = int(((Matrix_data[row,col]-Min_matrix[row])*(Nb_bins-2))/(Ampl_matrix[row]))+1
         else:    
             Matrix_data2[row,col] = 0
         workbook_data["DOPA_SAMPLED"].cell(row = (row+2), column = (col+2)).value = Matrix_data2[row,col]          
        
         





# Fonction factorielle 

def factorial(x):
     if x < 2:
         return 1
     else:
         return x * factorial(x-1)
         
# Fonction coeficient binomial (nombre de combinaison de k elements dans [1,..,n]) 
   
def binomial(n,k):
    return factorial(n)/(factorial(k)*factorial(n-k))
    
    
    
def compute_Ninfomut_cond(Nentropie_input):    
#    Ninfomut={}   
    for x,y in Nentropie_input.items():
        for k in range(1, len(x)+1):
           for subset in itertools.combinations(x, k):
#              print(subset)  
              Ninfomut[x]=Ninfomut.get(x,0)+ ((-1)**(len(subset)+1))*Nentropie_input[subset] 
    return (Ninfomut)     
            

###############################################################
### AFFICHAGE ET ENREGISTREMENT DES DATA RESAMPLEES  ##########
###                  ET RESCALEES            ##################
###############################################################

num_fig=0

workbook_data.save(os.path.join(directory_path,"Sampled_DATA.xlsx"))     
#plt.figure(1)
num_fig=num_fig+1
plt.matshow(Matrix_data, cmap='autumn')
if save_results == True: 
   if Format_SVG_SAVE_Figure == False: 
      plt.savefig(os.path.join(directory_path,'Matrix_RAW_VALUE'), format="png")
   if Format_SVG_SAVE_Figure == True:       
      plt.savefig(os.path.join(directory_path,'Matrix_RAW_VALUE'), format="svg")  

#plt.figure(2)
plt.matshow(Matrix_data2, cmap='autumn')
num_fig=num_fig+1
if save_results == True:
   if Format_SVG_SAVE_Figure == False: 
      plt.savefig(os.path.join(directory_path,'Matrix_RESCALED_VALUE'), format="png")
   if Format_SVG_SAVE_Figure == True:       
      plt.savefig(os.path.join(directory_path,'Matrix_RESCALED_VALUE'), format="svg")  

num_fig=num_fig+1
plt.figure(num_fig)
hist, bin_edges = np.histogram(Matrix_data2, bins = Nb_bins,range=(0,Nb_bins))
plt.bar(bin_edges[:-1], hist, width = 1)
plt.xlim(min(bin_edges), max(bin_edges))
if display_figure == True:
   plt.show()
if save_results == True: 
   if Format_SVG_SAVE_Figure == False: 
      plt.savefig(os.path.join(directory_path,'histogram_RESCALED_VALUE'), format="png")
   if Format_SVG_SAVE_Figure == True:       
      plt.savefig(os.path.join(directory_path,'histogram_RESCALED_VALUE'), format="svg")  

num_fig=num_fig+1   
plt.figure(num_fig)
hist, bin_edges = np.histogram(Matrix_data, bins = 52,range=(0,26))
plt.bar(bin_edges[:-1], hist, width = 0.5)
plt.xlim(min(bin_edges), max(bin_edges))
if display_figure == True:
   plt.show()
if save_results == True: 
   if Format_SVG_SAVE_Figure == False: 
      plt.savefig(os.path.join(directory_path,'histogram_RAW_VALUE'), format="png")
   if Format_SVG_SAVE_Figure == True:       
      plt.savefig(os.path.join(directory_path,'histogram_RAW_VALUE'), format="svg")  
   


###############################################################
###      LOAD SAVED FILE  ##########
###                  ET RESCALEES            ##################
###############################################################




if load_results_ENTROPY == True: 
    name_object= 'ENTROPY'
    Nentropie = load_obj(name_object)  
    

if load_results_ENTROPY_ORDERED == True: 
    name_object= 'ENTROPY_ORDERED'
    Nentropy_per_order_ordered = load_obj(name_object) 
  

if load_results_INFOMUT == True: 
   name_object= 'INFOMUT'
   Ninfomut = load_obj(name_object) 

if load_results_ENTROPY_SUM == True:  
    name_object= 'ENTROPY_SUM'
    entropy_sum_order = load_obj(name_object)
    

if load_results_INFOMUT_ORDERED == True: 
    name_object= 'INFOMUT_ORDERED'
    Ninfomut_per_order_ordered = load_obj(name_object)
    name_object= 'INFOMUT_ORDEREDList'
    infomut_per_order = load_obj(name_object)

if load_results_INFOMUT_SUM == True: 
    name_object= 'INFOMUT_SUM'
    infomut_sum_order = load_obj(name_object)

#(infomut_sum_order_abs,infotot_absbis) = compute_infomut_sum_order_abs(Ninfomut)


#for x in range(0,(2**Nb_var_tot)-1):     
#    print('les infos mutuelles à ', x ,'sont:')
#    print(Ninfomut_per_order_ordered[x])    
#print('la somme des val abs de toutes le infomut est:')
#print(infotot_absbis)         
#print('les valeur abs(info mutuelles) sommées par ordre sont:')
#print(infomut_sum_order_abs)

print('les info mutuelles sommées par ordre sont:')
print(infomut_sum_order) 

#########################################################################
#########################################################################
######      GRAPHICS AND DISPLAY           ##############################
#########################################################################
#########################################################################
# Visualisation Graphique Distribution des valeurs

def save_FIGURE(num_fig2, name, Format_SVG_SAVE_Fig ):
#    with open('obj/'+ name + '.pkl', 'wb') as f:
#   mng = plt.get_current_fig_manager(num_fig2)                                         
#   mng.window.showMaximized(num_fig2)
   plt.tight_layout(num_fig2)   
   if Format_SVG_SAVE_Fig == False: 
         plt.savefig(os.path.join(directory_path,name), format="png")
   if Format_SVG_SAVE_Fig == True:       
         plt.savefig(os.path.join(directory_path,name), format="svg") 
    
 


# Joint-ENtropy Values 

if SHOW_results_ENTROPY == True: 
   
   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   if SHOW_HARD_DISPLAY_medium== True:
      for x,y in Nentropie.items():
         plt.plot(len(x), y, 'ro')
         if y>maxordonnee:
             maxordonnee=y
         if y<minordonnee:
             minordonnee=y
#      xp = np.linspace(minordonnee-((maxordonnee-minordonnee)*0.03), maxordonnee+((maxordonnee-minordonnee)*0.03), 100)
#      plt.plot(len(x), y, 'ro',xp, polyn_fit(xp), '-')
       
      plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
      plt.ylabel('(Bits)')
      plt.title('Entropy')
      plt.grid(True) 
#      if display_figure == True: 
#         plt.show(num_fig)
#      else: plt.close(num_fig)        
      if save_results == True: 
          save_FIGURE(num_fig, 'DIAGRAM_ENTROPY', Format_SVG_SAVE_Figure )
   
   
# Sum of Joint-ENtropy Values per degrees n

   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   for x,y in entropy_sum_order.items():
      plt.plot(x, y, 'ro',x, y, 'b-')
      if y>maxordonnee:
          maxordonnee=y
      if y<minordonnee:
          minordonnee=y
   plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(bits)')
   plt.title('Sum entropy per order')
   plt.grid(True)
#   if display_figure == True: 
#      plt.show(num_fig)
#   else: plt.close(num_fig)        
   if save_results == True: 
       save_FIGURE(num_fig, 'DIAGRAM_ENTROPY', Format_SVG_SAVE_Figure )

# Average Entropy rate 

   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   for x,y in entropy_sum_order.items():
      rate=y/x
      plt.plot(x, rate, 'ro',x, y, 'b-')
      if rate>maxordonnee:
          maxordonnee=rate
      if rate<minordonnee:
          minordonnee=rate 
   plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(Bits/symbols)')
   plt.title('Average Entropy Rate')
   plt.grid(True)
#   if display_figure == True: 
#      plt.show(num_fig)
#   else: plt.close(num_fig)        
   if save_results == True: 
       save_FIGURE(num_fig, 'Average Entropy Rate', Format_SVG_SAVE_Figure )

# Average entropy normalised by Binomial coefficients (MEASURE OF CONTRIBUTION PER ORDER)

   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   average_entropy=[]
   for x,y in entropy_sum_order.items():
      rate=y/binomial(Nb_var,x)
      plt.plot(x, rate,  linestyle='--', marker='o', color='b')
      average_entropy.append(rate)
#      plt.plot(x, rate, 'ro',x, y, 'b-')
      if rate>maxordonnee:
          maxordonnee=rate
      if rate<minordonnee:
          minordonnee=rate 
   plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(Bits/symbols)')
   plt.title('Average entropy normalised by Binomial coef')
   plt.grid(True)
   print('Average entropy normalised by Binomial coef')
   print(average_entropy)
#   if display_figure == True: 
#      plt.show(num_fig)
#   else: plt.close(num_fig)        
   if save_results == True: 
       save_FIGURE(num_fig, 'Average entropy normalised by Binomial coef', Format_SVG_SAVE_Figure )




# Entropy rate 

#plt.subplot(323)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Nentropie.items():
#   rate=y/len(x)
#   plt.plot(len(x), rate, 'ro')
#   if rate>maxordonnee:
#       maxordonnee=rate
#   if rate<minordonnee:
#       minordonnee=rate 
#plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(Bits/symbols)')
#plt.title('Entropy Rate')
#plt.grid(True)

# Sum of Joint-entropy values per degrees n


# MULTI-INFO  
#plt.subplot(324)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Nentropie.items():
#    SumH1=0.00
##    print 'le code de sequence encours est:', x
#    for v in range(0,len(x)):
#        Mierda=(x[v],)
#        print 'les composantes marg :', x[v]
 #       print 'lentropy de la marg :', Nentropie[Mierda]
#        SumH1=SumH1+Nentropie[Mierda]
#    print 'la somme tot des H1 :',SumH1 
#    print 'lentropy tot de sequence en cours :',y
#    multiInfo=SumH1-y    
#    plt.plot(len(x), multiInfo, 'ro')
#    if multiInfo>maxordonnee:
#       maxordonnee=multiInfo
#    if multiInfo<minordonnee:
#       minordonnee=multiInfo
#plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(bits)')
#plt.title('MultiInfo = sumH(1)-Htot')
#plt.grid(True)



   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   for x,y in entropy_sum_order.items():
      dualdiff=entropy_sum_order.get(Nb_var-x+1,0)-y
      plt.plot(x, dualdiff, 'ro')
      if dualdiff>maxordonnee:
          maxordonnee=dualdiff
      if dualdiff<minordonnee:
          minordonnee=dualdiff
   plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(bits)')
   plt.title('Order Entropy Duality: H(n-k)-H(k)')
   plt.grid(True)
#   if display_figure == True: 
#      plt.show(num_fig)
#   else: plt.close(num_fig)        
   if save_results == True: 
       save_FIGURE(num_fig, 'Order Entropy Duality: H(n-k)-H(k)', Format_SVG_SAVE_Figure )


# AVERAGE INFORMATION EFFICIENCY 
   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   for x,y in entropy_sum_order.items():
# Hmax is = nlog(nb_bin) ...     
      redundancy=1-(y/(binomial(Nb_var,x)*x*math.log(Nb_bins)/math.log(2) ) )
      plt.plot(x,redundancy, 'ro')
      if redundancy>maxordonnee:
          maxordonnee=redundancy
      if redundancy<minordonnee:
          minordonnee=redundancy
   plt.axis([0, Nb_var+1,minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('[0,1]')
   plt.title('Average Efficiency - Redundancy')
   plt.grid(True)
#   if display_figure == True: 
#      plt.show(num_fig)
 #  else: plt.close(num_fig)        
   if save_results == True: 
       save_FIGURE(num_fig, 'Average Efficiency - Redundancy', Format_SVG_SAVE_Figure )

# INFORMATION EFFICIENCY 
#plt.subplot(323)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Nentropie.items():
# Hmax is = nlog(nb_bin) ...     
#   redundancy=1-(y/(len(x)*math.log(Nb_bins)/math.log(2) ) )
#   plt.plot(len(x),redundancy, 'ro')
#   if redundancy>maxordonnee:
#       maxordonnee=redundancy
#   if redundancy<minordonnee:
#       minordonnee=redundancy
#plt.axis([0, Nb_var+1,minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('[0,1]')
#plt.title('Efficiency - Redundancy')
#plt.grid(True)




# FIGURE 2 MUTUAL INFORMATIONS




# MUTUAL INFORMATIONS

if SHOW_results_INFOMUT == True: 
   num_fig=num_fig+1
   plt.figure(num_fig)
   if SHOW_HARD_DISPLAY_medium== True:
      import matplotlib as mpl
      mpl.rcParams['agg.path.chunksize'] = 3000000  
      maxordonnee=-1000000.00
      minordonnee=1000000.00
      x_absss = np.array([])
      y_absss = np.array([]) 
      for x,y in Ninfomut.items():
         if len(x)>0:
#             print(len(x))
#             print(y)
             x_absss=np.append(x_absss,[len(x)])
             y_absss=np.append(y_absss,[y])
#             plt.plot(len(x), y, 'ro',len(x), y, 'b-',len(x),0,'r-')
             if y>maxordonnee:
                maxordonnee=y
             if y<minordonnee:
                minordonnee=y
#      plt.plot(x_absss, y_absss, 'ro',len(x), y, 'b-',len(x),0,'r-')     
#      plt.plot(x_absss, y_absss, 'ro',alpha=(-1.8*x_absss)/(Nb_var-2)+(Nb_var-0.2)/(Nb_var-2))          
      plt.scatter(x_absss, y_absss, marker= "_", s= 1000)
#      print(x_absss)
#      print(y_absss)
      plt.draw()       
      plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
      plt.ylabel('(Bits)')
      plt.title('Mutual Information')
      plt.grid(True)
#      if display_figure == True: 
#        plt.show(num_fig)        
#      else: plt.close(num_fig)        
      if save_results == True: 
         save_FIGURE(num_fig, 'DIAGRAM_INFOMUT', Format_SVG_SAVE_Figure )

# SUM OF MUTUAL INFORMATIONS Per degrees

   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   for x,y in infomut_sum_order.items():
       if x>1:
           plt.plot(x, y, 'ro',x, y, 'b-',x, 0, 'r-')
           if y>maxordonnee:
              maxordonnee=y
           if y<minordonnee:
              minordonnee=y
   plt.axis([1, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(Bits)')
   plt.title('Sum Mutual Information per Order')
   plt.grid(True)
#   print('tamerenslip-2')
#   if display_figure == True: 
#        plt.show(num_fig)
#        print('tamerenslip-1')
#   else: plt.close(num_fig)        
   if save_results == True: 
         save_FIGURE(num_fig, 'Sum Mutual Information per Order', Format_SVG_SAVE_Figure )

# Average MUTUAL INFORMATIONS RATE Per degrees 


   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   for x,y in infomut_sum_order.items():
       rate=y/x
       if x>0:
           plt.plot(x, rate, 'ro',x, y, 'b-',x, 0, 'r-')
           if rate>maxordonnee:
              maxordonnee=rate
           if rate<minordonnee:
              minordonnee=rate
   plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(Bits)')
   plt.title('Averaged Mutual Information rate')
   plt.grid(True)
   print('tamerenslip0')
#   if display_figure == True: 
#        plt.show(num_fig)
 #       print('tamerenslip1')
 #  else: plt.close(num_fig)        
   if save_results == True: 
         save_FIGURE(num_fig, 'Averaged Mutual Information rate', Format_SVG_SAVE_Figure )


# Average MUTUAL INFORMATIONS NORMALISED Per degrees 
   print('Robertwasthere')  
   num_fig=num_fig+1
   plt.figure(num_fig)
   maxordonnee=-1000000.00
   minordonnee=1000000.00
   normed_informut={} 
   for x,y in infomut_sum_order.items():
       rate=y/binomial(Nb_var,x)
       normed_informut[x]=y/binomial(Nb_var,x)
       if x>0:
           plt.plot(x, rate, 'ro',x, y, 'b-',x, 0, 'r-')
           if rate>maxordonnee:
              maxordonnee=rate
           if rate<minordonnee:
              minordonnee=rate
   plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
   plt.ylabel('(Bits)')
   plt.title('Averaged Mutual Information Normalised by Binomial coef')
   plt.grid(True)   
   plt.show(num_fig)
   print('Robertwasthere1')
#   if display_figure == True: 
#        plt.show(num_fig)
#        print('tamerenslip')
#   else: plt.close(num_fig)        
   if save_results == True: 
         save_FIGURE(num_fig, 'Averaged Mutual Information Normalised by Binomial coef', Format_SVG_SAVE_Figure )
         save_obj(normed_informut,'normed_informut'+str(Nb_bins))

# Information distance-pseudovol d= H-I 

   if SHOW_HARD_DISPLAY== True:
      num_fig=num_fig+1
      plt.figure(num_fig)   
      maxordonnee=-1000000.00
      minordonnee=1000000.00
      for x,y in Ninfomut.items():
          dist=Nentropie[x]-Ninfomut[x]
          if len(x)>0:
             plt.plot(len(x), dist, 'ro',len(x), dist, 'b-',len(x),0,'r-')
             if dist>maxordonnee:
                maxordonnee=dist
             if dist<minordonnee:
                minordonnee=dist
      plt.axis([0, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
      plt.ylabel('(Bits)')
      plt.title('Info 2-Distance - n-Volume')
      plt.grid(True)
#      if display_figure == True: 
#           plt.show(num_fig)
#      else: plt.close(num_fig)        
      if save_results == True: 
            save_FIGURE(num_fig, 'Info 2-Distance - n-Volume', Format_SVG_SAVE_Figure )
         
#plt.show()
 
 
 
#plt.subplot(323)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Ninfomut.items():
#   if len(x)>1:
#       percent=100*math.fabs(y)/infomut_sum_order_abs[len(x)]
#       plt.plot(len(x), percent, 'ro')
#       if percent>maxordonnee:
#          maxordonnee=percent
#       if percent<minordonnee:
#          minordonnee=percent
#plt.axis([1, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(%)')
#plt.title('InfoMut(n)/SumInfomutOrder(n)')
#plt.grid(True)

# Calcul de Ninfomut_contrib:
# % de participation de |In| par rapport a la somme de 
# toutes les |Ik| (k=<n) avec les mêmes variables internes

#Ninfomut_contrib={}
#for x,y in Ninfomut.items():
#   Denominator=0.00 
#   for w,z in Ninfomut.items():
#         test_in=0         
#         for v in range(0,len(w)):
#             if  w[v] in x: 
#                test_in= test_in+0
#             else: 
#                test_in= test_in+1 
#         if test_in!= 0 :
#             Denominator=Denominator
#            Ninfomut[x]=Ninfomut.get(x,0) 
# I suppose the last lign was not very useful     pass?      
#         else:
#             Denominator=Denominator+math.fabs(Ninfomut[w])
 #            Ninfomut[x]=Ninfomut.get(x,0)+ ((-1)**(len(w)+1))*Nentropie[w]
#   if Denominator!=0.00:
#       Ninfomut_contrib[x]=math.fabs(Ninfomut[x])/Denominator 
   

#plt.subplot(324)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Ninfomut_contrib.items():
#   if len(x)>1:
#      plt.plot(len(x), y, 'ro')
#      if y>maxordonnee:
#         maxordonnee=y
#      if y<minordonnee:
#         minordonnee=y
#plt.axis([1, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(%)')
#plt.title('abs(InfoMut(n))/Sum_abs(InfomutOrder)(1->n)')
#plt.grid(True)


#plt.subplot(325)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Ninfomut.items():
#   if len(x)>1:  
#      percent=100*math.fabs(y)/infotot_absbis
#      plt.plot(len(x), percent, 'ro')
#      if percent>maxordonnee:
#         maxordonnee=percent
#      if percent<minordonnee:
#         minordonnee=percent
#plt.axis([1, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(%)')
#plt.title('InfoMut(n)/SumInfomutTOT')
#plt.grid(True)

#plt.subplot(326)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Nentropie.items():
#    if len(x)>1:
#        SumH1=0.00
#        print 'le code de sequence encours est:', x
#        for v in range(0,len(x)):
#            Mierda=(x[v],)
#            print 'les composantes marg :', x[v]
 #           print 'lentropy de la marg :', Nentropie[Mierda]
#            SumH1=SumH1+Nentropie[Mierda]
#        print 'la somme tot des H1 :',SumH1 
#        print 'lentropy tot de sequence en cours :',y
#        multiInfo=((SumH1-y)**0.5)   
#        plt.plot(len(x), multiInfo, 'ro')
#        if multiInfo>maxordonnee:
#           maxordonnee=multiInfo
#        if multiInfo<minordonnee:
#           minordonnee=multiInfo
#plt.axis([1, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(sqrt(HoRh))')
#plt.title('MultiInfo = sqrt(sumHi-Htot)')
#plt.grid(True)


# FIGURE 3 ESSAIS

#plt.figure(7)

#plt.subplot(321)
#maxordonnee=-1000000.00
#minordonnee=1000000.00
#for x,y in Nentropie.items():
#    if len(x)>1:
#        SumH1=0.00
#        print 'le code de sequence encours est:', x
#        for v in range(0,len(x)):
#            Mierda=(x[v],)
#            print 'les composantes marg :', x[v]
 #           print 'lentropy de la marg :', Nentropie[Mierda]
#            SumH1=SumH1+Nentropie[Mierda]
#        print 'la somme tot des H1 :',SumH1 
#        print 'lentropy tot de sequence en cours :',y
#        multiInfo=(math.log(SumH1-y))   
#        plt.plot(len(x), multiInfo, 'ro')
#        if multiInfo>maxordonnee:
#           maxordonnee=multiInfo
#        if multiInfo<minordonnee:
#           minordonnee=multiInfo
#plt.axis([1, Nb_var+1, minordonnee-((maxordonnee-minordonnee)*0.07),maxordonnee+((maxordonnee-minordonnee)*0.07)])
#plt.ylabel('(log(HoRh)/n)')
#plt.title('MultiInfo = log(sumHi-Htot)/n')
#plt.grid(True)



# FIGURE 3.5 CONDITIONAL INFORMATION
# Nentropy and Ninfomut are dictionaries (x,y) with x a list of kind (1,2,5) and y a value in bit 


if SHOW_results_COND_INFOMUT == True: 
   num_fig=num_fig+1
   plt.figure(num_fig)
   moyenne={}
   nbpoint={}
   matrix_distrib_infomut=np.array([])
   x_absss = np.array([])
   y_absss = np.array([]) 
# Display every Histo with its own scales:   
#   minim={}
#  maxim={}
   ListInfomutcond={}
   maxima_tot=-1000000.00
   minima_tot=1000000.00
   Ninfomut_COND=[]
   Ninfomut_CONDtot=[]
   Ninfomut_COND_perorder=[]
   for i in range(1,Nb_var+1):
      ListInfomutcond[i]=[]
      Ninfomut_CONDperORDER=[]
      Ninfomut_COND.append(Ninfomut_CONDperORDER)
      dicobis={} 
      Ninfomut_CONDtot.append(dicobis)
      for j in range(1,Nb_var+1):
          dico={} 
          Ninfomut_COND[i-1].append(dico)

   for i in range(1,Nb_var+1):
      for x,y in Ninfomut.items():
         if len(x)>1: 
            for b in x:
                 if (b==i):
                    xbis= tuple(a for a in x if (a!=i)) 
#                    print('x:',x,'xbis:', xbis)
                    cond= Ninfomut[xbis]-y
                    if cond>maxima_tot:
                       maxima_tot=cond
                    if cond<minima_tot:
                       minima_tot=cond 
# for conditioning per degree                       
                    ListInfomutcond[len(x)-1].append(cond)                  
# for conditioning per variable                    
#                    ListInfomutcond[i].append(cond) 
                    Ninfomut_COND[len(x)-1][i-1][xbis]=cond
                    xter = xbis + ((i),)
                    Ninfomut_CONDtot[len(x)-1][xter]=cond
# The last term in the tuple is the conditionning variable                    
#      moyenne[len(x)]=moyenne.get(len(x),0)+y
#      nbpoint[len(x)]=nbpoint.get(len(x),0)+1
         if SHOW_HARD_DISPLAY_medium== True:
             if len(x)>0:
#             print(len(x))
#             print(y)
                    x_absss=np.append(x_absss,[len(x)-0.5])
                    y_absss=np.append(y_absss,[y])
   

 

   for a in range(1,Nb_var+1):
       if Nb_var<9 :
           plt.subplot(3,3,a)
       else :    
           if Nb_var<=16 :
               plt.subplot(4,4,a)
           else : 
              if Nb_var<=20 : 
                   plt.subplot(5,4,a) 
              else :
                  plt.subplot(5,5,a)
       ListInfomutcond[a].append(minima_tot-0.1)
       ListInfomutcond[a].append(maxima_tot+0.1)           
#       Numgraph=
       resultion_histo=100  
       n, bins, patches = plt.hist(ListInfomutcond[a], resultion_histo, facecolor='r')
       plt.axis([minima_tot, maxima_tot,0,n.max()])
#      print(n)
       if a==1 :
          matrix_distrib_infomut=n
       else: 
          matrix_distrib_infomut=np.c_[matrix_distrib_infomut,n]
#          print(matrix_distrib_infomut)
#       matrix_distrib_infomut=np.vstack((matrix_distrib_infomut,n))
#      plt.axis([minim[a]-((maxim[a]-minim[a])*0.07), maxim[a]+((maxim[a]-minim[a])*0.07), 0,nbpoint[a]+10000])
#       plt.ylabel('(count)')
#       plt.title('Histogram of Mutual info at order '+str(a))
#       plt.text(moyenne[a]/nbpoint[a], 1,moyenne[a]/nbpoint[a])
       plt.grid(True)
#   if display_figure == True: 
#   print(matrix_distrib_infomut)   
    
#   mng = plt.get_current_fig_manager(num_fig)                                         
#   mng.window.showMaximized(num_fig)
#   plt.show(num_fig)  
   print('onestdes')    
#   plt.tight_layout(num_fig)    
   if save_results == True:    
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="svg")  
   
   
   num_fig=num_fig+1
   plt.figure(num_fig)  
   matrix_distrib_infomut=np.flipud(matrix_distrib_infomut)
   multiply_by_info_values= False
   if multiply_by_info_values :
        for i in range(0,resultion_histo):    
           matrix_distrib_infomut[i,:]=matrix_distrib_infomut[i,:] * ((maxima_tot-minima_tot+0.2) *(resultion_histo-i)/(resultion_histo-1) +minima_tot-0.1)
           print((maxima_tot-minima_tot+0.2) *(resultion_histo-i) /(resultion_histo-1) +minima_tot-0.1)
           print('i est ')
           print((resultion_histo-i))
           print(matrix_distrib_infomut[i,:])
#   print('min')
#   print(minima_tot-0.1)
#   print(maxima_tot+0.1)
   
#   plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=Nb_var/(2*resultion_histo),extent=[1,Nb_var,minima_tot-0.1,maxima_tot+0.1])
#   norm = matplotlib.colors.Normalize(vmin = 1, vmax = np.max(matrix_distrib_infomut))
   if multiply_by_info_values :
        if ((-1)*matrix_distrib_infomut.min()) <= (matrix_distrib_infomut.max()): 
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=-matrix_distrib_infomut.max(), vmax=matrix_distrib_infomut.max())
        else :
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=matrix_distrib_infomut.min(), vmax=-matrix_distrib_infomut.min())
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=norm)
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   else :
#        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm(vmin=1, vmax=200000))
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm())
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   plt.colorbar()
   plt.grid(False)     
   if SHOW_HARD_DISPLAY_medium== True:        
       plt.scatter(x_absss, y_absss, marker= "o" , facecolor= 'w')
   for i in range(1,6):
       print('ORDER =',(i-1))
       Ninfomut_CONDtot[i-1]=OrderedDict(sorted(Ninfomut_CONDtot[i-1].items(), key=lambda t: t[1]))
       print(Ninfomut_CONDtot[i-1])
#       for j  in range(1,Nb_var+1):
#              print('COND PER =',(j))
#             Ninfomut_COND[i-1][j-1]=OrderedDict(sorted(Ninfomut_COND[i-1][j-1].items(), key=lambda t: t[1]))
#              print(Ninfomut_COND[i-1][j-1])
#   plt.show(num_fig)       
#   print('tamere12')
#else: plt.close(num_fig)    
     





# FIGURE 4 Histogramms ENTROPY



if SHOW_results_ENTROPY_HISTO == True:
   num_fig=num_fig+1
   plt.figure(num_fig)  
   moyenne={}
   nbpoint={}
   matrix_distrib_info=np.array([])
   x_absss = np.array([])
   y_absss = np.array([]) 
# Display every Histo with its own scales:   
#   minim={}
#   maxim={}
   
   maxima_tot=-1000000.00
   minima_tot=1000000.00   
   ListEntropyordre={}

   for i in range(1,Nb_var+1):
     ListEntropyordre[i]=[]

   for x,y in Nentropie.items():
      ListEntropyordre[len(x)].append(y)
      moyenne[len(x)]=moyenne.get(len(x),0)+y
      nbpoint[len(x)]=nbpoint.get(len(x),0)+1
      if SHOW_HARD_DISPLAY_medium== True:
          if len(x)>0:
#             print(len(x))
#             print(y)
                 x_absss=np.append(x_absss,[len(x)-0.5])
                 y_absss=np.append(y_absss,[y])
# Display every Histo with its own scales:          
#      minim[len(x)]=minim.get(len(x),1000000.00)
#      maxim[len(x)]=maxim.get(len(x),-1000000.00)
# Display every Histo with its own scales:      
#      if y>maxim[len(x)]:
#          maxim[len(x)]=y
#     if y<minim[len(x)]:
#          minim[len(x)]=y    
      if y>maxima_tot:
          maxima_tot=y
      if y<minima_tot:
          minima_tot=y     
   for a in range(1,Nb_var+1):
       if Nb_var<=9 :
#       Numgraph=330+a
           plt.subplot(3,3,a)
       else :    
           if Nb_var<=16 :
               plt.subplot(4,4,a)
           else : 
              if Nb_var<=20 : 
                   plt.subplot(5,4,a) 
              else :
                  plt.subplot(5,5,a)
       ListEntropyordre[a].append(minima_tot-0.1)
       ListEntropyordre[a].append(maxima_tot+0.1)           
#    print 'listentropie à lordre:',a
#    print ListEntropyordre[a]
       resultion_histo=100  
       n, bins, patches = plt.hist(ListEntropyordre[a], resultion_histo, facecolor='g')
       plt.axis([minima_tot, maxima_tot,0,n.max()])
       if a==1 :
          matrix_distrib_info=n
       else: 
          matrix_distrib_info=np.c_[matrix_distrib_info,n]
#    plt.axis([minim[a]-((maxim[a]-minim[a])*0.07), maxim[a]+((maxim[a]-minim[a])*0.07), 0,nbpoint[a]+10000])
#       plt.ylabel('(count)')
#       plt.title('Histogram of Entropy at order '+str(a))
#       plt.text(moyenne[a]/nbpoint[a], 1,moyenne[a]/nbpoint[a])
       plt.grid(True)
    
#   mng = plt.get_current_fig_manager(num_fig)                                         
#   mng.window.showMaximized(num_fig)
#   plt.tight_layout(num_fig) 
   if save_results == True: 
      plt.savefig(os.path.join(directory_path,'Histogram_ENTROPY'))    
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'Histogram_ENTROPY'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'Histogram_ENTROPY'), format="svg")    
         
   num_fig=num_fig+1
   plt.figure(num_fig)  
   matrix_distrib_info=np.flipud(matrix_distrib_info)
   multiply_by_info_values= False
   if multiply_by_info_values :
        for i in range(0,resultion_histo):    
           matrix_distrib_info[i,:]=matrix_distrib_info[i,:] * ((maxima_tot-minima_tot+0.2) *(resultion_histo-i)/(resultion_histo-1) +minima_tot-0.1)
           print((maxima_tot-minima_tot+0.2) *(resultion_histo-i) /(resultion_histo-1) +minima_tot-0.1)
           print('i est ')
           print((resultion_histo-i))
           print(matrix_distrib_info[i,:])
#   print('min')
#   print(minima_tot-0.1)
#   print(maxima_tot+0.1)
   
#   plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=Nb_var/(2*resultion_histo),extent=[1,Nb_var,minima_tot-0.1,maxima_tot+0.1])
#   norm = matplotlib.colors.Normalize(vmin = 1, vmax = np.max(matrix_distrib_infomut))
   if multiply_by_info_values :
        if ((-1)*matrix_distrib_info.min()) <= (matrix_distrib_info.max()): 
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=-matrix_distrib_info.max(), vmax=matrix_distrib_info.max())
        else :
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=matrix_distrib_info.min(), vmax=-matrix_distrib_info.min())
        plt.matshow(matrix_distrib_info, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=norm)
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   else :
#        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm(vmin=1, vmax=200000))
        plt.matshow(matrix_distrib_info, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm())
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   plt.colorbar()
   plt.grid(False)     
   if SHOW_HARD_DISPLAY_medium== True:        
       plt.scatter(x_absss, y_absss, marker= "o" , facecolor= 'w')

#   plt.show(num_fig)       
#   print('tamere12')
#else: plt.close(num_fig)         

# FIGURE  Histogramms MUTUAL INFORMATION



if SHOW_results_INFOMUT_HISTO == True:
    
   num_fig=num_fig+1
   plt.figure(num_fig)   
   moyenne={}
   nbpoint={}
   matrix_distrib_infomut=np.array([])
   x_absss = np.array([])
   y_absss = np.array([]) 
# Display every Histo with its own scales:   
#   minim={}
#  maxim={}
   ListInfomutordre={}
   maxima_tot=-1000000.00
   minima_tot=1000000.00
   for i in range(1,Nb_var+1):
     ListInfomutordre[i]=[]

   for x,y in Ninfomut.items():
      ListInfomutordre[len(x)].append(y)
      moyenne[len(x)]=moyenne.get(len(x),0)+y
      nbpoint[len(x)]=nbpoint.get(len(x),0)+1
      if SHOW_HARD_DISPLAY_medium== True:
          if len(x)>0:
#             print(len(x))
#             print(y)
                 x_absss=np.append(x_absss,[len(x)-0.5])
                 y_absss=np.append(y_absss,[y])
# Display every Histo with its own scales:
      
#      minim[len(x)]=minim.get(len(x),1000000.00)
 #     maxim[len(x)]=maxim.get(len(x),-1000000.00)
# Display every Histo with its own scales:

#      if y>maxim[len(x)]:
#          maxim[len(x)]=y
 #     if y<minim[len(x)]:
 #         minim[len(x)]=y  
      if y>maxima_tot:
          maxima_tot=y
      if y<minima_tot:
          minima_tot=y 
       

 

   for a in range(1,Nb_var+1):
       if Nb_var<9 :
           plt.subplot(3,3,a)
       else :    
           if Nb_var<=16 :
               plt.subplot(4,4,a)
           else : 
              if Nb_var<=20 : 
                   plt.subplot(5,4,a) 
              else :
                  plt.subplot(5,5,a)
       ListInfomutordre[a].append(minima_tot-0.1)
       ListInfomutordre[a].append(maxima_tot+0.1)           
#       Numgraph=
       resultion_histo=100  
       n, bins, patches = plt.hist(ListInfomutordre[a], resultion_histo, facecolor='r')
       plt.axis([minima_tot, maxima_tot,0,n.max()])
#      print(n)
       if a==1 :
          matrix_distrib_infomut=n
       else: 
          matrix_distrib_infomut=np.c_[matrix_distrib_infomut,n]
#          print(matrix_distrib_infomut)
#       matrix_distrib_infomut=np.vstack((matrix_distrib_infomut,n))
#      plt.axis([minim[a]-((maxim[a]-minim[a])*0.07), maxim[a]+((maxim[a]-minim[a])*0.07), 0,nbpoint[a]+10000])
#       plt.ylabel('(count)')
#       plt.title('Histogram of Mutual info at order '+str(a))
#       plt.text(moyenne[a]/nbpoint[a], 1,moyenne[a]/nbpoint[a])
       plt.grid(True)
#   if display_figure == True: 
#   print(matrix_distrib_infomut)   
    
#   mng = plt.get_current_fig_manager(num_fig)                                         
#   mng.window.showMaximized(num_fig)
#   plt.show(num_fig)  
   print('onestdes')    
#   plt.tight_layout(num_fig)    
   if save_results == True:    
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="svg")  
   
   
   num_fig=num_fig+1
   plt.figure(num_fig)  
   matrix_distrib_infomut=np.flipud(matrix_distrib_infomut)
   multiply_by_info_values= False
   if multiply_by_info_values :
        for i in range(0,resultion_histo):    
           matrix_distrib_infomut[i,:]=matrix_distrib_infomut[i,:] * ((maxima_tot-minima_tot+0.2) *(resultion_histo-i)/(resultion_histo-1) +minima_tot-0.1)
           print((maxima_tot-minima_tot+0.2) *(resultion_histo-i) /(resultion_histo-1) +minima_tot-0.1)
           print('i est ')
           print((resultion_histo-i))
           print(matrix_distrib_infomut[i,:])
#   print('min')
#   print(minima_tot-0.1)
#   print(maxima_tot+0.1)
   
#   plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=Nb_var/(2*resultion_histo),extent=[1,Nb_var,minima_tot-0.1,maxima_tot+0.1])
#   norm = matplotlib.colors.Normalize(vmin = 1, vmax = np.max(matrix_distrib_infomut))
   if multiply_by_info_values :
        if ((-1)*matrix_distrib_infomut.min()) <= (matrix_distrib_infomut.max()): 
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=-matrix_distrib_infomut.max(), vmax=matrix_distrib_infomut.max())
        else :
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=matrix_distrib_infomut.min(), vmax=-matrix_distrib_infomut.min())
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=norm)
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   else :
#        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm(vmin=1, vmax=200000))
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm())
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   plt.colorbar()
   plt.grid(False)     
   if SHOW_HARD_DISPLAY_medium== True:        
       plt.scatter(x_absss, y_absss, marker= "o" , facecolor= 'w')

#   plt.show(num_fig)       
#   print('tamere12')
#else: plt.close(num_fig)         


###############################################################
########          FIGURE INFORMATION PATH            ##########
###                                                   #########
######                                                #########
###############################################################    

# Nentropy and Ninfomut are dictionaries (x,y) with x a list of kind (1,2,5) and y a value in bit 


if SHOW_results_INFOMUT_path == True: 
   num_fig=num_fig+1
   plt.figure(num_fig)
   moyenne={}
   nbpoint={}
   matrix_distrib_infomut=np.array([])
   x_absss = np.array([])
   y_absss = np.array([]) 
# Display every Histo with its own scales:   
#   minim={}
#  maxim={}
   ListInfomutcond={}
   maxima_tot=-1000000.00
   minima_tot=1000000.00
   Ninfomut_COND=[]
   Ninfomut_CONDtot=[]
   Ninfomut_COND_perorder=[]
   
   infocond=1000000.00
   listartbis=[] 
   infomutmax_path_VAR=[] 
   infomutmax_path_VALUE=[]
   infomutmin_path_VAR=[] 
   infomutmin_path_VALUE=[]
   number_of_max_and_min=21 #explore the 2 max an min marginals (of degree 1 information) 
   items = list(Ninfomut_per_order_ordered[1].items())
   for inforank in range(0,number_of_max_and_min): 
       infomutmax_path_VAR.append([])
       infomutmax_path_VALUE.append([])
       infomutmin_path_VAR.append([])
       infomutmin_path_VALUE.append([])
       liststart=[]
#       print(items[inforank])
#       print(items[Nb_var-inforank-1])
       xstart=items[inforank][0]
       xstartmin=items[Nb_var-inforank-1][0]
#       print(xstart)
#       print(xstart[0])
       infomutmax_path_VAR[-1].append(xstart[0])
       infomutmax_path_VALUE[-1].append(items[inforank][1])
       infomutmin_path_VAR[-1].append(xstartmin[0])
       infomutmin_path_VALUE[-1].append(items[Nb_var-inforank-1][1])
       degree=1
       infocond=1000000.00
       while infocond >=0 :
          maxima_tot=-1000000.00
          minima_tot=1000000.00 
          degree=degree+1 
          for i in range(1,Nb_var+1) :
             if i in infomutmax_path_VAR[-1] :     
                del listartbis[:]  
             else:    
                del listartbis[:]
                listartbis=infomutmax_path_VAR[-1][:]
                listartbis.append(i)
                listartbis.sort() 
#                print('listartbis')
#                print(listartbis)
                tuplestart=tuple(listartbis)
                if infomut_per_order[degree][tuplestart]>maxima_tot:
                      maxima_tot=infomut_per_order[degree][tuplestart]
                      tuplemax=tuplestart
                      igood= i
#          print(tuplemax)
#          print(maxima_tot)
          infomutmax_path_VAR[-1].append(igood)
          infomutmax_path_VALUE[-1].append(maxima_tot)
          infocond= infomutmax_path_VALUE[-1][-2]- infomutmax_path_VALUE[-1][-1]
#          print('infocond')
#          print(infocond)
       del infomutmax_path_VAR[-1][-1]
       del infomutmax_path_VALUE[-1][-1]
       print('The path of maximal mutual-info Nb',inforank+1,' is :')   
       print(infomutmax_path_VAR[-1])   
       
       degree=1
       infocond=1000000.00
       while infocond >=0 :
          maxima_tot=-1000000.00
          minima_tot=1000000.00 
          degree=degree+1 
          for i in range(1,Nb_var+1) :
             if i in infomutmin_path_VAR[-1] :     
                del listartbis[:]  
             else:    
                del listartbis[:]
                listartbis=infomutmin_path_VAR[-1][:]
                listartbis.append(i)
                listartbis.sort() 
#                print('listartbis')
#                print(listartbis)
                tuplestart=tuple(listartbis)
                if infomut_per_order[degree][tuplestart]<minima_tot:
                      minima_tot=infomut_per_order[degree][tuplestart]
                      tuplemax=tuplestart
                      igood= i
#          print(tuplemax)
#          print(minima_tot)
          infomutmin_path_VAR[-1].append(igood)
          infomutmin_path_VALUE[-1].append(minima_tot)
          infocond= infomutmin_path_VALUE[-1][-2]- infomutmin_path_VALUE[-1][-1]
#          print('infocond')
#          print(infocond)
       del infomutmin_path_VAR[-1][-1]
       del infomutmin_path_VALUE[-1][-1]   
       print('The path of minimal mutual-info Nb',inforank+1,' is :')   
       print(infomutmin_path_VAR[-1])    

# COMPUTE THE HISTOGRAMS OF INFORMATION           
# Display every Histo with its own scales:   
#  minim={}
#  maxim={}
   ListInfomutordre={}
   maxima_tot=-1000000.00
   minima_tot=1000000.00
   for i in range(1,Nb_var+1):
     ListInfomutordre[i]=[]

   for x,y in Ninfomut.items():
      ListInfomutordre[len(x)].append(y)
      moyenne[len(x)]=moyenne.get(len(x),0)+y
      nbpoint[len(x)]=nbpoint.get(len(x),0)+1
      if SHOW_HARD_DISPLAY_medium== True:
          if len(x)>0:
#             print(len(x))
#             print(y)
                 x_absss=np.append(x_absss,[len(x)-0.5])
                 y_absss=np.append(y_absss,[y])
# Display every Histo with its own scales:
      
      if y>maxima_tot:
          maxima_tot=y
      if y<minima_tot:
          minima_tot=y 

   for a in range(1,Nb_var+1):
       if Nb_var<9 :
           plt.subplot(3,3,a)
       else :    
           if Nb_var<=16 :
               plt.subplot(4,4,a)
           else : 
              if Nb_var<=20 : 
                   plt.subplot(5,4,a) 
              else :
                  plt.subplot(5,5,a)
       ListInfomutordre[a].append(minima_tot-0.1)
       ListInfomutordre[a].append(maxima_tot+0.1)           
#       Numgraph=
       resultion_histo=100  
       n, bins, patches = plt.hist(ListInfomutordre[a], resultion_histo, facecolor='r')
       plt.axis([minima_tot, maxima_tot,0,n.max()])
#      print(n)
       if a==1 :
          matrix_distrib_infomut=n
       else: 
          matrix_distrib_infomut=np.c_[matrix_distrib_infomut,n]
#          print(matrix_distrib_infomut)
#       matrix_distrib_infomut=np.vstack((matrix_distrib_infomut,n))
#      plt.axis([minim[a]-((maxim[a]-minim[a])*0.07), maxim[a]+((maxim[a]-minim[a])*0.07), 0,nbpoint[a]+10000])
#       plt.ylabel('(count)')
#       plt.title('Histogram of Mutual info at order '+str(a))
#       plt.text(moyenne[a]/nbpoint[a], 1,moyenne[a]/nbpoint[a])
       plt.grid(True)
#   if display_figure == True: 
#   print(matrix_distrib_infomut)   
    
#   mng = plt.get_current_fig_manager(num_fig)                                         
#   mng.window.showMaximized(num_fig)
#   plt.show(num_fig)  
   print('onestdes')    
#   plt.tight_layout(num_fig)    
   if save_results == True:    
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'Histogram_INFOMUT'), format="svg")  

# COMPUTE THE MATRIX OF INFORMATION LANDSACPES   
   
   num_fig=num_fig+1
#   plt.figure(num_fig)  
   plt.figure(num_fig,figsize=(50, 30))   
   matrix_distrib_infomut=np.flipud(matrix_distrib_infomut)
   multiply_by_info_values= False
   if multiply_by_info_values :
        for i in range(0,resultion_histo):    
           matrix_distrib_infomut[i,:]=matrix_distrib_infomut[i,:] * ((maxima_tot-minima_tot+0.2) *(resultion_histo-i)/(resultion_histo-1) +minima_tot-0.1)
           print((maxima_tot-minima_tot+0.2) *(resultion_histo-i) /(resultion_histo-1) +minima_tot-0.1)
           print('i est ')
           print((resultion_histo-i))
           print(matrix_distrib_infomut[i,:])

   if multiply_by_info_values :
        if ((-1)*matrix_distrib_infomut.min()) <= (matrix_distrib_infomut.max()): 
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=-matrix_distrib_infomut.max(), vmax=matrix_distrib_infomut.max())
        else :
            norm=mpl.colors.SymLogNorm(linthresh=1, vmin=matrix_distrib_infomut.min(), vmax=-matrix_distrib_infomut.min())
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=norm)
        plt.axis([0,Nb_var,minima_tot,maxima_tot])
   else :
# for figure paper with colar scale up to 200000       
        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm(vmin=1, vmax=200000))
# for autoscaled color scale - general case 
#        plt.matshow(matrix_distrib_infomut, cmap='jet', aspect=3, extent=[0,Nb_var,minima_tot-0.1,maxima_tot+0.1], norm=LogNorm())
        plt.axis([0,Nb_var,minima_tot,maxima_tot])  
#   plt.axes().set_aspect(0.5)     
   plt.colorbar()
   plt.grid(False)     
   if SHOW_HARD_DISPLAY_medium== True:        
       plt.scatter(x_absss, y_absss, marker= "o" , facecolor= 'w')
 

# COMPUTE THE INFORMATION PATHS    
   x_infomax=[] 
   x_infomin=[]    
   maxima_x=-10
   maxima_tot=-1000000.00
   minima_tot=1000000.00
   for inforank in range(0,number_of_max_and_min): 
       x_infomax.append([])
       j=-0.5
       for y in  range(0,len(infomutmax_path_VALUE[inforank])): 
          j=j+1 
          x_infomax[-1].append(j) 
          if j > maxima_x:
              maxima_x=j
          if infomutmax_path_VALUE[inforank][int(j-0.5)]>maxima_tot :
              maxima_tot=infomutmax_path_VALUE[inforank][int(j-0.5)]
          if infomutmax_path_VALUE[inforank][int(j-0.5)]<minima_tot :
              minima_tot=infomutmax_path_VALUE[inforank][int(j-0.5)]    
              
       x_infomin.append([])
       j=-0.5
       for y in  range(0,len(infomutmin_path_VALUE[inforank])): 
          j=j+1 
          x_infomin[-1].append(j) 
          if j > maxima_x:
              maxima_x=j
          if infomutmin_path_VALUE[inforank][int(j-0.5)]>maxima_tot :
              maxima_tot=infomutmin_path_VALUE[inforank][int(j-0.5)]
          if infomutmin_path_VALUE[inforank][int(j-0.5)]<minima_tot :
              minima_tot=infomutmin_path_VALUE[inforank][int(j-0.5)]     
              
       plt.plot(x_infomax[inforank], infomutmax_path_VALUE[inforank], marker='o', color='red')
       plt.plot(x_infomin[inforank], infomutmin_path_VALUE[inforank], marker='o',color='blue')
       plt.axis([0,maxima_x+0.5,minima_tot-0.2,maxima_tot+0.2])
      
#       plt.xlim(0, maxima_x+0.5)
   display_labelnodes=False
   if display_labelnodes :    
       for inforank in range(0,number_of_max_and_min):
           for label, x,y in zip(infomutmax_path_VAR[inforank], x_infomax[inforank], infomutmax_path_VALUE[inforank]):
                plt.annotate(label,xy=(x, y), xytext=(0, 0),textcoords='offset points')
           for label, x,y in zip(infomutmin_path_VAR[inforank], x_infomin[inforank], infomutmin_path_VALUE[inforank]):
                plt.annotate(label,xy=(x, y), xytext=(0, 0),textcoords='offset points')    
                
               
 
       




###############################################################
########          Ring representation               ##########
###                Mutual information               #########
###### http://networkx.readthedocs.org/en/stable/   #########
###############################################################    
#import networkx as nx

#plt.figure(9)


#if SHOW_results_SCAFOLD == True:

#   netring = nx.Graph()
#   list_of_node=[]
#   list_of_size=[]
#   list_of_edge=[]
#   list_of_width=[]


#   for x,y in infomut_per_order[1].items():
#        tuple_interim=(workbook_data[Name_worksheet].cell(row = x[0]+1, column = 1).value )
#        list_of_size.append((infomut_per_order[1][x]**2) *500)
#        list_of_node.append(tuple_interim)

#   tuple_interim=()
#   for x,y in infomut_per_order[2].items():    
#        tuple_interim=(workbook_data[Name_worksheet].cell(row = x[0]+1, column = 1).value,workbook_data[Name_worksheet].cell(row = x[1]+1, column = 1).value)
#        list_of_edge.append(tuple_interim)
#        netring.add_edge(tuple_interim[0], tuple_interim[1], weight= (infomut_per_order[2][x]) )
#        list_of_width.append((infomut_per_order[2][x]*10))
#     list_of_width.append(0.5*(infomut_per_order[2][x]*10)**2)

    
#   netring.add_nodes_from(list_of_node)
    

#   nx.draw_circular(netring, with_labels=True, nodelist = list_of_node,edgelist = list_of_edge, width= list_of_width, node_size = list_of_size)

#   mng = plt.get_current_fig_manager(9)                                         
#   mng.window.showMaximized(9)
#plt.tight_layout()     
#   if save_results == True: 
#      plt.savefig(os.path.join(directory_path,'SCAFOLD_INFOMUT_PAIR'), format="svg") 
#      if Format_SVG_SAVE_Figure == False: 
#         plt.savefig(os.path.join(directory_path,'SCAFOLD_INFOMUT_PAIR'), format="png")
#      if Format_SVG_SAVE_Figure == True:       
#         plt.savefig(os.path.join(directory_path,'SCAFOLD_INFOMUT_PAIR'), format="svg")  
#   if display_figure == True:
#      plt.show(9)
#   else: plt.close(9)  
#else: plt.close(9)    


import networkx as nx



if SHOW_results_SCAFOLD == True:
   num_fig=num_fig+1
   plt.figure(num_fig)  

   netring = nx.Graph()
   list_of_node=[]
   list_of_size=[]
   list_of_edge=[]
   list_of_width=[]
   list_of_labels={}
   mapping={}
   
   for x in range(1,Nb_var+1):
       tuple_interim=(workbook_data[Name_worksheet].cell(row = x, column = 1).value )
       list_of_node.append(x)
       list_of_labels[x]=tuple_interim
       cagoenlaputa=[x]
       mapping[x]=tuple_interim
       code_var=tuple(cagoenlaputa)
       list_of_size.append((infomut_per_order[1][code_var]**2) *500)
       netring.add_node(x)
       
#   for x,y in infomut_per_order[1].items():
#       tuple_interim=(workbook_data[Name_worksheet].cell(row = x[0]+1, column = 1).value )
#        list_of_size.append((infomut_per_order[1][x]**2) *500)
#        list_of_node.append(tuple_interim)

   tuple_interim=()
   for x,y in infomut_per_order[2].items():    
        tuple_interim=(workbook_data[Name_worksheet].cell(row = x[0]+1, column = 1).value,workbook_data[Name_worksheet].cell(row = x[1]+1, column = 1).value)
#        list_of_edge.append(tuple_interim)
        list_of_edge.append(x)
        var_1=x[0]
        var_2=x[1]
        netring.add_edge(var_1, var_2, weight= (infomut_per_order[2][x]) )
        list_of_width.append((infomut_per_order[2][x]*10))
#     list_of_width.append(0.5*(infomut_per_order[2][x]*10)**2)

   print(list_of_node) 
#   netring=nx.complete_graph(Nb_var+5)
#   netring.add_nodes_from(list_of_node)
  
  
#   G=nx.relabel_nodes(netring,mapping)
   nx.draw_circular(netring, with_labels= True, nodelist = list_of_node,edgelist = list_of_edge, width= list_of_width, node_size = list_of_size)
#   G=nx.relabel_nodes(G,mapping)

#   mng = plt.get_current_fig_manager(10)                                         
#   mng.window.showMaximized(10)
#plt.tight_layout()     
   if save_results == True: 
      plt.savefig(os.path.join(directory_path,'SCAFOLD_INFOMUT_PAIR'), format="svg") 
      if Format_SVG_SAVE_Figure == False: 
         plt.savefig(os.path.join(directory_path,'SCAFOLD_INFOMUT_PAIR'), format="png")
      if Format_SVG_SAVE_Figure == True:       
         plt.savefig(os.path.join(directory_path,'SCAFOLD_INFOMUT_PAIR'), format="svg")  
   if display_figure == True:
      plt.show()
   else: plt.close(num_fig)  
else: plt.close(num_fig)    


##############################
#infomut_per_order=[]
#Ninfomut_per_order_ordered=[]
#for x in range(Nb_var+1):
#    infomut_dicoperoder={} 
#    infomut_per_order.append(infomut_dicoperoder)
#    Ninfomut_per_order_ordered.append(infomut_dicoperoder)
    
#for x,y in Ninfomut.items():
#    print('lenghtx=',len(x))
#    infomut_per_order[len(x)][x]=Ninfomut[x]

#for x in range(Nb_var+1):
#    Ninfomut_per_order_ordered[x]=OrderedDict(sorted(infomut_per_order[x].items(), key=lambda t: t[1]))
#    print('les infos mutuelles à ', x ,'sont:')
#    print(Ninfomut_per_order_ordered[x])

#Ninfomut_ordered=OrderedDict(sorted(Ninfomut.items(), key=lambda t: t[1]))
    

print('boulot fini')